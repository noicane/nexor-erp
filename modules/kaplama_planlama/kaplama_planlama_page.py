# -*- coding: utf-8 -*-
"""
NEXOR ERP - Kaplama Hatti Haftalik Planlama
[KURUMSAL UI - v2.0]

Aciklama:
- Sol panel: Ozet kartlari, hat istatistik, banyo kartlari, urun listesi
- Sag panel: Toolbar, Gantt cizelgesi, Hat canli gorunum, Kapasite analizi
"""
from datetime import date, timedelta
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QSplitter, QLabel, QFrame,
    QPushButton, QScrollArea, QTableWidget, QTableWidgetItem,
    QHeaderView, QDateEdit, QDialog, QFormLayout, QLineEdit,
    QComboBox, QSpinBox, QMessageBox, QAbstractItemView, QGridLayout,
    QTabWidget, QSizePolicy, QFileDialog, QTimeEdit
)
from PySide6.QtCore import Qt, QTimer, QDate, QTime
from PySide6.QtGui import QFont, QColor

from components.base_page import BasePage
from .models import (
    KaplamaUrun, PlanGorev,
    BARA_SAYISI, VARDIYA_SURE_DK, VARDIYA_SAYISI, GUN_SAYISI
)
from .gantt_widget import GanttWidget
from .planlama_motoru import otomatik_planla
from .widgets import (
    BanyoCard, HatCanliWidget, BaraDurumStrip,
    HatIstatistikCard, ReceteAdimWidget, KapasiteGorselWidget,
    CevrimPlanWidget, optimize_cevrim_plan
)
from . import db_operations as db
from core.nexor_brand import brand


# ══════════════════════════════════════════════════════════════
#  SABITLER
# ══════════════════════════════════════════════════════════════

MARGIN = 20
SPACING = 16
CARD_SPACING = 12
CARD_RADIUS = 10
INPUT_RADIUS = 8
BUTTON_RADIUS = 8
TITLE_SIZE = 22
SUBTITLE_SIZE = 13
BODY_SIZE = 13
SMALL_SIZE = 11
LABEL_SIZE = 10


def _monday_of_week(d: date) -> date:
    return d - timedelta(days=d.weekday())


def _get_style(theme: dict) -> dict:
    return {
        'card_bg': theme.get('bg_card', '#151B23'),
        'input_bg': theme.get('bg_input', '#1A1A2E'),
        'border': theme.get('border', '#1E2736'),
        'text': theme.get('text', '#E8ECF1'),
        'text_secondary': theme.get('text_secondary', '#8896A6'),
        'text_muted': theme.get('text_muted', '#5C6878'),
        'primary': theme.get('primary', '#DC2626'),
        'success': theme.get('success', '#10B981'),
        'warning': theme.get('warning', '#F59E0B'),
        'error': theme.get('error', '#EF4444'),
        'info': theme.get('info', '#3B82F6'),
    }


def _table_style(s: dict) -> str:
    return f"""
        QTableWidget {{
            background: {s['card_bg']};
            color: {s['text']};
            border: 1px solid {s['border']};
            border-radius: {CARD_RADIUS}px;
            gridline-color: {s['border']};
            font-size: {BODY_SIZE}px;
        }}
        QTableWidget::item {{
            padding: 8px 10px;
            border-bottom: 1px solid {s['border']};
        }}
        QTableWidget::item:selected {{
            background: rgba(196, 30, 30, 0.15);
            color: {s['text']};
        }}
        QTableWidget::item:hover {{
            background: rgba(196, 30, 30, 0.06);
        }}
        QHeaderView::section {{
            background: #111822;
            color: {s['text_secondary']};
            padding: 10px 8px;
            border: none;
            border-bottom: 2px solid {s['primary']};
            font-weight: 600;
            font-size: {LABEL_SIZE + 1}px;
        }}
        QScrollBar:vertical {{
            background: transparent;
            width: 8px;
        }}
        QScrollBar::handle:vertical {{
            background: {s['border']};
            border-radius: 4px;
            min-height: 30px;
        }}
        QScrollBar::handle:vertical:hover {{
            background: #2A3545;
        }}
    """


def _input_style(s: dict) -> str:
    return f"""
        QLineEdit, QComboBox, QDateEdit, QSpinBox {{
            background: {s['input_bg']};
            color: {s['text']};
            border: 1px solid {s['border']};
            border-radius: {INPUT_RADIUS}px;
            padding: 8px 12px;
            font-size: {BODY_SIZE}px;
        }}
        QLineEdit:focus, QComboBox:focus, QSpinBox:focus {{
            border-color: {s['primary']};
        }}
    """


def _primary_btn_style(s: dict) -> str:
    return f"""
        QPushButton {{
            background: {s['primary']};
            color: white;
            border: none;
            border-radius: {BUTTON_RADIUS}px;
            padding: 8px 18px;
            font-weight: 600;
            font-size: {BODY_SIZE}px;
        }}
        QPushButton:hover {{ background: #D42A2A; }}
        QPushButton:pressed {{ background: #9B1818; }}
        QPushButton:disabled {{ background: {s['border']}; color: {s['text_muted']}; }}
    """


def _secondary_btn_style(s: dict) -> str:
    return f"""
        QPushButton {{
            background: transparent;
            color: {s['text']};
            border: 1px solid {s['border']};
            border-radius: {BUTTON_RADIUS}px;
            padding: 8px 18px;
            font-size: {BODY_SIZE}px;
        }}
        QPushButton:hover {{ background: {s['border']}; }}
        QPushButton:pressed {{ background: {s['input_bg']}; }}
    """


def _success_btn_style(s: dict) -> str:
    return f"""
        QPushButton {{
            background: {s['success']};
            color: white;
            border: none;
            border-radius: {BUTTON_RADIUS}px;
            padding: 8px 18px;
            font-weight: 600;
            font-size: {BODY_SIZE}px;
        }}
        QPushButton:hover {{ background: #059669; }}
        QPushButton:pressed {{ background: #047857; }}
    """


def _danger_btn_style(s: dict) -> str:
    return f"""
        QPushButton {{
            background: transparent;
            color: {s['error']};
            border: 1px solid {s['error']}44;
            border-radius: {BUTTON_RADIUS}px;
            padding: 6px 14px;
            font-size: {SMALL_SIZE}px;
        }}
        QPushButton:hover {{ background: {s['error']}22; }}
        QPushButton:pressed {{ background: {s['error']}33; }}
    """


def _card_style(s: dict, accent_color: str = None) -> str:
    border_left = f"border-left: 3px solid {accent_color};" if accent_color else ""
    return f"""
        QFrame {{
            background: {s['card_bg']};
            border: 1px solid {s['border']};
            {border_left}
            border-radius: {CARD_RADIUS}px;
        }}
    """


# ══════════════════════════════════════════════════════════════
#  ANA SAYFA
# ══════════════════════════════════════════════════════════════

class KaplamaPlanlamaPage(BasePage):
    """Kaplama Hatti Haftalik Planlama Sayfasi"""

    def __init__(self, theme: dict):
        super().__init__(theme)
        self.s = _get_style(theme)
        self.plan_id: int = 0
        self.urunler: list[KaplamaUrun] = []
        self.gorevler: list[PlanGorev] = []
        self._plc_data: list = []
        self._setup_ui()
        QTimer.singleShot(200, self._init_data)

    # ══════════════════════════════════════════════════════
    #  UI KURULUMU
    # ══════════════════════════════════════════════════════

    def _setup_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        splitter = QSplitter(Qt.Horizontal)
        splitter.setStyleSheet(f"QSplitter::handle {{ background: {brand.BORDER}; width: 2px; }}")

        left_widget = self._build_left_panel()
        splitter.addWidget(left_widget)

        right_widget = self._build_right_panel()
        splitter.addWidget(right_widget)

        splitter.setSizes([380, 850])
        splitter.setStretchFactor(0, 0)
        splitter.setStretchFactor(1, 1)

        main_layout.addWidget(splitter)

    # ── SOL PANEL ──

    def _build_left_panel(self) -> QWidget:
        s = self.s
        container = QWidget()
        container.setStyleSheet(f"background: {s['card_bg']};")

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet(f"""
            QScrollArea {{ border: none; background: transparent; }}
            QScrollBar:vertical {{ background: transparent; width: 8px; }}
            QScrollBar::handle:vertical {{ background: {s['border']}; border-radius: 4px; min-height: 30px; }}
            QScrollBar::handle:vertical:hover {{ background: #2A3545; }}
        """)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)

        content = QWidget()
        layout = QVBoxLayout(content)
        layout.setContentsMargins(MARGIN, MARGIN, MARGIN, MARGIN)
        layout.setSpacing(SPACING)

        # Header: Accent bar + baslik
        header = QHBoxLayout()
        header.setSpacing(0)

        accent = QFrame()
        accent.setFixedSize(4, 36)
        accent.setStyleSheet(f"background: {s['primary']}; border-radius: 2px;")
        header.addWidget(accent)

        title_section = QVBoxLayout()
        title_section.setSpacing(2)

        title = QLabel("Kaplama Planlama")
        title.setStyleSheet(f"""
            color: {s['text']};
            font-size: {TITLE_SIZE}px;
            font-weight: 600;
            margin-left: 12px;
        """)
        title_section.addWidget(title)

        self.lbl_subtitle = QLabel("Haftalik uretim planlama")
        self.lbl_subtitle.setStyleSheet(f"""
            color: {s['text_secondary']};
            font-size: {SUBTITLE_SIZE}px;
            margin-left: 12px;
        """)
        title_section.addWidget(self.lbl_subtitle)

        header.addLayout(title_section)
        header.addStretch()
        layout.addLayout(header)

        # Ozet kartlari
        self._build_summary_cards(layout)

        # Hat istatistik kartlari
        self._build_hat_istatistik(layout)

        # Banyo kartlari
        self._build_banyo_cards(layout)

        # Urun listesi
        self._build_urun_table(layout)

        # Recete goruntuleme
        self._build_recete_viewer(layout)

        layout.addStretch()
        scroll.setWidget(content)

        outer = QVBoxLayout(container)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.addWidget(scroll)
        return container

    def _build_summary_cards(self, parent_layout: QVBoxLayout):
        self._section_label(parent_layout, "Ozet")
        grid = QGridLayout()
        grid.setSpacing(CARD_SPACING)

        self.lbl_toplam_aski = self._stat_card("Toplam Yukleme", "0", brand.INFO)
        self.lbl_toplam_parca = self._stat_card("Toplam Parca", "0", brand.SUCCESS)
        self.lbl_bara_kullanim = self._stat_card("Bara Kullanim", "%0", brand.WARNING)
        self.lbl_plan_durum = self._stat_card("Durum", "Taslak", brand.TEXT_DIM)

        grid.addWidget(self.lbl_toplam_aski, 0, 0)
        grid.addWidget(self.lbl_toplam_parca, 0, 1)
        grid.addWidget(self.lbl_bara_kullanim, 1, 0)
        grid.addWidget(self.lbl_plan_durum, 1, 1)
        parent_layout.addLayout(grid)

    def _build_hat_istatistik(self, parent_layout: QVBoxLayout):
        self._section_label(parent_layout, "Hat Durumu (Canli)")
        hat_layout = QHBoxLayout()
        hat_layout.setSpacing(CARD_SPACING)

        self.card_ktl = HatIstatistikCard(self.s)
        self.card_cinko = HatIstatistikCard(self.s)

        hat_layout.addWidget(self.card_ktl)
        hat_layout.addWidget(self.card_cinko)
        parent_layout.addLayout(hat_layout)

    def _build_banyo_cards(self, parent_layout: QVBoxLayout):
        s = self.s
        self._section_label(parent_layout, "Banyolar (Canli PLC)")

        self.banyo_scroll = QScrollArea()
        self.banyo_scroll.setWidgetResizable(True)
        self.banyo_scroll.setFixedHeight(190)
        self.banyo_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        self.banyo_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.banyo_scroll.setStyleSheet(f"""
            QScrollArea {{ border: none; background: transparent; }}
            QScrollBar:horizontal {{
                height: 8px; background: transparent;
            }}
            QScrollBar::handle:horizontal {{
                background: {s['border']}; border-radius: 4px; min-width: 30px;
            }}
            QScrollBar::handle:horizontal:hover {{ background: #2A3545; }}
        """)

        self.banyo_container = QWidget()
        self.banyo_cards_layout = QHBoxLayout(self.banyo_container)
        self.banyo_cards_layout.setContentsMargins(0, 0, 0, 0)
        self.banyo_cards_layout.setSpacing(CARD_SPACING)

        self.banyo_cards: list[BanyoCard] = []
        self.banyo_scroll.setWidget(self.banyo_container)
        parent_layout.addWidget(self.banyo_scroll)

    def _build_urun_table(self, parent_layout: QVBoxLayout):
        s = self.s
        header_layout = QHBoxLayout()
        self._section_label(header_layout, "Urun Ihtiyaclari")
        header_layout.addStretch()

        btn_ekle = QPushButton("+ Urun Ekle")
        btn_ekle.setCursor(Qt.PointingHandCursor)
        btn_ekle.setStyleSheet(_primary_btn_style(s))
        btn_ekle.clicked.connect(self._on_urun_ekle)
        header_layout.addWidget(btn_ekle)
        parent_layout.addLayout(header_layout)

        self.urun_table = QTableWidget()
        self.urun_table.setColumnCount(7)
        self.urun_table.setHorizontalHeaderLabels(["Ref", "Tip", "Ihtiyac", "Aski", "Bara", "Cevrim", "Oncelik"])
        self.urun_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.urun_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.urun_table.verticalHeader().setVisible(False)
        self.urun_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.urun_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.urun_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.urun_table.setMaximumHeight(200)
        self.urun_table.setAlternatingRowColors(False)
        self.urun_table.setStyleSheet(_table_style(s))
        self.urun_table.clicked.connect(self._on_urun_table_clicked)
        parent_layout.addWidget(self.urun_table)

        btn_sil = QPushButton("Secili Urunu Sil")
        btn_sil.setCursor(Qt.PointingHandCursor)
        btn_sil.setStyleSheet(_danger_btn_style(s))
        btn_sil.clicked.connect(self._on_urun_sil)
        parent_layout.addWidget(btn_sil)

    def _build_recete_viewer(self, parent_layout: QVBoxLayout):
        self._section_label(parent_layout, "Recete Adimlari")
        self.recete_widget = ReceteAdimWidget()
        parent_layout.addWidget(self.recete_widget)

        self.lbl_recete_toplam = QLabel("Urun secin")
        self.lbl_recete_toplam.setStyleSheet(f"color: {brand.TEXT_DIM}; font-size: {LABEL_SIZE}px;")
        parent_layout.addWidget(self.lbl_recete_toplam)

    # ── SAG PANEL ──

    def _build_right_panel(self) -> QWidget:
        s = self.s
        container = QWidget()
        layout = QVBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        # Toolbar
        toolbar = self._build_toolbar()
        layout.addWidget(toolbar)

        # Bara durum strip
        self.bara_strip = BaraDurumStrip()
        layout.addWidget(self.bara_strip)

        # Tab widget
        self.tab_widget = QTabWidget()
        self.tab_widget.setStyleSheet(f"""
            QTabWidget::pane {{ border: none; background: {s['card_bg']}; }}
            QTabBar::tab {{
                background: {s['input_bg']}; color: {s['text_secondary']};
                padding: 10px 22px; border: none; border-bottom: 2px solid transparent;
                font-size: {BODY_SIZE}px; font-weight: 600;
            }}
            QTabBar::tab:selected {{
                color: {s['text']}; border-bottom: 2px solid {s['primary']};
                background: {s['card_bg']};
            }}
            QTabBar::tab:hover {{ color: {s['text']}; }}
        """)

        # Tab 1: Gantt Planlama
        gantt_container = QWidget()
        gantt_layout = QVBoxLayout(gantt_container)
        gantt_layout.setContentsMargins(0, 0, 0, 0)
        gantt_layout.setSpacing(0)
        self.gantt = GanttWidget()
        self.gantt.gorev_silindi.connect(self._on_gorev_silindi)
        gantt_layout.addWidget(self.gantt, 1)
        self.tab_widget.addTab(gantt_container, "Haftalik Plan")

        # Tab 2: Hat Canli Gorunum
        hat_container = QWidget()
        hat_layout = QVBoxLayout(hat_container)
        hat_layout.setContentsMargins(MARGIN, SPACING, MARGIN, SPACING)
        hat_layout.setSpacing(SPACING)

        lbl_ktl = QLabel("KTL Hatti (101-143)")
        lbl_ktl.setStyleSheet(f"color: {s['info']}; font-size: 14px; font-weight: 600;")
        hat_layout.addWidget(lbl_ktl)
        self.hat_ktl = HatCanliWidget()
        self.hat_ktl.setMinimumHeight(220)
        hat_layout.addWidget(self.hat_ktl)

        lbl_cinko = QLabel("Cinko-Nikel Hatti (201-247)")
        lbl_cinko.setStyleSheet(f"color: {s['warning']}; font-size: 14px; font-weight: 600;")
        hat_layout.addWidget(lbl_cinko)
        self.hat_cinko = HatCanliWidget()
        self.hat_cinko.setMinimumHeight(220)
        hat_layout.addWidget(self.hat_cinko)

        hat_layout.addStretch()
        self.tab_widget.addTab(hat_container, "Hat Canli Gorunum")

        # Tab 3: Kapasite Analizi
        self._build_kapasite_tab()

        layout.addWidget(self.tab_widget, 1)

        # Alt bilgi bari
        bottom_bar = self._build_bottom_bar()
        layout.addWidget(bottom_bar)

        return container

    def _build_kapasite_tab(self):
        s = self.s
        kap_container = QWidget()
        kap_layout = QVBoxLayout(kap_container)
        kap_layout.setContentsMargins(MARGIN, SPACING, MARGIN, SPACING)
        kap_layout.setSpacing(SPACING)

        # Parametre satiri
        kap_param = QHBoxLayout()
        kap_param.setSpacing(CARD_SPACING)

        lbl_v = QLabel("Vardiya:")
        lbl_v.setStyleSheet(f"color: {s['text']}; font-size: {BODY_SIZE}px;")
        kap_param.addWidget(lbl_v)

        self.spn_kap_vardiya = QSpinBox()
        self.spn_kap_vardiya.setRange(1, 3)
        self.spn_kap_vardiya.setValue(3)
        self.spn_kap_vardiya.setStyleSheet(_input_style(s))
        self.spn_kap_vardiya.setFixedWidth(70)
        kap_param.addWidget(self.spn_kap_vardiya)

        lbl_g = QLabel("Gun:")
        lbl_g.setStyleSheet(f"color: {s['text']}; font-size: {BODY_SIZE}px; margin-left: 8px;")
        kap_param.addWidget(lbl_g)

        self.spn_kap_gun = QSpinBox()
        self.spn_kap_gun.setRange(1, 7)
        self.spn_kap_gun.setValue(5)
        self.spn_kap_gun.setStyleSheet(_input_style(s))
        self.spn_kap_gun.setFixedWidth(70)
        kap_param.addWidget(self.spn_kap_gun)

        self.lbl_kap_toplam = QLabel("")
        self.lbl_kap_toplam.setStyleSheet(f"color: {s['text_secondary']}; font-size: {SMALL_SIZE}px; margin-left: 12px;")
        kap_param.addWidget(self.lbl_kap_toplam)

        lbl_pb = QLabel("Paralel Bara:")
        lbl_pb.setStyleSheet(f"color: {s['text']}; font-size: {BODY_SIZE}px; margin-left: 8px;")
        kap_param.addWidget(lbl_pb)

        self.spn_kap_paralel = QSpinBox()
        self.spn_kap_paralel.setRange(1, 30)
        self.spn_kap_paralel.setValue(15)
        self.spn_kap_paralel.setStyleSheet(_input_style(s))
        self.spn_kap_paralel.setFixedWidth(70)
        kap_param.addWidget(self.spn_kap_paralel)

        lbl_cv = QLabel("Çevrim (dk):")
        lbl_cv.setStyleSheet(f"color: {s['text']}; font-size: {BODY_SIZE}px; margin-left: 8px;")
        kap_param.addWidget(lbl_cv)

        self.spn_kap_cevrim = QSpinBox()
        self.spn_kap_cevrim.setRange(10, 300)
        self.spn_kap_cevrim.setValue(90)
        self.spn_kap_cevrim.setStyleSheet(_input_style(s))
        self.spn_kap_cevrim.setFixedWidth(70)
        kap_param.addWidget(self.spn_kap_cevrim)

        lbl_vb = QLabel("Vardiya Başlangıç:")
        lbl_vb.setStyleSheet(f"color: {s['text']}; font-size: {BODY_SIZE}px; margin-left: 8px;")
        kap_param.addWidget(lbl_vb)

        self.time_vardiya_bas = QTimeEdit()
        self.time_vardiya_bas.setTime(QTime(7, 30))
        self.time_vardiya_bas.setDisplayFormat("HH:mm")
        self.time_vardiya_bas.setStyleSheet(_input_style(s))
        self.time_vardiya_bas.setFixedWidth(80)
        kap_param.addWidget(self.time_vardiya_bas)

        lbl_st = QLabel("Setup (dk):")
        lbl_st.setStyleSheet(f"color: {s['text']}; font-size: {BODY_SIZE}px; margin-left: 8px;")
        kap_param.addWidget(lbl_st)

        self.spn_kap_setup = QSpinBox()
        self.spn_kap_setup.setRange(1, 60)
        self.spn_kap_setup.setValue(10)
        self.spn_kap_setup.setStyleSheet(_input_style(s))
        self.spn_kap_setup.setFixedWidth(60)
        kap_param.addWidget(self.spn_kap_setup)

        lbl_ktl = QLabel("KTL Bara/Gün:")
        lbl_ktl.setStyleSheet(f"color: {s['text']}; font-size: {BODY_SIZE}px; margin-left: 8px;")
        kap_param.addWidget(lbl_ktl)

        self.spn_ktl_bara = QSpinBox()
        self.spn_ktl_bara.setRange(0, 500)
        self.spn_ktl_bara.setValue(130)
        self.spn_ktl_bara.setStyleSheet(_input_style(s))
        self.spn_ktl_bara.setFixedWidth(70)
        self.spn_ktl_bara.setToolTip("KTL hattından günlük gelen bara sayısı\n(ortak yağ alma kapasitesini düşürür)")
        kap_param.addWidget(self.spn_ktl_bara)

        lbl_sya = QLabel("SYA Tank:")
        lbl_sya.setStyleSheet(f"color: {s['text']}; font-size: {BODY_SIZE}px; margin-left: 8px;")
        kap_param.addWidget(lbl_sya)

        self.spn_sya_tank = QSpinBox()
        self.spn_sya_tank.setRange(1, 20)
        self.spn_sya_tank.setValue(4)
        self.spn_sya_tank.setStyleSheet(_input_style(s))
        self.spn_sya_tank.setFixedWidth(50)
        self.spn_sya_tank.setToolTip("SYA (Sıcak Yağ Alma) tank sayısı\nK5, K6, K7, K8 = 4 tank")
        kap_param.addWidget(self.spn_sya_tank)

        lbl_yag = QLabel("Yağ Alma (dk):")
        lbl_yag.setStyleSheet(f"color: {s['text']}; font-size: {BODY_SIZE}px; margin-left: 8px;")
        kap_param.addWidget(lbl_yag)

        self.spn_yag_alma = QSpinBox()
        self.spn_yag_alma.setRange(1, 120)
        self.spn_yag_alma.setValue(15)
        self.spn_yag_alma.setStyleSheet(_input_style(s))
        self.spn_yag_alma.setFixedWidth(60)
        self.spn_yag_alma.setToolTip("Her baranın yağ almada kalma süresi (dk)")
        kap_param.addWidget(self.spn_yag_alma)

        btn_kap_hesapla = QPushButton("Hesapla")
        btn_kap_hesapla.setCursor(Qt.PointingHandCursor)
        btn_kap_hesapla.setStyleSheet(_primary_btn_style(s))
        btn_kap_hesapla.clicked.connect(self._hesapla_kapasite)
        kap_param.addWidget(btn_kap_hesapla)

        btn_excel = QPushButton("Excel Yükle")
        btn_excel.setCursor(Qt.PointingHandCursor)
        btn_excel.setStyleSheet(_success_btn_style(s))
        btn_excel.clicked.connect(self._on_excel_yukle)
        kap_param.addWidget(btn_excel)

        btn_excel_aktar = QPushButton("Excel Aktar")
        btn_excel_aktar.setCursor(Qt.PointingHandCursor)
        btn_excel_aktar.setStyleSheet(_secondary_btn_style(s))
        btn_excel_aktar.clicked.connect(self._on_plan_excel_aktar)
        kap_param.addWidget(btn_excel_aktar)

        btn_is_emri = QPushButton("İş Emri Oluştur")
        btn_is_emri.setCursor(Qt.PointingHandCursor)
        btn_is_emri.setStyleSheet(_primary_btn_style(s))
        btn_is_emri.clicked.connect(self._on_is_emri_olustur)
        kap_param.addWidget(btn_is_emri)

        kap_param.addStretch()
        kap_layout.addLayout(kap_param)

        # ── Alt tab: Kapasite Görsel + Çevrim Plan ──
        _scroll_style = f"""
            QScrollArea {{ border: none; background: transparent; }}
            QScrollBar:vertical {{ background: transparent; width: 8px; }}
            QScrollBar::handle:vertical {{ background: {s['border']}; border-radius: 4px; min-height: 30px; }}
        """

        self.kap_alt_tabs = QTabWidget()
        self.kap_alt_tabs.setStyleSheet(f"""
            QTabWidget::pane {{ border: none; background: {s['card_bg']}; }}
            QTabBar::tab {{
                background: {s['input_bg']}; color: {s['text_secondary']};
                padding: 8px 18px; border: none; border-bottom: 2px solid transparent;
                font-size: {SMALL_SIZE}px;
            }}
            QTabBar::tab:selected {{
                color: {s['text']}; border-bottom: 2px solid {s['success']};
                background: {s['card_bg']};
            }}
        """)

        # Alt Tab 1: Kapasite bar chart
        scroll_gorsel = QScrollArea()
        scroll_gorsel.setWidgetResizable(True)
        scroll_gorsel.setStyleSheet(_scroll_style)
        self.kapasite_gorsel = KapasiteGorselWidget()
        scroll_gorsel.setWidget(self.kapasite_gorsel)
        self.kap_alt_tabs.addTab(scroll_gorsel, "Kapasite Özet")

        # Alt Tab 2: Çevrim planı
        scroll_cevrim = QScrollArea()
        scroll_cevrim.setWidgetResizable(True)
        scroll_cevrim.setStyleSheet(_scroll_style)
        self.cevrim_plan_widget = CevrimPlanWidget()
        scroll_cevrim.setWidget(self.cevrim_plan_widget)
        self.kap_alt_tabs.addTab(scroll_cevrim, "Çevrim Planı")

        kap_layout.addWidget(self.kap_alt_tabs, 2)

        # Kapasite ozet kartlari
        self.kap_cards_layout = QHBoxLayout()
        self.kap_cards_layout.setSpacing(CARD_SPACING)
        kap_layout.addLayout(self.kap_cards_layout)

        # Recete kapasite tablosu
        self.tbl_kapasite = QTableWidget()
        self.tbl_kapasite.setSelectionBehavior(QTableWidget.SelectRows)
        self.tbl_kapasite.setSelectionMode(QTableWidget.SingleSelection)
        self.tbl_kapasite.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl_kapasite.verticalHeader().setVisible(False)
        self.tbl_kapasite.setAlternatingRowColors(False)
        self.tbl_kapasite.setStyleSheet(_table_style(s))
        kap_layout.addWidget(self.tbl_kapasite)

        # Darbogaz bilgisi
        self.lbl_darbogaz = QLabel("")
        self.lbl_darbogaz.setWordWrap(True)
        self.lbl_darbogaz.setStyleSheet(f"""
            color: {s['warning']};
            font-size: {SMALL_SIZE}px;
            padding: 10px;
            background: {s['card_bg']};
            border: 1px solid {s['warning']}33;
            border-left: 3px solid {s['warning']};
            border-radius: {CARD_RADIUS}px;
        """)
        kap_layout.addWidget(self.lbl_darbogaz)

        self.tab_widget.addTab(kap_container, "Kapasite Analizi")

    def _build_toolbar(self) -> QFrame:
        s = self.s
        toolbar = QFrame()
        toolbar.setFixedHeight(54)
        toolbar.setStyleSheet(f"""
            QFrame {{
                background: {s['card_bg']};
                border-bottom: 1px solid {s['border']};
            }}
        """)

        layout = QHBoxLayout(toolbar)
        layout.setContentsMargins(MARGIN, 8, MARGIN, 8)
        layout.setSpacing(CARD_SPACING)

        lbl = QLabel("Hafta:")
        lbl.setStyleSheet(f"color: {s['text_secondary']}; font-size: {BODY_SIZE}px;")
        layout.addWidget(lbl)

        self.date_edit = QDateEdit()
        self.date_edit.setCalendarPopup(True)
        monday = _monday_of_week(date.today())
        self.date_edit.setDate(QDate(monday.year, monday.month, monday.day))
        self.date_edit.setDisplayFormat("dd.MM.yyyy")
        self.date_edit.setStyleSheet(_input_style(s))
        self.date_edit.setFixedWidth(130)
        self.date_edit.dateChanged.connect(self._on_hafta_degisti)
        layout.addWidget(self.date_edit)

        # Canli gosterge
        self.lbl_canli = QLabel("CANLI")
        self._set_canli_style(False)
        layout.addWidget(self.lbl_canli)

        layout.addStretch()

        # Butonlar
        for text, style_fn, handler in [
            ("Temizle", _secondary_btn_style, self._on_temizle),
            ("Otomatik Planla", _primary_btn_style, self._on_otomatik_planla),
            ("Kaydet", _success_btn_style, self._on_kaydet),
        ]:
            btn = QPushButton(text)
            btn.setCursor(Qt.PointingHandCursor)
            btn.setStyleSheet(style_fn(s))
            btn.clicked.connect(handler)
            layout.addWidget(btn)

        return toolbar

    def _build_bottom_bar(self) -> QFrame:
        s = self.s
        bar = QFrame()
        bar.setFixedHeight(40)
        bar.setStyleSheet(f"""
            QFrame {{
                background: {s['card_bg']};
                border-top: 1px solid {s['border']};
            }}
        """)
        layout = QHBoxLayout(bar)
        layout.setContentsMargins(MARGIN, 4, MARGIN, 4)
        layout.setSpacing(MARGIN)

        self.lbl_planlanan = QLabel("Planlanan: 0 adet")
        self.lbl_planlanan.setStyleSheet(f"color: {s['text_secondary']}; font-size: {SMALL_SIZE}px;")
        layout.addWidget(self.lbl_planlanan)

        self.lbl_bekleyen = QLabel("Bekleyen: 0dk")
        self.lbl_bekleyen.setStyleSheet(f"color: {s['warning']}; font-size: {SMALL_SIZE}px;")
        layout.addWidget(self.lbl_bekleyen)

        self.lbl_acil = QLabel("Acil: 0")
        self.lbl_acil.setStyleSheet(f"color: {s['error']}; font-size: {SMALL_SIZE}px;")
        layout.addWidget(self.lbl_acil)

        layout.addStretch()
        for label, color in [("Cinko", "#F59E0B"), ("Cinko-Nikel", "#3B82F6"), ("Acil", "#EF4444")]:
            dot = QLabel(f"* {label}")
            dot.setStyleSheet(f"color: {color}; font-size: {SMALL_SIZE}px;")
            layout.addWidget(dot)

        return bar

    # ── YARDIMCI UI ──

    def _section_label(self, parent, text: str):
        lbl = QLabel(text.upper())
        lbl.setStyleSheet(f"""
            color: {brand.TEXT_DIM};
            font-size: {LABEL_SIZE}px;
            font-weight: 600;
            letter-spacing: 1px;
        """)
        if isinstance(parent, (QVBoxLayout, QHBoxLayout)):
            parent.addWidget(lbl)

    def _stat_card(self, label: str, value: str, color: str) -> QFrame:
        s = self.s
        frame = QFrame()
        frame.setStyleSheet(_card_style(s, accent_color=color))

        layout = QHBoxLayout(frame)
        layout.setContentsMargins(14, 10, 14, 10)
        layout.setSpacing(10)

        text_layout = QVBoxLayout()
        text_layout.setSpacing(2)

        lbl = QLabel(label.upper())
        lbl.setStyleSheet(f"""
            color: {s['text_muted']};
            font-size: {LABEL_SIZE}px;
            font-weight: 600;
            letter-spacing: 1px;
            border: none;
        """)
        text_layout.addWidget(lbl)

        val = QLabel(value)
        val.setObjectName("stat_value")
        val.setStyleSheet(f"""
            color: {color};
            font-size: 18px;
            font-weight: 700;
            border: none;
        """)
        text_layout.addWidget(val)

        layout.addLayout(text_layout)
        layout.addStretch()
        return frame

    def _update_stat(self, frame: QFrame, value: str):
        lbl = frame.findChild(QLabel, "stat_value")
        if lbl:
            lbl.setText(value)

    def _set_canli_style(self, flash: bool = False):
        try:
            if not self.lbl_canli or not self.lbl_canli.isVisible():
                return
            s = self.s
            alpha_bg = "33" if flash else "22"
            alpha_border = "66" if flash else "44"
            self.lbl_canli.setStyleSheet(f"""
                color: {s['success']};
                font-size: {LABEL_SIZE}px;
                font-weight: 600;
                background: {s['success']}{alpha_bg};
                border: 1px solid {s['success']}{alpha_border};
                border-radius: 4px;
                padding: 3px 10px;
                letter-spacing: 1px;
            """)
        except RuntimeError:
            pass

    # ══════════════════════════════════════════════════════
    #  VERI ISLEMLERI
    # ══════════════════════════════════════════════════════

    def _init_data(self):
        db.ensure_tables()
        db.seed_recete_tanimlari()
        db.fix_hat_kodlari()
        db.sync_recete_sureleri()
        self._refresh_plc_data()
        self._load_plan()

        # 10sn'de bir PLC verisi yenile
        self._refresh_timer = QTimer(self)
        self._refresh_timer.timeout.connect(self._refresh_plc_data)
        self._refresh_timer.start(10000)

    def _refresh_plc_data(self):
        try:
            self._plc_data = db.get_plc_canli()

            hat_stats = db.get_hat_istatistik()
            self.card_ktl.set_data("KTL Hatti", hat_stats.get('KTL', {}))
            self.card_cinko.set_data("Cinko-Nikel Hatti", hat_stats.get('ZNNI', {}))

            self._update_banyo_cards()

            ktl_kazanlar = [k for k in self._plc_data if k.get('hat_kodu') == 'KTL']
            cinko_kazanlar = [k for k in self._plc_data if k.get('hat_kodu') == 'ZNNI']
            self.hat_ktl.set_kazanlar(ktl_kazanlar)
            self.hat_cinko.set_kazanlar(cinko_kazanlar)

            self._update_bara_strip()

            self._set_canli_style(True)
            QTimer.singleShot(500, lambda: self._set_canli_style(False))

        except Exception as e:
            print(f"[KaplamaPlanlama] PLC veri guncelleme hatasi: {e}")

    def _update_banyo_cards(self):
        for card in self.banyo_cards:
            card.deleteLater()
        self.banyo_cards.clear()

        for kz in self._plc_data:
            kazan_no = kz.get('kazan_no', 0)
            if kazan_no < 100:
                continue

            banyo_adi = kz.get('banyo_adi', f"K{kazan_no}")
            recete_adi = kz.get('recete_adi', '')
            recete_acik = kz.get('recete_aciklama', '')

            card = BanyoCard(self.s)
            card.set_data({
                'ad': banyo_adi,
                'kazan_no': kazan_no,
                'sicaklik': kz.get('sicaklik', 0),
                'sicaklik_min': kz.get('sicaklik_min', 0),
                'sicaklik_max': kz.get('sicaklik_max', 100),
                'sicaklik_hedef': kz.get('sicaklik_hedef', 0),
                'durum': kz.get('durum', 'BELIRSIZ'),
                'durum_dakika': kz.get('durum_dakika', 0),
                'recete_no': kz.get('recete_no'),
                'recete_adi': recete_adi,
                'recete_aciklama': recete_acik,
                'recete_sure_dk': kz.get('recete_sure_dk'),
                'akim': kz.get('akim', 0),
                'son_bara': kz.get('son_bara'),
                'hat_kodu': kz.get('hat_kodu', ''),
                'aktif_aski': 1 if kz.get('durum') == 'AKTIF' else 0,
                'max_aski': 4,
            })
            self.banyo_cards.append(card)
            self.banyo_cards_layout.addWidget(card)

        if not self.banyo_cards:
            lbl = QLabel("PLC verisi bekleniyor...")
            lbl.setStyleSheet(f"color: {brand.TEXT_DIM}; font-size: {SMALL_SIZE}px; padding: 20px;")
            lbl.setAlignment(Qt.AlignCenter)
            self.banyo_cards_layout.addWidget(lbl)

    def _update_bara_strip(self):
        bara_data = []
        for i in range(1, BARA_SAYISI + 1):
            aktif_gorev = None
            for g in self.gorevler:
                if g.bara_no == i:
                    aktif_gorev = g
                    break

            if aktif_gorev:
                bara_data.append({
                    'bara_no': i,
                    'durum': 'PLANLANDI',
                    'urun_ref': aktif_gorev.urun_ref,
                })
            else:
                bara_data.append({'bara_no': i, 'durum': 'BOS'})

        self.bara_strip.set_data(bara_data)

    def _get_selected_monday(self) -> date:
        qd = self.date_edit.date()
        d = date(qd.year(), qd.month(), qd.day())
        return _monday_of_week(d)

    def _load_plan(self):
        monday = self._get_selected_monday()
        self.plan_id = db.get_or_create_plan(monday) or 0
        if self.plan_id:
            self.urunler = db.load_urunler(self.plan_id)
            self.gorevler = db.load_gorevler(self.plan_id)
        else:
            self.urunler = []
            self.gorevler = []
        self._refresh_ui()

    def _refresh_ui(self):
        self._refresh_urun_table()
        self.gantt.set_gorevler(self.gorevler)
        self._refresh_summary()
        self._update_bara_strip()

    def _refresh_urun_table(self):
        self.urun_table.setRowCount(len(self.urunler))
        for i, u in enumerate(self.urunler):
            self.urun_table.setItem(i, 0, QTableWidgetItem(u.ref))

            tip_item = QTableWidgetItem("ZnNi" if u.tip == "zn-ni" else "Zn")
            tip_item.setForeground(QColor("#3B82F6") if u.tip == "zn-ni" else QColor("#F59E0B"))
            self.urun_table.setItem(i, 1, tip_item)

            self.urun_table.setItem(i, 2, QTableWidgetItem(f"{u.haftalik_ihtiyac:,}"))

            aski_item = QTableWidgetItem(f"{u.stok_aski}x{u.kapasite}")
            if not u.aski_yeterli:
                aski_item.setForeground(QColor("#EF4444"))
            self.urun_table.setItem(i, 3, aski_item)

            self.urun_table.setItem(i, 4, QTableWidgetItem(f"{u.gerekli_bara}"))
            self.urun_table.setItem(i, 5, QTableWidgetItem(f"{u.toplam_cevrim}x{u.cevrim_suresi}dk"))

            oncelik_item = QTableWidgetItem(u.oncelik.upper())
            if u.oncelik == "acil":
                oncelik_item.setForeground(QColor("#EF4444"))
            self.urun_table.setItem(i, 6, oncelik_item)

    def _refresh_summary(self):
        toplam_aski_yukleme = sum(g.aski_sayisi for g in self.gorevler)
        toplam_parca = 0
        for u in self.urunler:
            urun_gorevleri = [g for g in self.gorevler if g.urun_ref == u.ref]
            toplam_gorev_sure = sum(g.sure_dk for g in urun_gorevleri)
            cevrim_sayisi = toplam_gorev_sure // u.cevrim_suresi if u.cevrim_suresi > 0 else 0
            toplam_parca += cevrim_sayisi * u.stok_aski * u.kapasite

        kullanilan_slotlar = set()
        for g in self.gorevler:
            kullanilan_slotlar.add((g.bara_no, g.gun, g.vardiya))
        max_slots = BARA_SAYISI * GUN_SAYISI * VARDIYA_SAYISI
        kullanim = int((len(kullanilan_slotlar) / max_slots * 100)) if max_slots > 0 else 0

        self._update_stat(self.lbl_toplam_aski, f"{toplam_aski_yukleme:,}")
        self._update_stat(self.lbl_toplam_parca, f"{toplam_parca:,}")
        self._update_stat(self.lbl_bara_kullanim, f"%{kullanim}")
        durum = db.get_plan_durum(self.plan_id) if self.plan_id else "taslak"
        self._update_stat(self.lbl_plan_durum, durum.capitalize())

        planlanan_sure = sum(g.sure_dk for g in self.gorevler)
        toplam_ihtiyac_sure = sum(u.toplam_sure_dk for u in self.urunler)
        bekleyen_sure = max(0, toplam_ihtiyac_sure - planlanan_sure)
        acil_sayisi = sum(1 for g in self.gorevler if g.acil)

        self.lbl_planlanan.setText(f"Planlanan: {toplam_parca:,} adet")
        self.lbl_bekleyen.setText(f"Bekleyen: {bekleyen_sure}dk")
        self.lbl_acil.setText(f"Acil: {acil_sayisi}")

    # ══════════════════════════════════════════════════════
    #  OLAY ISLEYICILERI
    # ══════════════════════════════════════════════════════

    def _on_hafta_degisti(self):
        self._load_plan()

    def _on_urun_ekle(self):
        dialog = UrunEkleDialog(self.s, self)
        if dialog.exec() == QDialog.Accepted:
            urun = dialog.get_urun()
            urun.id = len(self.urunler) + 1
            self.urunler.append(urun)
            self._refresh_ui()

    def _on_urun_sil(self):
        row = self.urun_table.currentRow()
        if row < 0:
            return
        ref = self.urunler[row].ref
        self.gorevler = [g for g in self.gorevler if g.urun_ref != ref]
        self.urunler.pop(row)
        self._refresh_ui()

    def _on_urun_table_clicked(self):
        row = self.urun_table.currentRow()
        if row < 0 or row >= len(self.urunler):
            return
        urun = self.urunler[row]
        recete = db.get_urun_recete(urun.ref)
        if recete:
            self.recete_widget.set_adimlar(recete)
            toplam_sn = sum(a['sure_sn'] for a in recete)
            self.lbl_recete_toplam.setText(
                f"{urun.ref}: {len(recete)} adim, toplam {toplam_sn}sn ({toplam_sn // 60}dk {toplam_sn % 60}sn)"
            )
        else:
            self.recete_widget.set_adimlar([])
            self.lbl_recete_toplam.setText(f"{urun.ref}: Recete bulunamadi")

    def _on_temizle(self):
        reply = QMessageBox.question(self, "Temizle", "Tum gorevler silinecek. Emin misiniz?",
                                     QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            self.gorevler = []
            self._refresh_ui()

    def _on_otomatik_planla(self):
        if not self.urunler:
            QMessageBox.warning(self, "Uyari", "Planlamak icin once urun ekleyin.")
            return
        sonuc = otomatik_planla(self.urunler)
        self.gorevler = sonuc.gorevler
        self._refresh_ui()
        if sonuc.uyarilar:
            msg = "\n".join(sonuc.uyarilar)
            QMessageBox.information(self, "Planlama Sonucu", msg)

    def _on_kaydet(self):
        if not self.plan_id:
            QMessageBox.warning(self, "Hata", "Plan olusturulamadi.")
            return
        ok1 = db.save_urunler(self.plan_id, self.urunler)
        if ok1:
            self.urunler = db.load_urunler(self.plan_id)
            ok2 = db.save_gorevler(self.plan_id, self.urunler, self.gorevler)
            if ok2:
                QMessageBox.information(self, "Basarili", "Plan kaydedildi.")
                self._load_plan()
                return
        QMessageBox.warning(self, "Hata", "Plan kaydedilemedi.")

    def _on_gorev_silindi(self, gorev: PlanGorev):
        self.gorevler = self.gantt.get_gorevler()
        self._refresh_summary()

    # ── Kapasite Analizi ──

    def _hesapla_kapasite(self):
        s = self.s
        vardiya = self.spn_kap_vardiya.value()
        gun = self.spn_kap_gun.value()

        try:
            data = db.hesapla_kapasite(vardiya, gun)
        except Exception as e:
            QMessageBox.warning(self, "Hata", f"Kapasite hesaplanamadi:\n{e}")
            return

        haftalik_dk = data.get('haftalik_dk', 0)
        self.lbl_kap_toplam.setText(
            f"Haftalik: {haftalik_dk} dk  ({haftalik_dk // 60} saat)  |  "
            f"{vardiya} vardiya x {gun} gun"
        )

        # Ozet kartlarini temizle ve yeniden olustur
        while self.kap_cards_layout.count():
            item = self.kap_cards_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        hat_renk = {'KTL': s['info'], 'ZNNI': s['warning'], 'ON': s['success']}
        hatlar = data.get('hatlar', {})
        receteler = data.get('receteler', [])

        for hat_tipi in ['KTL', 'ZNNI', 'ON']:
            hat_info = hatlar.get(hat_tipi, {})
            rec_sayisi = hat_info.get('recete_sayisi', 0)
            hat_receteler = [r for r in receteler if r['hat_tipi'] == hat_tipi]
            aktif = sum(1 for r in hat_receteler if r.get('gercek_cevrim_7gun', 0) > 0)
            color = hat_renk.get(hat_tipi, s['border'])

            card = QFrame()
            card.setStyleSheet(_card_style(s, accent_color=color))

            cl = QVBoxLayout(card)
            cl.setContentsMargins(14, 10, 14, 10)
            cl.setSpacing(4)

            lbl_title = QLabel(f"{hat_tipi} Hatti")
            lbl_title.setStyleSheet(f"color: {color}; font-size: {BODY_SIZE}px; font-weight: 600; border: none;")
            cl.addWidget(lbl_title)

            lbl_det = QLabel(f"{rec_sayisi} recete  |  {aktif} aktif (son 7 gun)")
            lbl_det.setStyleSheet(f"color: {s['text_secondary']}; font-size: {LABEL_SIZE}px; border: none;")
            cl.addWidget(lbl_det)

            self.kap_cards_layout.addWidget(card)

        self.kap_cards_layout.addStretch()

        # Tablo doldur
        cols = ["Recete", "Adi", "Hat", "Cevrim (dk)", "Maks/Hafta", "Gercek (7gn)", "Haftalik Ort", "Kullanim %"]
        self.tbl_kapasite.setColumnCount(len(cols))
        self.tbl_kapasite.setHorizontalHeaderLabels(cols)
        self.tbl_kapasite.setRowCount(len(receteler))

        for i, r in enumerate(receteler):
            vals = [
                str(r['recete_no']),
                r.get('recete_adi') or '-',
                r.get('hat_tipi', ''),
                f"{r.get('cevrim_dk', 0):.1f}",
                str(r.get('max_cevrim_haftalik', 0)),
                str(r.get('gercek_cevrim_7gun', 0)),
                str(r.get('gercek_haftalik_ort', 0)),
                f"{r.get('kullanim_pct', 0):.1f}",
            ]
            for j, v in enumerate(vals):
                item = QTableWidgetItem(v)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                if j >= 3:
                    item.setTextAlignment(Qt.AlignCenter)
                if j == 7:
                    pct = r.get('kullanim_pct', 0)
                    if pct >= 80:
                        item.setForeground(QColor(s['error']))
                    elif pct >= 40:
                        item.setForeground(QColor(s['warning']))
                    elif pct > 0:
                        item.setForeground(QColor(s['success']))
                if j == 2:
                    c = hat_renk.get(r.get('hat_tipi', ''), s['text'])
                    item.setForeground(QColor(c))
                self.tbl_kapasite.setItem(i, j, item)

        header = self.tbl_kapasite.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.Stretch)
        for c in range(2, len(cols)):
            header.setSectionResizeMode(c, QHeaderView.ResizeToContents)

        # Darbogaz bilgisi
        darbogaz = data.get('darbogaz_kazanlar', [])
        if darbogaz:
            top5 = darbogaz[:5]
            lines = ["Darbogaz Kazanlar (son 7 gun, en yogun):"]
            for d in top5:
                lines.append(
                    f"  K{d['kazan_no']}: {d['recete_sayisi']} recete, "
                    f"{d['toplam_islem']} islem, "
                    f"doluluk %{d['doluluk_pct']}"
                )
            self.lbl_darbogaz.setText("\n".join(lines))
        else:
            self.lbl_darbogaz.setText("Son 7 gunde darbogaz verisi bulunamadi.")

        # Görsel widget'ı ürün listesinden besle
        self._update_kapasite_gorsel()

    def _update_kapasite_gorsel(self, gorsel_data=None):
        """Kapasite görsel widget'ını güncelle"""
        import math

        if gorsel_data is None:
            gorsel_data = []
            for u in self.urunler:
                if u.haftalik_ihtiyac <= 0:
                    continue
                bara_adeti = u.kapasite * u.stok_aski if u.kapasite > 0 and u.stok_aski > 0 else 1
                bara_aski = u.bara_aski if u.bara_aski > 0 else 1
                cevrim_uretim = bara_adeti * bara_aski
                gerekli_cevrim = math.ceil(u.haftalik_ihtiyac / max(cevrim_uretim, 1))
                bara_gerekli = gerekli_cevrim * bara_aski
                gorsel_data.append({
                    'adi': u.ref[:20],
                    'tip': 'cinko' if u.tip == 'zn' else 'nikel',
                    'ihtiyac': u.haftalik_ihtiyac,
                    'bara_adeti': bara_adeti,
                    'bara_aski': bara_aski,
                    'bara_gerekli': bara_gerekli,
                })

        paralel = self.spn_kap_paralel.value()
        cevrim = self.spn_kap_cevrim.value()
        vardiya = self.spn_kap_vardiya.value()
        vardiya_bas = self.time_vardiya_bas.time().toString("HH:mm")
        setup = self.spn_kap_setup.value()
        ktl_bara = self.spn_ktl_bara.value()

        if hasattr(self, 'kapasite_gorsel'):
            self.kapasite_gorsel.set_data(gorsel_data, paralel, cevrim, vardiya)

        # Çevrim planını da güncelle
        if hasattr(self, 'cevrim_plan_widget') and gorsel_data:
            sya_tank = self.spn_sya_tank.value()
            yag_alma = self.spn_yag_alma.value()
            plan = optimize_cevrim_plan(gorsel_data, paralel, cevrim, vardiya,
                                        vardiya_baslangic=vardiya_bas,
                                        setup_dk=setup,
                                        ktl_bara_gun=ktl_bara,
                                        sya_tank=sya_tank,
                                        yag_alma_dk=yag_alma)
            self.cevrim_plan_widget.set_plan(plan, paralel, cevrim)

    def _on_excel_yukle(self):
        """Excel dosyasından ürün/ihtiyaç verisi yükle ve görsele aktar"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Excel Dosyası Seç", "",
            "Excel Dosyaları (*.xlsx *.xls);;Tüm Dosyalar (*)"
        )
        if not file_path:
            return

        try:
            import openpyxl
            import math

            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb[wb.sheetnames[0]]

            gorsel_data = []
            # Excel format: Firma | Stok Kodu | Stok Adı | Proses | Günlük İhtiyaç | Bara Adeti | Askı Sayısı
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                if not row or len(row) < 6:
                    continue
                stok_adi = row[2]
                proses = str(row[3] or '').strip().lower()
                ihtiyac = row[4]
                bara_adeti = row[5]
                aski_sayisi = row[6] if len(row) > 6 else None

                if not stok_adi or not ihtiyac:
                    continue

                try:
                    ihtiyac = int(ihtiyac)
                    bara_adeti = int(bara_adeti) if bara_adeti else 1
                    aski_sayisi = int(aski_sayisi) if aski_sayisi else 0
                except (ValueError, TypeError):
                    continue

                if ihtiyac <= 0:
                    continue

                # Tip belirleme
                if 'inko' in proses or 'zn' in proses:
                    tip = 'cinko'
                else:
                    tip = 'nikel'

                bara_aski = aski_sayisi if aski_sayisi > 0 else 1
                cevrim_uretim = bara_adeti * bara_aski
                gerekli_cevrim = math.ceil(ihtiyac / max(cevrim_uretim, 1))
                bara_gerekli = gerekli_cevrim * bara_aski

                gorsel_data.append({
                    'adi': str(stok_adi)[:20],
                    'tip': tip,
                    'ihtiyac': ihtiyac,
                    'bara_adeti': bara_adeti,
                    'bara_aski': bara_aski,
                    'bara_gerekli': bara_gerekli,
                })

            if not gorsel_data:
                QMessageBox.warning(self, "Uyarı", "Excel'de geçerli ürün verisi bulunamadı.")
                return

            self._update_kapasite_gorsel(gorsel_data)

            toplam_cevrim = sum(
                math.ceil(u['ihtiyac'] / max(u['bara_adeti'] * u['bara_aski'], 1))
                for u in gorsel_data
            )
            QMessageBox.information(
                self, "Başarılı",
                f"{len(gorsel_data)} ürün yüklendi.\n"
                f"Toplam: {toplam_cevrim} çevrim ihtiyaç"
            )

        except ImportError:
            QMessageBox.warning(self, "Hata", "openpyxl kütüphanesi bulunamadı.\npip install openpyxl")
        except Exception as e:
            QMessageBox.warning(self, "Hata", f"Excel okuma hatası:\n{e}")

    # ──────────────────────────────────────────────────────────
    #  EXCEL AKTAR - Saatlik çevrim planını Excel'e yaz
    # ──────────────────────────────────────────────────────────
    def _on_plan_excel_aktar(self):
        """Çevrim planını saatlik detaylı Excel'e aktarır"""
        if not hasattr(self, 'cevrim_plan_widget') or not self.cevrim_plan_widget._plan:
            QMessageBox.warning(self, "Uyarı", "Önce Hesapla ile çevrim planı oluşturun.")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self, "Excel Kaydet", f"cevrim_plani_{QDate.currentDate().toString('yyyyMMdd')}.xlsx",
            "Excel Dosyaları (*.xlsx)"
        )
        if not file_path:
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            plan = self.cevrim_plan_widget._plan
            cevrimler = plan['cevrimler']
            baralar = plan.get('baralar', [])
            zaman = plan.get('zaman', [])
            vardiyalar = plan.get('vardiyalar', [])
            aski_analiz = plan['aski_analiz']
            ozet = plan['ozet']
            paralel = self.spn_kap_paralel.value()

            wb = openpyxl.Workbook()

            # Stiller
            baslik_font = Font(bold=True, size=12, color="FFFFFF")
            baslik_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
            header_font = Font(bold=True, size=10, color="FFFFFF")
            header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
            vardiya_font = Font(bold=True, size=10, color="F59E0B")
            vardiya_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
            ref_font = Font(bold=True, size=10, color="3B82F6")
            ref_fill = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            center = Alignment(horizontal='center', vertical='center')

            # ══════════════════════════════════════════════════
            # Sayfa 1: Bara Detay (referans referans, her bara ayrı satır)
            # ══════════════════════════════════════════════════
            ws1 = wb.active
            ws1.title = "Bara Detay"

            # Başlık
            ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
            c = ws1.cell(row=1, column=1, value="KAPLAMA HATTI - BARA BAZLI PLAN")
            c.font = baslik_font
            c.fill = baslik_fill
            c.alignment = center

            setup_dk = ozet.get('setup_dk', 10)
            ktl_bara = ozet.get('ktl_bara_gun', 0)
            ws1.cell(row=2, column=1, value=f"Başlangıç: {ozet.get('baslangic', '07:30')}")
            ws1.cell(row=2, column=3, value=f"Setup: {setup_dk} dk")
            ws1.cell(row=2, column=5, value=f"Toplam: {ozet.get('toplam_bara', 0)} bara")
            ws1.cell(row=2, column=7, value=f"Süre: {ozet.get('toplam_sure_dk', 0)} dk")
            if ktl_bara > 0:
                ws1.cell(row=3, column=1,
                         value=f"KTL Bara/Gün: {ktl_bara}")
                kalan = ozet.get('yag_alma_kalan', 0)
                c3 = ws1.cell(row=3, column=3,
                              value=f"Yağ Alma Kalan: {kalan} bara")
                if kalan < ozet.get('toplam_bara', 0):
                    c3.font = Font(color="EF4444", bold=True)

            headers = ["Sıra", "Referans", "Tip", "Bara No", "Çevrim",
                        "Vardiya", "Giriş", "Çıkış", "Adet/Bara", "Askı"]
            hdr_row = 5
            for ci, h in enumerate(headers, 1):
                c = ws1.cell(row=hdr_row, column=ci, value=h)
                c.font = header_font
                c.fill = header_fill
                c.alignment = center
                c.border = thin_border

            row = hdr_row + 1
            prev_ref = None
            prev_vardiya = 0
            tarih_str = QDate.currentDate().toString('yyyyMMdd')

            for bi, b in enumerate(baralar):
                v_no = b['vardiya']

                # Vardiya ayırıcı
                if v_no != prev_vardiya:
                    prev_vardiya = v_no
                    vd = next((v for v in vardiyalar if v['no'] == v_no), None)
                    vd_txt = f"── Vardiya {v_no}"
                    if vd:
                        vd_txt += f": {vd['baslangic']} - {vd['bitis']} ({vd.get('bara_adet', 0)} bara)"
                    vd_txt += " ──"
                    ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
                    c = ws1.cell(row=row, column=1, value=vd_txt)
                    c.font = vardiya_font
                    c.fill = vardiya_fill
                    c.alignment = center
                    row += 1

                # Referans değişim ayırıcı
                if b['adi'] != prev_ref:
                    prev_ref = b['adi']
                    a = next((x for x in aski_analiz if x['idx'] == b['uidx']), None)
                    ref_txt = f"▸ {b['adi']}"
                    if a:
                        ref_txt += f"  (İhtiyaç: {a['ihtiyac']:,}  |  Çevrim Üretim: {a['cevrim_uretim']:,}  |  {a['gerekli_cevrim']} çevrim)"
                    ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
                    c = ws1.cell(row=row, column=1, value=ref_txt)
                    c.font = ref_font
                    c.fill = ref_fill
                    row += 1

                # Bara detay satırı
                a = next((x for x in aski_analiz if x['idx'] == b['uidx']), None)
                adet_bara = a['bara_adeti'] if a else 0

                ws1.cell(row=row, column=1, value=bi + 1).alignment = center
                ws1.cell(row=row, column=2, value=b['adi'])
                ws1.cell(row=row, column=3, value=b['tip']).alignment = center
                ws1.cell(row=row, column=4, value=b['bara_no']).alignment = center
                ws1.cell(row=row, column=5, value=f"Ç{b['cevrim'] + 1}").alignment = center
                ws1.cell(row=row, column=6, value=f"V{v_no}").alignment = center
                ws1.cell(row=row, column=7, value=b['giris']).alignment = center
                ws1.cell(row=row, column=8, value=b['cikis']).alignment = center
                ws1.cell(row=row, column=9, value=adet_bara).alignment = center
                ws1.cell(row=row, column=10, value=a['bara_aski'] if a else 0).alignment = center

                for col in range(1, 11):
                    ws1.cell(row=row, column=col).border = thin_border
                row += 1

            for col, w in [(1, 6), (2, 22), (3, 10), (4, 10), (5, 10),
                           (6, 10), (7, 10), (8, 10), (9, 12), (10, 8)]:
                ws1.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

            # ══════════════════════════════════════════════════
            # Sayfa 2: Çevrim Özet (çevrim bazlı)
            # ══════════════════════════════════════════════════
            ws2 = wb.create_sheet("Çevrim Özet")

            headers2 = ["Çevrim", "Vardiya", "İlk Giriş", "Son Çıkış",
                         "Bara", "Ürünler", "Doluluk %"]
            for ci, h in enumerate(headers2, 1):
                c = ws2.cell(row=1, column=ci, value=h)
                c.font = header_font
                c.fill = header_fill
                c.alignment = center
                c.border = thin_border

            for ci, slot in enumerate(cevrimler):
                r = ci + 2
                z = zaman[ci] if ci < len(zaman) else {}
                toplam_aski = sum(x[1] for x in slot)
                doluluk = toplam_aski / max(paralel, 1) * 100
                urun_str = ", ".join(f"{adi}({aski})" for adi, aski, tip, uidx in slot)

                ws2.cell(row=r, column=1, value=f"Ç{ci+1}").alignment = center
                ws2.cell(row=r, column=2, value=f"V{z.get('vardiya', '?')}").alignment = center
                ws2.cell(row=r, column=3, value=z.get('ilk_giris', '')).alignment = center
                ws2.cell(row=r, column=4, value=z.get('son_cikis', '')).alignment = center
                ws2.cell(row=r, column=5, value=toplam_aski).alignment = center
                ws2.cell(row=r, column=6, value=urun_str)
                dc = ws2.cell(row=r, column=7, value=f"%{doluluk:.0f}")
                dc.alignment = center
                if doluluk >= 80:
                    dc.font = Font(color="10B981", bold=True)
                elif doluluk >= 40:
                    dc.font = Font(color="F59E0B", bold=True)
                else:
                    dc.font = Font(color="EF4444", bold=True)
                for col in range(1, 8):
                    ws2.cell(row=r, column=col).border = thin_border

            for col, w in [(1, 10), (2, 10), (3, 12), (4, 12), (5, 8), (6, 50), (7, 12)]:
                ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

            # ══════════════════════════════════════════════════
            # Sayfa 3: Ürün Analizi
            # ══════════════════════════════════════════════════
            ws3 = wb.create_sheet("Ürün Analizi")

            headers3 = ["Ürün", "Tip", "İhtiyaç", "Bara Adeti", "Askı",
                         "Çevrim Üretim", "Gerekli Çevrim", "Toplam Bara", "Verimlilik"]
            for ci, h in enumerate(headers3, 1):
                c = ws3.cell(row=1, column=ci, value=h)
                c.font = header_font
                c.fill = header_fill
                c.alignment = center
                c.border = thin_border

            for ri, a in enumerate(aski_analiz, 2):
                ws3.cell(row=ri, column=1, value=a['adi'])
                ws3.cell(row=ri, column=2, value=a['tip'])
                ws3.cell(row=ri, column=3, value=a['ihtiyac'])
                ws3.cell(row=ri, column=4, value=a['bara_adeti'])
                ws3.cell(row=ri, column=5, value=a['bara_aski'])
                ws3.cell(row=ri, column=6, value=a['cevrim_uretim'])
                ws3.cell(row=ri, column=7, value=a['gerekli_cevrim'])
                ws3.cell(row=ri, column=8, value=a['toplam_slot'])
                ws3.cell(row=ri, column=9, value=round(a['verimlilik'], 3))
                for col in range(1, 10):
                    ws3.cell(row=ri, column=col).border = thin_border

            for col, w in [(1, 22), (2, 10), (3, 12), (4, 12), (5, 8), (6, 14), (7, 14), (8, 12), (9, 12)]:
                ws3.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

            wb.save(file_path)
            toplam_bara = ozet.get('toplam_bara', 0)
            QMessageBox.information(
                self, "Başarılı",
                f"Plan Excel'e aktarıldı:\n{file_path}\n\n"
                f"Sayfa 1: Bara Detay ({toplam_bara} bara, referans referans)\n"
                f"Sayfa 2: Çevrim Özet ({ozet['toplam_cevrim']} çevrim)\n"
                f"Sayfa 3: Ürün Analizi ({len(aski_analiz)} ürün)"
            )

        except ImportError:
            QMessageBox.warning(self, "Hata", "openpyxl kütüphanesi bulunamadı.\npip install openpyxl")
        except Exception as e:
            QMessageBox.warning(self, "Hata", f"Excel aktarım hatası:\n{e}")

    # ──────────────────────────────────────────────────────────
    #  İŞ EMRİ OLUŞTUR - Çevrim planından DB'ye iş emirleri yaz
    # ──────────────────────────────────────────────────────────
    def _on_is_emri_olustur(self):
        """Çevrim planından siparis.is_emirleri + uretim.planlama kayıtları oluştur"""
        if not hasattr(self, 'cevrim_plan_widget') or not self.cevrim_plan_widget._plan:
            QMessageBox.warning(self, "Uyarı", "Önce Hesapla ile çevrim planı oluşturun.")
            return

        plan = self.cevrim_plan_widget._plan
        cevrimler = plan['cevrimler']
        baralar = plan.get('baralar', [])
        zaman = plan.get('zaman', [])
        aski_analiz = plan['aski_analiz']
        ozet = plan['ozet']

        toplam_bara = ozet.get('toplam_bara', len(baralar))

        # Her ürün×çevrim için 1 iş emri (aynı ürünün çevrimdeki baraları gruplanır)
        urun_cevrim_gruplari = {}
        for b in baralar:
            key = (b['cevrim'], b['adi'], b['uidx'])
            if key not in urun_cevrim_gruplari:
                urun_cevrim_gruplari[key] = []
            urun_cevrim_gruplari[key].append(b)

        toplam_ie = len(urun_cevrim_gruplari)

        ret = QMessageBox.question(
            self, "İş Emri Onayı",
            f"Toplam {toplam_ie} iş emri ({toplam_bara} bara) oluşturulacak.\n"
            f"{ozet['toplam_cevrim']} çevrim, {ozet['vardiya_gerekli']} vardiya\n\n"
            f"Devam edilsin mi?",
            QMessageBox.Yes | QMessageBox.No
        )
        if ret != QMessageBox.Yes:
            return

        try:
            from core.database import get_db_connection
            from datetime import datetime

            conn = get_db_connection()
            cursor = conn.cursor()

            cursor.execute("SELECT ISNULL(MAX(id), 0) FROM siparis.is_emirleri")
            max_id = cursor.fetchone()[0]

            tarih_str = datetime.now().strftime('%Y%m')
            ie_no_base = max_id
            olusturulan = 0
            hatalar = []

            for (ci, adi, uidx), grup_baralar in sorted(urun_cevrim_gruplari.items()):
                a = next((x for x in aski_analiz if x['idx'] == uidx), None)
                if not a:
                    continue

                v_no = grup_baralar[0]['vardiya']
                ilk_giris = grup_baralar[0]['giris']
                son_cikis = grup_baralar[-1]['cikis']
                aski = a['bara_aski']
                cevrim_uretim = a.get('cevrim_uretim', 0)
                sira = ci + 1

                ie_no_base += 1
                is_emri_no = f"KP-{tarih_str}-{ie_no_base:04d}"

                try:
                    cursor.execute("""
                        INSERT INTO siparis.is_emirleri
                        (uuid, is_emri_no, tarih, stok_adi, kaplama_tipi,
                         planlanan_miktar, toplam_miktar, birim, toplam_bara,
                         aski_adet, tahmini_sure_dk,
                         termin_tarihi, durum, uretim_notu,
                         olusturma_tarihi, guncelleme_tarihi, silindi_mi)
                        OUTPUT INSERTED.id
                        VALUES (NEWID(), ?, GETDATE(), ?, ?,
                                ?, ?, 'ADET', ?,
                                ?, ?,
                                GETDATE(), 'PLANLANDI', ?,
                                GETDATE(), GETDATE(), 0)
                    """, (
                        is_emri_no, adi, a['tip'],
                        cevrim_uretim, cevrim_uretim, aski,
                        aski, self.spn_kap_cevrim.value(),
                        f"V{v_no} Ç{ci+1} | {ilk_giris}-{son_cikis} | {aski} bara | Setup:{ozet.get('setup_dk', 10)}dk"
                    ))
                    is_emri_id = cursor.fetchone()[0]

                    from core.yetki_manager import YetkiManager
                    _olusturan_id = YetkiManager._current_user_id
                    cursor.execute("""
                        INSERT INTO uretim.planlama
                        (uuid, tarih, vardiya_id, is_emri_id, sira_no,
                         planlanan_bara, durum, olusturan_id, olusturma_tarihi)
                        VALUES (NEWID(), GETDATE(), ?, ?, ?,
                                ?, 'PLANLANDI', ?, GETDATE())
                    """, (v_no, is_emri_id, sira, aski, _olusturan_id))

                    olusturulan += 1
                except Exception as ie_err:
                    hatalar.append(f"{is_emri_no}: {ie_err}")

            conn.commit()
            conn.close()

            # Bildirim: Kaplama planlamasından iş emirleri oluşturuldu
            if olusturulan > 0:
                try:
                    from core.bildirim_tetikleyici import BildirimTetikleyici
                    BildirimTetikleyici.is_emri_olusturuldu(
                        ie_id=0,
                        ie_no=f"{olusturulan} IE (Kaplama Planlama)",
                        musteri_adi='',
                    )
                except Exception as bt_err:
                    print(f"Bildirim hatasi: {bt_err}")

            msg = f"{olusturulan} iş emri oluşturuldu."
            if hatalar:
                msg += f"\n\n{len(hatalar)} hata:\n" + "\n".join(hatalar[:5])
            QMessageBox.information(self, "İş Emri", msg)

        except Exception as e:
            QMessageBox.warning(self, "Hata", f"İş emri oluşturma hatası:\n{e}")


# ══════════════════════════════════════════════════════════════
#  URUN EKLE DIALOG
# ══════════════════════════════════════════════════════════════

class UrunEkleDialog(QDialog):
    """Urun ekleme dialog'u - Stok kartindan arama ile"""

    def __init__(self, style: dict, parent=None):
        super().__init__(parent)
        self.s = style
        self.selected_urun = None
        self._sonuclar = []
        self._search_timer = QTimer()
        self._search_timer.setSingleShot(True)
        self._search_timer.setInterval(400)
        self._search_timer.timeout.connect(self._do_search)
        self.setWindowTitle("Stok Kartindan Urun Ekle")
        self.setMinimumSize(650, 580)
        self.setModal(True)
        self.setStyleSheet(f"QDialog {{ background: {brand.BG_CARD}; color: {brand.TEXT}; }}")
        self._setup_ui()

    def _setup_ui(self):
        s = self.s
        layout = QVBoxLayout(self)
        layout.setSpacing(SPACING)
        layout.setContentsMargins(MARGIN + 4, MARGIN, MARGIN + 4, MARGIN)

        inp_style = _input_style(s)
        lbl_style = f"color: {s['text_muted']}; font-size: {LABEL_SIZE}px; font-weight: 600; letter-spacing: 1px;"

        # Baslik
        header = QHBoxLayout()
        accent = QFrame()
        accent.setFixedSize(4, 28)
        accent.setStyleSheet(f"background: {s['primary']}; border-radius: 2px;")
        header.addWidget(accent)

        lbl_ara = QLabel("Urun Ara")
        lbl_ara.setStyleSheet(f"color: {s['text']}; font-size: 18px; font-weight: 600; margin-left: 10px;")
        header.addWidget(lbl_ara)
        header.addStretch()
        layout.addLayout(header)

        self.txt_arama = QLineEdit()
        self.txt_arama.setPlaceholderText("Urun kodu veya adi yazin...")
        self.txt_arama.setStyleSheet(inp_style)
        self.txt_arama.textChanged.connect(lambda: self._search_timer.start())
        layout.addWidget(self.txt_arama)

        # Sonuc tablosu
        self.sonuc_table = QTableWidget()
        self.sonuc_table.setColumnCount(5)
        self.sonuc_table.setHorizontalHeaderLabels(["Urun Kodu", "Urun Adi", "Kaplama", "Aski", "Bara"])
        self.sonuc_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.sonuc_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.sonuc_table.verticalHeader().setVisible(False)
        self.sonuc_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.sonuc_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.sonuc_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.sonuc_table.setMaximumHeight(160)
        self.sonuc_table.setAlternatingRowColors(False)
        self.sonuc_table.setStyleSheet(_table_style(s))
        self.sonuc_table.cellClicked.connect(self._on_sonuc_secildi)
        layout.addWidget(self.sonuc_table)

        # Secilen urun
        self.lbl_secilen = QLabel("Henuz urun secilmedi")
        self.lbl_secilen.setStyleSheet(f"color: {s['text_muted']}; font-size: {SMALL_SIZE}px; padding: 2px;")
        layout.addWidget(self.lbl_secilen)

        # Recete adimlari
        self.recete_preview = ReceteAdimWidget()
        layout.addWidget(self.recete_preview)

        # Stok karti bilgileri
        info_frame = QFrame()
        info_frame.setStyleSheet(_card_style(s))
        info_layout = QGridLayout(info_frame)
        info_layout.setContentsMargins(14, 10, 14, 10)
        info_layout.setSpacing(8)

        def add_info(row, col, label_text, obj_name):
            lbl = QLabel(label_text.upper())
            lbl.setStyleSheet(lbl_style)
            val = QLabel("-")
            val.setObjectName(obj_name)
            val.setStyleSheet(f"color: {s['text']}; font-size: {SMALL_SIZE}px; font-weight: 600; border: none;")
            info_layout.addWidget(lbl, row, col * 2)
            info_layout.addWidget(val, row, col * 2 + 1)

        add_info(0, 0, "Recete:", "info_recete")
        add_info(0, 1, "Kaplama:", "info_kaplama")
        add_info(1, 0, "Aski Tipi:", "info_aski_tip")
        add_info(1, 1, "Aski Adedi:", "info_aski_adedi")

        layout.addWidget(info_frame)

        # Recete secimi
        rec_label = QLabel("RECETE SEC")
        rec_label.setStyleSheet(lbl_style)
        layout.addWidget(rec_label)

        self.cmb_recete = QComboBox()
        self.cmb_recete.setStyleSheet(inp_style)
        self.cmb_recete.currentIndexChanged.connect(self._on_recete_secildi)
        layout.addWidget(self.cmb_recete)

        self.lbl_cevrim_info = QLabel("")
        self.lbl_cevrim_info.setStyleSheet(f"color: {s['text_muted']}; font-size: {LABEL_SIZE}px;")
        layout.addWidget(self.lbl_cevrim_info)

        # Kullanici alanlari
        form_layout = QFormLayout()
        form_layout.setSpacing(8)

        self.spn_cevrim = QSpinBox()
        self.spn_cevrim.setRange(1, 480)
        self.spn_cevrim.setValue(45)
        self.spn_cevrim.setSuffix(" dk")
        self.spn_cevrim.setStyleSheet(inp_style)

        self.spn_stok_aski = QSpinBox()
        self.spn_stok_aski.setRange(0, 9999)
        self.spn_stok_aski.setValue(0)
        self.spn_stok_aski.setStyleSheet(inp_style)

        self.spn_bara_aski = QSpinBox()
        self.spn_bara_aski.setRange(1, 99)
        self.spn_bara_aski.setValue(2)
        self.spn_bara_aski.setStyleSheet(inp_style)

        self.spn_kapasite = QSpinBox()
        self.spn_kapasite.setRange(1, 9999)
        self.spn_kapasite.setValue(1)
        self.spn_kapasite.setToolTip("Bir askiya kac adet urun asilir")
        self.spn_kapasite.setStyleSheet(inp_style)

        self.spn_ihtiyac = QSpinBox()
        self.spn_ihtiyac.setRange(0, 999999)
        self.spn_ihtiyac.setValue(100)
        self.spn_ihtiyac.setStyleSheet(inp_style)

        self.cmb_oncelik = QComboBox()
        self.cmb_oncelik.addItems(["Normal", "Acil"])
        self.cmb_oncelik.setStyleSheet(inp_style)

        for label, widget in [
            ("Cevrim Suresi:", self.spn_cevrim),
            ("Aski Stoku (adet):", self.spn_stok_aski),
            ("Bara Basina Aski:", self.spn_bara_aski),
            ("Aski Kapasitesi:", self.spn_kapasite),
            ("Haftalik Ihtiyac:", self.spn_ihtiyac),
            ("Oncelik:", self.cmb_oncelik),
        ]:
            lbl = QLabel(label)
            lbl.setStyleSheet(f"color: {s['text_secondary']}; font-size: {SMALL_SIZE}px;")
            form_layout.addRow(lbl, widget)

        layout.addLayout(form_layout)

        # Alt butonlar
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(CARD_SPACING)
        btn_layout.addStretch()

        btn_iptal = QPushButton("Iptal")
        btn_iptal.setCursor(Qt.PointingHandCursor)
        btn_iptal.setStyleSheet(_secondary_btn_style(s))
        btn_iptal.clicked.connect(self.reject)
        btn_layout.addWidget(btn_iptal)

        btn_ekle = QPushButton("Ekle")
        btn_ekle.setCursor(Qt.PointingHandCursor)
        btn_ekle.setStyleSheet(_primary_btn_style(s))
        btn_ekle.clicked.connect(self._on_ekle)
        btn_layout.addWidget(btn_ekle)

        layout.addLayout(btn_layout)

    def _do_search(self):
        arama = self.txt_arama.text().strip()
        if len(arama) < 2:
            self.sonuc_table.setRowCount(0)
            return
        sonuclar = db.search_stok_kartlari(arama)
        self._sonuclar = sonuclar
        self.sonuc_table.setRowCount(len(sonuclar))
        for i, s in enumerate(sonuclar):
            self.sonuc_table.setItem(i, 0, QTableWidgetItem(s['urun_kodu']))
            self.sonuc_table.setItem(i, 1, QTableWidgetItem(s['urun_adi'] or ''))
            self.sonuc_table.setItem(i, 2, QTableWidgetItem(s['kaplama_adi'] or '-'))
            self.sonuc_table.setItem(i, 3, QTableWidgetItem(str(s['aski_adedi'])))
            self.sonuc_table.setItem(i, 4, QTableWidgetItem(str(s['bara_adedi'])))

    def _on_sonuc_secildi(self, row, col):
        if row < 0 or row >= len(self._sonuclar):
            return
        s = self._sonuclar[row]
        self.selected_urun = s

        self.lbl_secilen.setText(f"Secilen: {s['urun_kodu']} - {s['urun_adi'] or ''}")
        self.lbl_secilen.setStyleSheet(f"color: {brand.SUCCESS}; font-size: {SMALL_SIZE}px; font-weight: 600; padding: 2px;")

        self.findChild(QLabel, "info_recete").setText(s['recete_no'] or '-')
        self.findChild(QLabel, "info_kaplama").setText(s['kaplama_adi'] or '-')
        self.findChild(QLabel, "info_aski_tip").setText(s['aski_tip_adi'] or '-')
        self.findChild(QLabel, "info_aski_adedi").setText(str(s['aski_adedi']) if s['aski_adedi'] else '-')

        if s['aski_adedi'] and s['aski_adedi'] > 0:
            self.spn_stok_aski.setValue(s['aski_adedi'])
        if s['bara_adedi'] and s['bara_adedi'] > 0:
            self.spn_bara_aski.setValue(s['bara_adedi'])

        # Recete secimi icin combobox doldur
        self.cmb_recete.clear()
        recete_no_str = s.get('recete_no', '')
        if recete_no_str:
            try:
                rec_no = int(recete_no_str)
                tanim = db.get_recete_by_no(rec_no)
                if tanim:
                    self.cmb_recete.addItem(
                        f"R{rec_no} - {tanim.get('recete_adi', '')} / {tanim.get('recete_aciklama', '')} ({tanim.get('toplam_sure_dk', '?')}dk)",
                        rec_no
                    )
            except (ValueError, TypeError):
                pass

        kaplama = (s.get('kaplama_kodu') or '').lower()
        if 'ni' in kaplama or 'nikel' in kaplama:
            hat_tipi = 'ZNNI'
        elif 'ktl' in kaplama or 'kat' in kaplama:
            hat_tipi = 'KTL'
        else:
            hat_tipi = None

        if hat_tipi:
            tum_receteler = db.get_recete_tanimlari(hat_tipi)
            existing_nos = [self.cmb_recete.itemData(i) for i in range(self.cmb_recete.count())]
            for rt in tum_receteler:
                if rt['recete_no'] not in existing_nos:
                    sure_str = f"{rt['toplam_sure_dk']}dk" if rt['toplam_sure_dk'] else "?"
                    self.cmb_recete.addItem(
                        f"R{rt['recete_no']} - {rt['recete_adi']} / {rt['recete_aciklama']} ({sure_str})",
                        rt['recete_no']
                    )

        self._on_recete_secildi()

        recete = db.get_urun_recete(s['urun_kodu'])
        if recete:
            self.recete_preview.set_adimlar(recete)
        else:
            self.recete_preview.set_adimlar([])

    def _on_recete_secildi(self):
        rec_no = self.cmb_recete.currentData()
        if rec_no is None:
            return
        tanim = db.get_recete_by_no(rec_no)
        if tanim and tanim.get('toplam_sure_dk'):
            self.spn_cevrim.setValue(tanim['toplam_sure_dk'])
            self.lbl_cevrim_info.setText(
                f"PLC'den hesaplanan: {tanim['toplam_sure_dk']} dk (adim toplami)"
            )
            self.lbl_cevrim_info.setStyleSheet(f"color: {brand.SUCCESS}; font-size: {LABEL_SIZE}px;")
        else:
            self.lbl_cevrim_info.setText("PLC verisi bulunamadi - manuel girin")
            self.lbl_cevrim_info.setStyleSheet(f"color: {brand.WARNING}; font-size: {LABEL_SIZE}px;")

    def _on_ekle(self):
        if not self.selected_urun:
            QMessageBox.warning(self, "Uyari", "Listeden bir urun secin.")
            return
        self.accept()

    def get_urun(self) -> KaplamaUrun:
        s = self.selected_urun
        kaplama = (s.get('kaplama_kodu') or '').lower()
        tip = "zn-ni" if ('ni' in kaplama or 'nikel' in kaplama) else "zn"
        oncelik = "acil" if self.cmb_oncelik.currentIndex() == 1 else "normal"

        rec_no = self.cmb_recete.currentData()
        recete_str = str(rec_no) if rec_no else s.get('recete_no', '')

        return KaplamaUrun(
            ref=s['urun_kodu'],
            recete_no=recete_str,
            tip=tip,
            aski_tip=s.get('aski_tip_kodu', ''),
            kapasite=self.spn_kapasite.value(),
            cevrim_suresi=self.spn_cevrim.value(),
            stok_aski=self.spn_stok_aski.value(),
            bara_aski=self.spn_bara_aski.value(),
            haftalik_ihtiyac=self.spn_ihtiyac.value(),
            oncelik=oncelik,
        )
