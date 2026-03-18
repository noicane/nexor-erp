# -*- coding: utf-8 -*-
"""
REDLINE NEXOR ERP - Cari Detay Sayfası
Seçilen carinin tüm detayları ve ilişkili bilgileri
"""
from PySide6.QtWidgets import (
    QVBoxLayout, QHBoxLayout, QLabel, QFrame, QTableWidget, QTableWidgetItem,
    QLineEdit, QComboBox, QPushButton, QHeaderView, QMessageBox, QAbstractItemView,
    QTabWidget, QWidget, QFormLayout, QTextEdit, QGridLayout, QGroupBox
)
from PySide6.QtCore import Qt, QTimer, Signal
from PySide6.QtGui import QColor, QBrush
from datetime import datetime

from components.base_page import BasePage
from core.database import get_db_connection
from core.log_manager import LogManager


class CariDetayPage(BasePage):
    """Cari Detay Sayfası"""
    
    back_requested = Signal()
    
    def __init__(self, theme: dict):
        super().__init__(theme)
        self.current_cari_id = None
        self.cari_data = {}
        self._setup_ui()
    
    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(16)
        
        # Üst bar
        top_bar = self._create_top_bar()
        layout.addWidget(top_bar)
        
        # Bilgi kartı
        self.info_card = self._create_info_card()
        layout.addWidget(self.info_card)
        
        # Sekmeler
        self.tabs = self._create_tabs()
        layout.addWidget(self.tabs, 1)
        
        # Alt butonlar
        bottom = self._create_bottom_buttons()
        layout.addLayout(bottom)
    
    def _create_top_bar(self) -> QFrame:
        frame = QFrame()
        frame.setStyleSheet(f"QFrame {{ background: {self.theme['bg_card']}; border-radius: 12px; }}")
        layout = QHBoxLayout(frame)
        layout.setContentsMargins(16, 12, 16, 12)
        
        # Geri butonu
        back_btn = QPushButton("← Listeye Dön")
        back_btn.setCursor(Qt.PointingHandCursor)
        back_btn.setStyleSheet(f"""
            QPushButton {{ background: transparent; color: {self.theme['primary']}; border: none; font-size: 13px; font-weight: 500; padding: 8px 12px; }}
            QPushButton:hover {{ background: {self.theme['primary']}15; border-radius: 6px; }}
        """)
        back_btn.clicked.connect(self.back_requested.emit)
        layout.addWidget(back_btn)
        
        layout.addSpacing(20)
        
        # Cari seçici
        lbl = QLabel("Cari Seç:")
        lbl.setStyleSheet(f"color: {self.theme['text_secondary']}; font-size: 12px;")
        layout.addWidget(lbl)
        
        self.cari_combo = QComboBox()
        self.cari_combo.setFixedWidth(400)
        self.cari_combo.setEditable(True)
        self.cari_combo.setStyleSheet(f"""
            QComboBox {{ background: {self.theme['bg_main']}; color: {self.theme['text']}; border: 1px solid {self.theme['border']}; border-radius: 6px; padding: 8px 12px; font-size: 13px; }}
            QComboBox QAbstractItemView {{ background: {self.theme['bg_card']}; color: {self.theme['text']}; selection-background-color: {self.theme['primary']}; }}
        """)
        self.cari_combo.currentIndexChanged.connect(self._on_cari_selected)
        layout.addWidget(self.cari_combo)
        
        layout.addStretch()
        
        self.title_label = QLabel("Cari Detay")
        self.title_label.setStyleSheet(f"color: {self.theme['text']}; font-size: 18px; font-weight: bold;")
        layout.addWidget(self.title_label)
        
        return frame
    
    def _create_info_card(self) -> QFrame:
        frame = QFrame()
        frame.setStyleSheet(f"QFrame {{ background: {self.theme['bg_card']}; border-radius: 12px; }}")
        layout = QHBoxLayout(frame)
        layout.setContentsMargins(20, 16, 20, 16)
        layout.setSpacing(24)
        
        # Sol: Ana bilgiler
        left = QVBoxLayout()
        left.setSpacing(8)
        
        self.unvan_label = QLabel("-")
        self.unvan_label.setStyleSheet(f"color: {self.theme['text']}; font-size: 20px; font-weight: bold;")
        self.unvan_label.setWordWrap(True)
        left.addWidget(self.unvan_label)
        
        info_row = QHBoxLayout()
        self.kod_label = QLabel("-")
        self.kod_label.setStyleSheet(f"color: {self.theme['primary']}; background: {self.theme['primary']}20; padding: 4px 12px; border-radius: 4px; font-size: 12px; font-weight: 600;")
        info_row.addWidget(self.kod_label)
        
        self.tip_label = QLabel("-")
        self.tip_label.setStyleSheet(f"color: {self.theme['text_secondary']}; background: {self.theme['border']}; padding: 4px 12px; border-radius: 4px; font-size: 12px;")
        info_row.addWidget(self.tip_label)
        
        self.durum_label = QLabel("-")
        info_row.addWidget(self.durum_label)
        info_row.addStretch()
        left.addLayout(info_row)
        layout.addLayout(left, 2)
        
        # Orta: İletişim
        mid = QGridLayout()
        mid.setSpacing(8)
        items = [("📞", "Telefon:", "telefon"), ("📱", "Cep:", "cep"), ("📧", "E-posta:", "email"), ("🌐", "Web:", "web")]
        self.contact_labels = {}
        for i, (icon, label, key) in enumerate(items):
            mid.addWidget(QLabel(icon), i, 0)
            lbl = QLabel(label)
            lbl.setStyleSheet(f"color: {self.theme['text_secondary']}; font-size: 11px;")
            mid.addWidget(lbl, i, 1)
            val = QLabel("-")
            val.setStyleSheet(f"color: {self.theme['text']}; font-size: 12px;")
            self.contact_labels[key] = val
            mid.addWidget(val, i, 2)
        layout.addLayout(mid, 1)
        
        # Sağ: Vergi
        right = QGridLayout()
        right.setSpacing(8)
        tax_items = [("Vergi Dairesi:", "vd"), ("Vergi No:", "vn"), ("Şehir:", "sehir"), ("Vade:", "vade")]
        self.tax_labels = {}
        for i, (label, key) in enumerate(tax_items):
            lbl = QLabel(label)
            lbl.setStyleSheet(f"color: {self.theme['text_secondary']}; font-size: 11px;")
            right.addWidget(lbl, i, 0)
            val = QLabel("-")
            val.setStyleSheet(f"color: {self.theme['text']}; font-size: 12px; font-weight: 500;")
            self.tax_labels[key] = val
            right.addWidget(val, i, 1)
        layout.addLayout(right, 1)
        
        return frame
    
    def _create_tabs(self) -> QTabWidget:
        tabs = QTabWidget()
        tabs.setStyleSheet(f"""
            QTabWidget::pane {{ background: {self.theme['bg_card']}; border: none; border-radius: 10px; }}
            QTabBar::tab {{ background: {self.theme['bg_main']}; color: {self.theme['text_secondary']}; padding: 10px 20px; margin-right: 4px; border-top-left-radius: 8px; border-top-right-radius: 8px; font-size: 12px; }}
            QTabBar::tab:selected {{ background: {self.theme['bg_card']}; color: {self.theme['primary']}; font-weight: 600; }}
        """)
        
        tabs.addTab(self._create_general_tab(), "📋 Genel")
        tabs.addTab(self._create_address_tab(), "📍 Adresler")
        tabs.addTab(self._create_contact_tab(), "👤 Yetkililer")
        tabs.addTab(self._create_spec_tab(), "📐 Spesifikasyonlar")
        
        return tabs
    
    def _create_general_tab(self) -> QWidget:
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Adresler
        grp = QGroupBox("Adres Bilgileri")
        grp.setStyleSheet(f"QGroupBox {{ color: {self.theme['text']}; font-weight: bold; border: 1px solid {self.theme['border']}; border-radius: 8px; margin-top: 12px; }} QGroupBox::title {{ subcontrol-origin: margin; left: 16px; padding: 0 8px; }}")
        g_layout = QGridLayout(grp)
        g_layout.setContentsMargins(16, 24, 16, 16)
        
        g_layout.addWidget(QLabel("Fatura Adresi:"), 0, 0, Qt.AlignTop)
        self.fatura_text = QTextEdit()
        self.fatura_text.setReadOnly(True)
        self.fatura_text.setMaximumHeight(70)
        self.fatura_text.setStyleSheet(f"QTextEdit {{ background: {self.theme['bg_main']}; color: {self.theme['text']}; border: 1px solid {self.theme['border']}; border-radius: 6px; padding: 8px; }}")
        g_layout.addWidget(self.fatura_text, 0, 1)
        
        g_layout.addWidget(QLabel("Sevk Adresi:"), 1, 0, Qt.AlignTop)
        self.sevk_text = QTextEdit()
        self.sevk_text.setReadOnly(True)
        self.sevk_text.setMaximumHeight(70)
        self.sevk_text.setStyleSheet(self.fatura_text.styleSheet())
        g_layout.addWidget(self.sevk_text, 1, 1)
        
        layout.addWidget(grp)
        layout.addStretch()
        return widget
    
    def _create_address_tab(self) -> QWidget:
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(16, 16, 16, 16)
        
        self.addr_table = QTableWidget()
        self.addr_table.setColumnCount(5)
        self.addr_table.setHorizontalHeaderLabels(["Tip", "Adres Adı", "Şehir", "Yetkili", "Telefon"])
        self._style_table(self.addr_table)
        layout.addWidget(self.addr_table)
        
        return widget
    
    def _create_contact_tab(self) -> QWidget:
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(16, 16, 16, 16)
        
        self.yetkili_table = QTableWidget()
        self.yetkili_table.setColumnCount(5)
        self.yetkili_table.setHorizontalHeaderLabels(["Ad Soyad", "Ünvan", "Telefon", "E-posta", "Birincil"])
        self._style_table(self.yetkili_table)
        layout.addWidget(self.yetkili_table)
        
        return widget
    
    def _create_spec_tab(self) -> QWidget:
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(16, 16, 16, 16)
        
        self.spec_table = QTableWidget()
        self.spec_table.setColumnCount(5)
        self.spec_table.setHorizontalHeaderLabels(["Kaplama Türü", "Min (µm)", "Max (µm)", "Hedef (µm)", "Tuz Testi (saat)"])
        self._style_table(self.spec_table)
        layout.addWidget(self.spec_table)
        
        note = QLabel("💡 Müşteriye özel kaplama kalınlığı ve kalite gereksinimleri")
        note.setStyleSheet(f"color: {self.theme['text_secondary']}; font-size: 11px; font-style: italic;")
        layout.addWidget(note)
        
        return widget
    
    def _style_table(self, table: QTableWidget):
        table.setStyleSheet(f"""
            QTableWidget {{ background: {self.theme['bg_card']}; color: {self.theme['text']}; border: none; gridline-color: {self.theme['border']}; font-size: 12px; }}
            QTableWidget::item {{ padding: 8px; border-bottom: 1px solid {self.theme['border']}; }}
            QTableWidget::item:selected {{ background: {self.theme['primary']}30; }}
            QHeaderView::section {{ background: {self.theme['bg_main']}; color: {self.theme['text']}; padding: 10px 8px; border: none; border-bottom: 2px solid {self.theme['border']}; font-weight: bold; font-size: 11px; }}
        """)
        table.setSelectionBehavior(QAbstractItemView.SelectRows)
        table.verticalHeader().setVisible(False)
        table.horizontalHeader().setStretchLastSection(True)
        table.verticalHeader().setDefaultSectionSize(40)
    
    def _create_bottom_buttons(self) -> QHBoxLayout:
        layout = QHBoxLayout()
        
        self.update_label = QLabel("Son güncelleme: -")
        self.update_label.setStyleSheet(f"color: {self.theme['text_secondary']}; font-size: 11px;")
        layout.addWidget(self.update_label)
        layout.addStretch()
        
        for text, icon, callback in [("Yenile", "🔄", self._refresh), ("Excel", "📊", self._export)]:
            btn = QPushButton(f"{icon}  {text}")
            btn.setFixedHeight(36)
            btn.setCursor(Qt.PointingHandCursor)
            btn.setStyleSheet(f"""
                QPushButton {{ background: {self.theme['bg_card']}; color: {self.theme['text']}; border: 1px solid {self.theme['border']}; border-radius: 8px; padding: 0 16px; font-size: 12px; }}
                QPushButton:hover {{ background: {self.theme['bg_main']}; border-color: {self.theme['primary']}; }}
            """)
            btn.clicked.connect(callback)
            layout.addWidget(btn)
        
        return layout
    
    def load_cari_list(self):
        """Cari combo'sunu doldur"""
        try:
            conn = get_db_connection()
            if not conn:
                return
            cursor = conn.cursor()
            cursor.execute("SELECT id, hesap_kodu, unvani FROM dbo.CariKartlari WHERE aktif = 1 ORDER BY hesap_kodu")
            rows = cursor.fetchall()
            conn.close()
            
            self.cari_combo.blockSignals(True)
            self.cari_combo.clear()
            self.cari_combo.addItem("-- Cari Seçin --", None)
            for row in rows:
                txt = f"{row[1]} - {row[2][:40]}" if len(str(row[2])) > 40 else f"{row[1]} - {row[2]}"
                self.cari_combo.addItem(txt, row[0])
            self.cari_combo.blockSignals(False)
        except Exception as e:
            print(f"Hata: {e}")
    
    def load_cari(self, cari_id: int, hesap_kodu: str = None):
        """Belirtilen cariyi yükle"""
        self.current_cari_id = cari_id
        for i in range(self.cari_combo.count()):
            if self.cari_combo.itemData(i) == cari_id:
                self.cari_combo.blockSignals(True)
                self.cari_combo.setCurrentIndex(i)
                self.cari_combo.blockSignals(False)
                break
        self._load_data()
    
    def _on_cari_selected(self, idx):
        cari_id = self.cari_combo.currentData()
        if cari_id:
            self.current_cari_id = cari_id
            self._load_data()
    
    def _load_data(self):
        """Tüm verileri yükle"""
        if not self.current_cari_id:
            return
        
        try:
            conn = get_db_connection()
            if not conn:
                return
            cursor = conn.cursor()
            
            # Ana cari bilgileri
            cursor.execute("""
                SELECT hesap_kodu, unvani, kisa_unvan, tip_adi, sehir, ilce, telefon, cep_tel, 
                       email, web, vergi_dairesi, vergi_no, vade_gun, odeme_sekli, 
                       fatura_adresi, sevk_adresi, aktif
                FROM dbo.CariKartlari WHERE id = ?
            """, (self.current_cari_id,))
            row = cursor.fetchone()
            
            if row:
                self.kod_label.setText(row[0] or "-")
                self.unvan_label.setText(row[1] or "-")
                self.tip_label.setText(row[3] or "-")
                self.contact_labels['telefon'].setText(row[6] or "-")
                self.contact_labels['cep'].setText(row[7] or "-")
                self.contact_labels['email'].setText(row[8] or "-")
                self.contact_labels['web'].setText(row[9] or "-")
                self.tax_labels['vd'].setText(row[10] or "-")
                self.tax_labels['vn'].setText(row[11] or "-")
                self.tax_labels['sehir'].setText(row[4] or "-")
                self.tax_labels['vade'].setText(f"{row[12]} gün" if row[12] else "-")
                self.fatura_text.setPlainText(row[14] or "")
                self.sevk_text.setPlainText(row[15] or "")
                
                aktif = row[16] if row[16] is not None else True
                self.durum_label.setText("✅ Aktif" if aktif else "❌ Pasif")
                self.durum_label.setStyleSheet(f"color: {'#10b981' if aktif else '#ef4444'}; font-size: 12px; font-weight: 500;")
                
                self.title_label.setText(f"Cari Detay: {row[0]}")
            
            # Adresler (musteri.cari_adresler)
            self._load_addresses(cursor)
            
            # Yetkililer (musteri.cari_yetkililer)
            self._load_contacts(cursor)
            
            # Spesifikasyonlar (musteri.cari_spesifikasyonlar)
            self._load_specs(cursor)
            
            conn.close()
            self.update_label.setText(f"Son güncelleme: {datetime.now().strftime('%H:%M:%S')}")
            
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Veri yüklenirken hata:\n{str(e)}")
    
    def _load_addresses(self, cursor):
        """Adresleri yükle"""
        try:
            # Önce musteri.cariler'den cari_id bul (zirve_cari_kodu ile eşleştir)
            cursor.execute("""
                SELECT ca.adres_tipi, ca.adres_adi, s.ad as sehir, ca.yetkili_kisi, ca.telefon
                FROM musteri.cari_adresler ca
                LEFT JOIN tanim.sehirler s ON ca.sehir_id = s.id
                WHERE ca.cari_id IN (
                    SELECT c.id FROM musteri.cariler c
                    JOIN dbo.CariKartlari ck ON c.zirve_cari_kodu = ck.hesap_kodu
                    WHERE ck.id = ?
                ) AND ca.aktif_mi = 1
            """, (self.current_cari_id,))
            rows = cursor.fetchall()
            
            self.addr_table.setRowCount(len(rows))
            for i, row in enumerate(rows):
                for j, val in enumerate(row):
                    self.addr_table.setItem(i, j, QTableWidgetItem(str(val) if val else "-"))
        except Exception:
            self.addr_table.setRowCount(0)
    
    def _load_contacts(self, cursor):
        """Yetkilileri yükle"""
        try:
            cursor.execute("""
                SELECT cy.ad_soyad, cy.unvan, cy.telefon, cy.email, 
                       CASE WHEN cy.birincil_yetkili_mi = 1 THEN '✓' ELSE '' END
                FROM musteri.cari_yetkililer cy
                WHERE cy.cari_id IN (
                    SELECT c.id FROM musteri.cariler c
                    JOIN dbo.CariKartlari ck ON c.zirve_cari_kodu = ck.hesap_kodu
                    WHERE ck.id = ?
                ) AND cy.aktif_mi = 1
            """, (self.current_cari_id,))
            rows = cursor.fetchall()
            
            self.yetkili_table.setRowCount(len(rows))
            for i, row in enumerate(rows):
                for j, val in enumerate(row):
                    item = QTableWidgetItem(str(val) if val else "-")
                    if j == 4:  # Birincil
                        item.setTextAlignment(Qt.AlignCenter)
                        if val == '✓':
                            item.setForeground(QBrush(QColor("#10b981")))
                    self.yetkili_table.setItem(i, j, item)
        except Exception:
            self.yetkili_table.setRowCount(0)
    
    def _load_specs(self, cursor):
        """Spesifikasyonları yükle"""
        try:
            cursor.execute("""
                SELECT kt.ad as kaplama_turu, cs.min_kalinlik_um, cs.max_kalinlik_um, 
                       cs.hedef_kalinlik_um, cs.tuz_testi_saat
                FROM musteri.cari_spesifikasyonlar cs
                LEFT JOIN tanim.kaplama_turleri kt ON cs.kaplama_turu_id = kt.id
                WHERE cs.cari_id IN (
                    SELECT c.id FROM musteri.cariler c
                    JOIN dbo.CariKartlari ck ON c.zirve_cari_kodu = ck.hesap_kodu
                    WHERE ck.id = ?
                ) AND cs.aktif_mi = 1
            """, (self.current_cari_id,))
            rows = cursor.fetchall()
            
            self.spec_table.setRowCount(len(rows))
            for i, row in enumerate(rows):
                for j, val in enumerate(row):
                    item = QTableWidgetItem(str(val) if val else "-")
                    if j > 0:  # Sayısal alanlar
                        item.setTextAlignment(Qt.AlignCenter)
                    self.spec_table.setItem(i, j, item)
        except Exception:
            self.spec_table.setRowCount(0)
    
    def _refresh(self):
        self._load_data()
    
    def _export(self):
        """Excel'e aktar"""
        try:
            from openpyxl import Workbook
            from PySide6.QtWidgets import QFileDialog
            
            if not self.current_cari_id:
                QMessageBox.warning(self, "Uyarı", "Önce bir cari seçin!")
                return
            
            path, _ = QFileDialog.getSaveFileName(
                self, "Excel Kaydet",
                f"cari_detay_{self.kod_label.text()}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                "Excel (*.xlsx)"
            )
            if not path:
                return
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Cari Bilgileri"
            
            # Başlık
            ws['A1'] = "Cari Detay Raporu"
            ws['A2'] = f"Hesap Kodu: {self.kod_label.text()}"
            ws['A3'] = f"Ünvan: {self.unvan_label.text()}"
            ws['A4'] = f"Tarih: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
            
            wb.save(path)
            QMessageBox.information(self, "Başarılı", f"Dosya kaydedildi:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Excel hatası:\n{str(e)}")
    
    def refresh_data(self):
        """Dışarıdan çağrılabilir yenileme"""
        self.load_cari_list()
        if self.current_cari_id:
            self._load_data()
