# -*- coding: utf-8 -*-
"""
REDLINE NEXOR ERP - İK Vardiya Planlama
Personel vardiya planlaması - haftalık görünüm ve atama
"""
import os
from datetime import datetime, date, timedelta
from PySide6.QtWidgets import (
    QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QFrame,
    QTableWidget, QTableWidgetItem, QHeaderView, QComboBox,
    QAbstractItemView, QMessageBox, QDialog, QWidget,
    QCheckBox, QListWidget, QListWidgetItem, QSizePolicy
)
from PySide6.QtCore import Qt, QTimer
from PySide6.QtGui import QColor

from components.base_page import BasePage
from core.database import get_db_connection
from core.log_manager import LogManager
from config import REPORT_OUTPUT_DIR
from core.firma_bilgileri import get_firma_bilgileri
from utils.etiket_yazdir import _register_dejavu_fonts


# Vardiya adına göre renk eşleme
VARDIYA_RENKLERI = {
    "Sabah": QColor(34, 197, 94, 50),
    "Gündüz": QColor(34, 197, 94, 50),
    "Akşam": QColor(59, 130, 246, 50),
    "Gece": QColor(139, 92, 246, 50),
}

IZIN_BG = QColor(245, 158, 11, 50)
IZIN_FG = QColor("#f59e0b")

LEGEND_ITEMS = [
    ("Sabah", "#22c55e"),
    ("Akşam", "#3b82f6"),
    ("Gece", "#8b5cf6"),
    ("Diğer", "#f97316"),
    ("İzin", "#f59e0b"),
]


def _vardiya_renk(vardiya_adi: str) -> QColor | None:
    """Vardiya adından arka plan rengi döndür"""
    for anahtar, renk in VARDIYA_RENKLERI.items():
        if anahtar in vardiya_adi:
            return renk
    # Tanınmayan vardiyalar turuncu
    return QColor(249, 115, 22, 50)


class TopluAtamaDialog(QDialog):
    """Birden fazla personele toplu vardiya atama dialogu"""

    def __init__(self, personeller: list, vardiyalar: list, theme: dict, parent=None):
        """
        personeller: [(id, ad_soyad), ...]
        vardiyalar: [(id, ad), ...]
        """
        super().__init__(parent)
        self.personeller = personeller
        self.vardiyalar = vardiyalar
        self.theme = theme
        self.result_data = None
        self.setWindowTitle("Toplu Vardiya Atama")
        self.setMinimumSize(500, 520)
        self.setModal(True)
        self._setup_ui()

    def _setup_ui(self):
        bg = self.theme.get('bg_card', '#151B23')
        txt = self.theme.get('text', '#E8ECF1')
        border = self.theme.get('border', '#1E2736')
        bg_input = self.theme.get('bg_input', '#232C3B')
        primary = self.theme.get('primary', '#C41E1E')

        self.setStyleSheet(f"""
            QDialog {{ background: {bg}; }}
            QLabel {{ color: {txt}; }}
        """)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(12)

        # --- Personel seçimi ---
        layout.addWidget(QLabel("Personel Seçimi:"))

        # Tümünü seç / kaldır
        select_all_layout = QHBoxLayout()
        self.select_all_cb = QCheckBox("Tümünü Seç")
        self.select_all_cb.setStyleSheet(f"color: {txt};")
        self.select_all_cb.stateChanged.connect(self._toggle_select_all)
        select_all_layout.addWidget(self.select_all_cb)
        select_all_layout.addStretch()
        layout.addLayout(select_all_layout)

        self.personel_list = QListWidget()
        self.personel_list.setStyleSheet(f"""
            QListWidget {{
                background: {bg_input};
                border: 1px solid {border};
                border-radius: 6px;
                color: {txt};
            }}
            QListWidget::item {{ padding: 4px; }}
        """)
        for per_id, per_ad in self.personeller:
            item = QListWidgetItem(per_ad)
            item.setData(Qt.UserRole, per_id)
            item.setCheckState(Qt.Unchecked)
            self.personel_list.addItem(item)
        layout.addWidget(self.personel_list, 1)

        # --- Gün seçimi ---
        layout.addWidget(QLabel("Gün Seçimi:"))
        gun_layout = QHBoxLayout()
        gun_isimleri = ["Pzt", "Sal", "Çar", "Per", "Cum", "Cmt", "Paz"]
        self.gun_checkboxes = []
        for gun in gun_isimleri:
            cb = QCheckBox(gun)
            cb.setStyleSheet(f"color: {txt};")
            cb.setChecked(True)
            gun_layout.addWidget(cb)
            self.gun_checkboxes.append(cb)
        layout.addLayout(gun_layout)

        # --- Vardiya seçimi ---
        layout.addWidget(QLabel("Vardiya:"))
        self.vardiya_combo = QComboBox()
        self.vardiya_combo.setStyleSheet(f"""
            QComboBox {{
                background: {bg_input};
                border: 1px solid {border};
                border-radius: 6px;
                padding: 8px;
                color: {txt};
            }}
        """)
        self.vardiya_combo.addItem("– (Boş / Kaldır)", None)
        for v_id, v_ad in self.vardiyalar:
            self.vardiya_combo.addItem(v_ad, v_id)
        layout.addWidget(self.vardiya_combo)

        # --- Butonlar ---
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        cancel_btn = QPushButton("İptal")
        cancel_btn.setStyleSheet(f"""
            QPushButton {{
                background: {bg_input};
                color: {txt};
                border: 1px solid {border};
                border-radius: 6px;
                padding: 8px 20px;
            }}
        """)
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)

        apply_btn = QPushButton("Uygula")
        apply_btn.setStyleSheet(f"""
            QPushButton {{
                background: {primary};
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 20px;
                font-weight: bold;
            }}
        """)
        apply_btn.clicked.connect(self._apply)
        btn_layout.addWidget(apply_btn)

        layout.addLayout(btn_layout)

    def _toggle_select_all(self, state):
        check = Qt.Checked if state == Qt.Checked.value else Qt.Unchecked
        for i in range(self.personel_list.count()):
            self.personel_list.item(i).setCheckState(check)

    def _apply(self):
        # Seçili personeller
        secili_personeller = []
        for i in range(self.personel_list.count()):
            item = self.personel_list.item(i)
            if item.checkState() == Qt.Checked:
                secili_personeller.append(item.data(Qt.UserRole))

        if not secili_personeller:
            QMessageBox.warning(self, "Uyarı", "En az bir personel seçmelisiniz.")
            return

        # Seçili günler (0=Pzt, 6=Paz)
        secili_gunler = []
        for idx, cb in enumerate(self.gun_checkboxes):
            if cb.isChecked():
                secili_gunler.append(idx)

        if not secili_gunler:
            QMessageBox.warning(self, "Uyarı", "En az bir gün seçmelisiniz.")
            return

        vardiya_id = self.vardiya_combo.currentData()

        self.result_data = {
            "personeller": secili_personeller,
            "gunler": secili_gunler,
            "vardiya_id": vardiya_id,
        }
        self.accept()


class IKVardiyaPage(BasePage):
    """İK Vardiya Planlama Sayfası"""

    # Sabit sütun indeksleri
    COL_PERSONEL = 0
    COL_DEPARTMAN = 1
    COL_POZISYON = 2
    COL_DAY_START = 3  # 3..9 = 7 gün

    def __init__(self, theme: dict):
        super().__init__(theme)
        self.current_week_start = self._get_week_start(date.today())
        self.vardiyalar = []          # [(id, ad, baslangic_saati, bitis_saati), ...]
        self.personeller = []         # [(id, ad_soyad, dept_ad, poz_ad, varsayilan_vardiya_id), ...]
        self.izin_map = {}            # {(personel_id, date): True}
        self.plan_map = {}            # {(personel_id, date): vardiya_id}
        self.changes = {}             # {(personel_id, date): vardiya_id | None}
        self._loading = False         # Sinyal döngüsü engelleme
        self._setup_ui()
        QTimer.singleShot(100, self._initial_load)

    # ──────────────────── Yardımcı ────────────────────

    def _get_week_start(self, d: date) -> date:
        return d - timedelta(days=d.weekday())

    def _button_style(self):
        return f"""
            QPushButton {{
                background: {self.theme.get('bg_input')};
                color: {self.theme.get('text')};
                border: 1px solid {self.theme.get('border')};
                border-radius: 6px;
                padding: 8px 16px;
            }}
            QPushButton:hover {{ background: {self.theme.get('bg_hover')}; }}
        """

    def _combo_style(self):
        return f"""
            QComboBox {{
                background: {self.theme.get('bg_input')};
                border: 1px solid {self.theme.get('border')};
                border-radius: 6px;
                padding: 8px;
                color: {self.theme.get('text')};
            }}
        """

    def _table_style(self):
        return f"""
            QTableWidget {{
                background: {self.theme.get('bg_card')};
                border: 1px solid {self.theme.get('border')};
                border-radius: 8px;
                gridline-color: {self.theme.get('border')};
                color: {self.theme.get('text')};
            }}
            QTableWidget::item {{
                padding: 4px;
            }}
            QHeaderView::section {{
                background: {self.theme.get('bg_input')};
                color: {self.theme.get('text')};
                padding: 10px;
                border: none;
                border-bottom: 2px solid {self.theme.get('primary')};
                font-weight: bold;
            }}
        """

    # ──────────────────── UI Setup ────────────────────

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(12)

        # ─── Header ───
        header = QHBoxLayout()

        title = QLabel("Vardiya Planlama")
        title.setStyleSheet(f"font-size: 20px; font-weight: bold; color: {self.theme.get('text')};")
        header.addWidget(title)
        header.addStretch()

        prev_btn = QPushButton("◀ Önceki Hafta")
        prev_btn.setStyleSheet(self._button_style())
        prev_btn.clicked.connect(self._prev_week)
        header.addWidget(prev_btn)

        self.week_label = QLabel()
        self.week_label.setStyleSheet(
            f"color: {self.theme.get('primary')}; font-weight: bold; font-size: 14px; margin: 0 16px;"
        )
        header.addWidget(self.week_label)

        next_btn = QPushButton("Sonraki Hafta ▶")
        next_btn.setStyleSheet(self._button_style())
        next_btn.clicked.connect(self._next_week)
        header.addWidget(next_btn)

        today_btn = QPushButton("Bugün")
        today_btn.setStyleSheet(f"""
            QPushButton {{
                background: {self.theme.get('primary')};
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }}
            QPushButton:hover {{ background: {self.theme.get('primary_hover')}; }}
        """)
        today_btn.clicked.connect(self._go_today)
        header.addWidget(today_btn)

        layout.addLayout(header)

        # ─── Filtre Çubuğu ───
        filter_frame = QFrame()
        filter_frame.setStyleSheet(f"background: {self.theme.get('bg_card')}; border-radius: 8px;")
        filter_layout = QHBoxLayout(filter_frame)
        filter_layout.setContentsMargins(16, 12, 16, 12)

        # Departman filtresi
        lbl_dept = QLabel("Departman:")
        lbl_dept.setStyleSheet(f"color: {self.theme.get('text')};")
        filter_layout.addWidget(lbl_dept)

        self.dept_combo = QComboBox()
        self.dept_combo.setStyleSheet(self._combo_style())
        self.dept_combo.setMinimumWidth(150)
        self.dept_combo.currentIndexChanged.connect(self._on_dept_changed)
        filter_layout.addWidget(self.dept_combo)

        # Pozisyon filtresi
        lbl_poz = QLabel("Pozisyon:")
        lbl_poz.setStyleSheet(f"color: {self.theme.get('text')}; margin-left: 12px;")
        filter_layout.addWidget(lbl_poz)

        self.poz_combo = QComboBox()
        self.poz_combo.setStyleSheet(self._combo_style())
        self.poz_combo.setMinimumWidth(150)
        self.poz_combo.currentIndexChanged.connect(self._on_filter_changed)
        filter_layout.addWidget(self.poz_combo)

        # Vardiya filtresi
        lbl_var = QLabel("Vardiya:")
        lbl_var.setStyleSheet(f"color: {self.theme.get('text')}; margin-left: 12px;")
        filter_layout.addWidget(lbl_var)

        self.vardiya_combo = QComboBox()
        self.vardiya_combo.setStyleSheet(self._combo_style())
        self.vardiya_combo.setMinimumWidth(130)
        self.vardiya_combo.currentIndexChanged.connect(self._on_filter_changed)
        filter_layout.addWidget(self.vardiya_combo)

        filter_layout.addStretch()

        # Vardiya renk göstergesi (legend)
        legend_layout = QHBoxLayout()
        for vardiya_ad, renk in LEGEND_ITEMS:
            dot = QLabel("●")
            dot.setStyleSheet(f"color: {renk}; font-size: 16px;")
            legend_layout.addWidget(dot)
            lbl = QLabel(vardiya_ad)
            lbl.setStyleSheet(f"color: {self.theme.get('text_muted')}; font-size: 11px; margin-right: 12px;")
            legend_layout.addWidget(lbl)
        filter_layout.addLayout(legend_layout)

        filter_layout.addStretch()

        # Excel butonu
        excel_btn = QPushButton("Excel")
        excel_btn.setStyleSheet(f"""
            QPushButton {{
                background: {self.theme.get('bg_input')};
                color: {self.theme.get('text')};
                border: 1px solid {self.theme.get('border')};
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: bold;
            }}
            QPushButton:hover {{ background: {self.theme.get('bg_hover')}; }}
        """)
        excel_btn.clicked.connect(self._export_excel)
        filter_layout.addWidget(excel_btn)

        # PDF butonu
        pdf_btn = QPushButton("PDF")
        pdf_btn.setStyleSheet(f"""
            QPushButton {{
                background: {self.theme.get('bg_input')};
                color: {self.theme.get('text')};
                border: 1px solid {self.theme.get('border')};
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: bold;
            }}
            QPushButton:hover {{ background: {self.theme.get('bg_hover')}; }}
        """)
        pdf_btn.clicked.connect(self._export_pdf)
        filter_layout.addWidget(pdf_btn)

        # Toplu Atama butonu
        toplu_btn = QPushButton("Toplu Atama")
        toplu_btn.setStyleSheet(f"""
            QPushButton {{
                background: {self.theme.get('info', '#3b82f6')};
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: bold;
            }}
            QPushButton:hover {{ background: #2563eb; }}
        """)
        toplu_btn.clicked.connect(self._open_toplu_atama)
        filter_layout.addWidget(toplu_btn)

        # Kaydet butonu
        self.save_btn = QPushButton("Kaydet")
        self.save_btn.setStyleSheet(f"""
            QPushButton {{
                background: {self.theme.get('success', '#10B981')};
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 20px;
                font-weight: bold;
            }}
            QPushButton:hover {{ background: #059669; }}
        """)
        self.save_btn.clicked.connect(self._save_changes)
        filter_layout.addWidget(self.save_btn)

        layout.addWidget(filter_frame)

        # ─── Ana Tablo ───
        self.table = QTableWidget()
        self.table.setColumnCount(10)  # Personel + Departman + Pozisyon + 7 gün
        self.table.setStyleSheet(self._table_style())
        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.cellDoubleClicked.connect(self._on_cell_double_clicked)

        # Sütun genişlikleri
        self.table.horizontalHeader().setSectionResizeMode(self.COL_PERSONEL, QHeaderView.Stretch)
        self.table.setColumnWidth(self.COL_DEPARTMAN, 120)
        self.table.setColumnWidth(self.COL_POZISYON, 120)
        for i in range(self.COL_DAY_START, self.COL_DAY_START + 7):
            self.table.setColumnWidth(i, 110)

        layout.addWidget(self.table, 1)

        # ─── Canlı Vardiya Monitörü ───
        self._setup_monitor_panel(layout)

    # ──────────────────── Veri Yükleme ────────────────────

    def _initial_load(self):
        """İlk açılışta departman/pozisyon/vardiya listelerini yükle, ardından tablo"""
        self._load_departmanlar()
        self._load_pozisyonlar_all()
        self._load_vardiyalar()
        self._load_data()

        # Canlı monitör: ilk yükleme + 5 sn timer
        self._load_monitor_data()
        self._monitor_timer = QTimer(self)
        self._monitor_timer.timeout.connect(self._load_monitor_data)
        self._monitor_timer.start(5000)

    def _load_departmanlar(self):
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT id, ad FROM ik.departmanlar WHERE aktif_mi = 1 ORDER BY ad")
            rows = cursor.fetchall()
            conn.close()

            self._loading = True
            self.dept_combo.clear()
            self.dept_combo.addItem("Tümü", None)
            for row in rows:
                self.dept_combo.addItem(row[1], row[0])
            self._loading = False
        except Exception as e:
            self._loading = False
            print(f"Departman yükleme hatası: {e}")

    def _load_pozisyonlar_all(self):
        """Tüm pozisyonları yükle (dahili cache)"""
        self._all_pozisyonlar = []  # [(id, ad, departman_id), ...]
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT id, ad, departman_id FROM ik.pozisyonlar WHERE aktif_mi = 1 ORDER BY ad")
            self._all_pozisyonlar = cursor.fetchall()
            conn.close()
        except Exception as e:
            print(f"Pozisyon yükleme hatası: {e}")

        self._update_poz_combo()

    def _update_poz_combo(self):
        """Pozisyon combo'sunu seçili departmana göre güncelle"""
        dept_id = self.dept_combo.currentData()

        self._loading = True
        self.poz_combo.clear()
        self.poz_combo.addItem("Tümü", None)
        for poz in self._all_pozisyonlar:
            if dept_id is None or poz[2] == dept_id:
                self.poz_combo.addItem(poz[1], poz[0])
        self._loading = False

    def _load_vardiyalar(self):
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute(
                "SELECT id, ad, baslangic_saati, bitis_saati FROM tanim.vardiyalar WHERE aktif_mi = 1"
            )
            self.vardiyalar = cursor.fetchall()
            conn.close()
        except Exception as e:
            print(f"Vardiya yükleme hatası: {e}")

        # Vardiya filtre combo'sunu doldur
        self._loading = True
        self.vardiya_combo.clear()
        self.vardiya_combo.addItem("Tümü", None)
        for v in self.vardiyalar:
            self.vardiya_combo.addItem(v[1], v[0])
        self._loading = False

    def _load_data(self):
        """Haftalık verileri yükle ve tabloyu doldur"""
        if self._loading:
            return

        try:
            week_end = self.current_week_start + timedelta(days=6)

            # Hafta etiketi
            self.week_label.setText(
                f"{self.current_week_start.strftime('%d.%m.%Y')} - {week_end.strftime('%d.%m.%Y')}"
            )

            # Tablo başlıkları
            gun_isimleri = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi", "Pazar"]
            headers = ["Personel", "Departman", "Pozisyon"]
            for i in range(7):
                gun = self.current_week_start + timedelta(days=i)
                headers.append(f"{gun_isimleri[i]}\n{gun.strftime('%d.%m')}")
            self.table.setHorizontalHeaderLabels(headers)

            conn = get_db_connection()
            cursor = conn.cursor()

            # ── Personeller ──
            where = ["p.aktif_mi = 1"]
            params = []

            dept_id = self.dept_combo.currentData()
            if dept_id:
                where.append("p.departman_id = ?")
                params.append(dept_id)

            poz_id = self.poz_combo.currentData()
            if poz_id:
                where.append("p.pozisyon_id = ?")
                params.append(poz_id)

            cursor.execute(f"""
                SELECT p.id, p.ad + ' ' + p.soyad AS ad_soyad,
                       ISNULL(d.ad, '') AS dept, ISNULL(poz.ad, '') AS pozisyon,
                       p.varsayilan_vardiya_id
                FROM ik.personeller p
                LEFT JOIN ik.departmanlar d ON p.departman_id = d.id
                LEFT JOIN ik.pozisyonlar poz ON p.pozisyon_id = poz.id
                WHERE {' AND '.join(where)}
                ORDER BY d.ad, poz.ad, p.ad
            """, params)
            self.personeller = cursor.fetchall()

            # ── Haftalık planlar ──
            cursor.execute("""
                SELECT personel_id, tarih, vardiya_id
                FROM ik.vardiya_planlama
                WHERE tarih BETWEEN ? AND ?
            """, (self.current_week_start, week_end))

            self.plan_map = {}
            for row in cursor.fetchall():
                self.plan_map[(row[0], row[1])] = row[2]

            # ── Onaylı izinler ──
            cursor.execute("""
                SELECT personel_id, baslangic_tarihi, bitis_tarihi
                FROM ik.izin_talepleri
                WHERE durum = 'ONAYLANDI'
                  AND baslangic_tarihi <= ? AND bitis_tarihi >= ?
            """, (week_end, self.current_week_start))

            self.izin_map = {}
            for row in cursor.fetchall():
                gun_sayisi = (row[2] - row[1]).days + 1
                for d in range(gun_sayisi):
                    izin_gun = row[1] + timedelta(days=d)
                    if self.current_week_start <= izin_gun <= week_end:
                        self.izin_map[(row[0], izin_gun)] = True

            conn.close()

            # Değişiklikleri sıfırla
            self.changes = {}

            # Tabloyu doldur
            self._populate_table()

        except Exception as e:
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "Hata", f"Veri yüklenemedi: {e}")

    def _get_filtered_personeller(self):
        """Vardiya filtresine göre personel listesini döndür"""
        filtre_vardiya_id = self.vardiya_combo.currentData() if hasattr(self, 'vardiya_combo') else None
        if filtre_vardiya_id is None:
            return list(self.personeller)

        filtered = []
        for per in self.personeller:
            per_id = per[0]
            varsayilan_vardiya_id = per[4]
            # Hafta içinde bu vardiyaya atanmış mı kontrol et
            match = False
            for day_idx in range(7):
                gun = self.current_week_start + timedelta(days=day_idx)
                key = (per_id, gun)
                plan_v = self.plan_map.get(key)
                if plan_v == filtre_vardiya_id:
                    match = True
                    break
            # Plan yoksa varsayılan vardiyaya bak
            if not match and varsayilan_vardiya_id == filtre_vardiya_id:
                # Hiçbir günde planı yoksa varsayılana göre göster
                has_any_plan = any(
                    (per_id, self.current_week_start + timedelta(days=d)) in self.plan_map
                    for d in range(7)
                )
                if not has_any_plan:
                    match = True
            if match:
                filtered.append(per)
        return filtered

    def _populate_table(self):
        """Tabloyu personel ve plan verileriyle doldur"""
        display_personeller = self._get_filtered_personeller()
        self.table.setRowCount(0)
        self.table.setRowCount(len(display_personeller))

        for row_idx, per in enumerate(display_personeller):
            per_id, ad_soyad, dept_ad, poz_ad, varsayilan_vardiya_id = per

            self.table.setRowHeight(row_idx, 42)

            # Personel adı
            name_item = QTableWidgetItem(ad_soyad)
            name_item.setData(Qt.UserRole, per_id)
            name_item.setFlags(name_item.flags() & ~Qt.ItemIsEditable)
            self.table.setItem(row_idx, self.COL_PERSONEL, name_item)

            # Departman
            dept_item = QTableWidgetItem(dept_ad)
            dept_item.setFlags(dept_item.flags() & ~Qt.ItemIsEditable)
            dept_item.setForeground(QColor(self.theme.get('text_muted')))
            self.table.setItem(row_idx, self.COL_DEPARTMAN, dept_item)

            # Pozisyon
            poz_item = QTableWidgetItem(poz_ad)
            poz_item.setFlags(poz_item.flags() & ~Qt.ItemIsEditable)
            poz_item.setForeground(QColor(self.theme.get('text_muted')))
            self.table.setItem(row_idx, self.COL_POZISYON, poz_item)

            # 7 gün
            for day_idx in range(7):
                gun = self.current_week_start + timedelta(days=day_idx)
                col = self.COL_DAY_START + day_idx
                key = (per_id, gun)

                cell = QTableWidgetItem()
                cell.setTextAlignment(Qt.AlignCenter)
                cell.setFlags(cell.flags() & ~Qt.ItemIsEditable)
                # Personel id ve tarih bilgisini sakla
                cell.setData(Qt.UserRole, per_id)
                cell.setData(Qt.UserRole + 1, gun)

                if key in self.izin_map:
                    cell.setText("İZİN")
                    cell.setBackground(IZIN_BG)
                    cell.setForeground(IZIN_FG)
                elif key in self.plan_map:
                    vardiya_id = self.plan_map[key]
                    vardiya_adi = self._vardiya_adi(vardiya_id)
                    cell.setText(vardiya_adi)
                    renk = _vardiya_renk(vardiya_adi)
                    if renk:
                        cell.setBackground(renk)
                else:
                    cell.setText("–")
                    cell.setForeground(QColor(self.theme.get('text_muted')))

                self.table.setItem(row_idx, col, cell)

    def _vardiya_adi(self, vardiya_id) -> str:
        """Vardiya ID'den adını bul"""
        for v in self.vardiyalar:
            if v[0] == vardiya_id:
                return v[1]
        return "–"

    # ──────────────────── Hafta Navigasyonu ────────────────────

    def _prev_week(self):
        if self.changes:
            if not self._confirm_discard():
                return
        self.current_week_start -= timedelta(days=7)
        self._load_data()

    def _next_week(self):
        if self.changes:
            if not self._confirm_discard():
                return
        self.current_week_start += timedelta(days=7)
        self._load_data()

    def _go_today(self):
        if self.changes:
            if not self._confirm_discard():
                return
        self.current_week_start = self._get_week_start(date.today())
        self._load_data()

    def _confirm_discard(self) -> bool:
        reply = QMessageBox.question(
            self, "Kaydedilmemiş Değişiklikler",
            "Kaydedilmemiş değişiklikler var. Devam ederseniz kaybolacak.\nDevam etmek istiyor musunuz?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        return reply == QMessageBox.Yes

    # ──────────────────── Filtre ────────────────────

    def _on_dept_changed(self):
        if self._loading:
            return
        self._update_poz_combo()
        self._load_data()

    def _on_filter_changed(self):
        if self._loading:
            return
        self._load_data()

    # ──────────────────── Hücre Düzenleme ────────────────────

    def _on_cell_double_clicked(self, row: int, col: int):
        """Gün hücresine çift tıklandığında QComboBox aç,
        personel adına çift tıklandığında tüm hafta vardiyasını değiştir"""
        if col == self.COL_PERSONEL:
            self._change_personel_week_vardiya(row)
            return

        if col < self.COL_DAY_START:
            return

        item = self.table.item(row, col)
        if not item:
            return

        per_id = item.data(Qt.UserRole)
        gun = item.data(Qt.UserRole + 1)

        # İzinli günde düzenleme yapma
        if (per_id, gun) in self.izin_map:
            return

        # Mevcut widget varsa kaldır
        existing = self.table.cellWidget(row, col)
        if existing:
            return

        combo = QComboBox()
        combo.setStyleSheet(self._combo_style())

        # Seçenekler: Boş + tüm vardiyalar
        combo.addItem("–", None)
        current_vardiya_id = self.changes.get((per_id, gun), self.plan_map.get((per_id, gun)))
        current_index = 0

        for idx, v in enumerate(self.vardiyalar):
            combo.addItem(v[1], v[0])
            if v[0] == current_vardiya_id:
                current_index = idx + 1

        combo.setCurrentIndex(current_index)

        # Seçim değiştiğinde
        combo.currentIndexChanged.connect(
            lambda _idx, r=row, c=col, p=per_id, g=gun: self._on_vardiya_selected(r, c, p, g)
        )

        # Odak kaybedince combo'yu kaldır
        combo.activated.connect(
            lambda _idx, r=row, c=col, p=per_id, g=gun: self._finalize_combo(r, c, p, g)
        )

        self.table.setCellWidget(row, col, combo)
        combo.showPopup()

    def _on_vardiya_selected(self, row: int, col: int, per_id: int, gun: date):
        """ComboBox'ta seçim yapıldığında değişikliği kaydet"""
        combo = self.table.cellWidget(row, col)
        if not combo:
            return
        vardiya_id = combo.currentData()
        self.changes[(per_id, gun)] = vardiya_id

    def _finalize_combo(self, row: int, col: int, per_id: int, gun: date):
        """ComboBox seçimi tamamlanınca widget'ı kaldır ve hücreyi güncelle"""
        combo = self.table.cellWidget(row, col)
        if not combo:
            return

        vardiya_id = combo.currentData()
        self.changes[(per_id, gun)] = vardiya_id

        self.table.removeCellWidget(row, col)

        # Hücreyi güncelle
        item = self.table.item(row, col)
        if not item:
            return

        if vardiya_id is not None:
            vardiya_adi = self._vardiya_adi(vardiya_id)
            item.setText(vardiya_adi)
            item.setForeground(QColor(self.theme.get('text')))
            renk = _vardiya_renk(vardiya_adi)
            if renk:
                item.setBackground(renk)
        else:
            item.setText("–")
            item.setForeground(QColor(self.theme.get('text_muted')))
            item.setBackground(QColor(self.theme.get('bg_card')))

    # ──────────────────── Tekil Vardiya Değiştirme ────────────────────

    def _change_personel_week_vardiya(self, row: int):
        """Personel adına çift tıklandığında tüm hafta vardiyasını değiştir"""
        name_item = self.table.item(row, self.COL_PERSONEL)
        if not name_item:
            return
        per_id = name_item.data(Qt.UserRole)
        per_ad = name_item.text()

        # Mevcut baskın vardiyayı bul
        current_vardiya_id = None
        for day_idx in range(7):
            gun = self.current_week_start + timedelta(days=day_idx)
            key = (per_id, gun)
            v_id = self.changes.get(key, self.plan_map.get(key))
            if v_id:
                current_vardiya_id = v_id
                break

        # Dialog oluştur
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Vardiya Değiştir - {per_ad}")
        dialog.setMinimumWidth(350)
        dialog.setModal(True)
        dialog.setStyleSheet(f"""
            QDialog {{ background: {self.theme.get('bg_card', '#151B23')}; }}
            QLabel {{ color: {self.theme.get('text', '#E8ECF1')}; }}
        """)

        dlg_layout = QVBoxLayout(dialog)
        dlg_layout.setContentsMargins(20, 20, 20, 20)
        dlg_layout.setSpacing(12)

        lbl = QLabel(f"{per_ad} için tüm haftanın vardiyasını seçin:")
        lbl.setWordWrap(True)
        lbl.setStyleSheet(f"font-size: 13px; color: {self.theme.get('text')};")
        dlg_layout.addWidget(lbl)

        combo = QComboBox()
        combo.setStyleSheet(self._combo_style())
        combo.addItem("– (Boş / Kaldır)", None)
        current_index = 0
        for idx, v in enumerate(self.vardiyalar):
            combo.addItem(v[1], v[0])
            if v[0] == current_vardiya_id:
                current_index = idx + 1
        combo.setCurrentIndex(current_index)
        dlg_layout.addWidget(combo)

        # Gün seçimi
        gun_lbl = QLabel("Uygulanacak günler:")
        gun_lbl.setStyleSheet(f"font-size: 11px; color: {self.theme.get('text_muted')};")
        dlg_layout.addWidget(gun_lbl)

        gun_isimleri = ["Pzt", "Sal", "Çar", "Per", "Cum", "Cmt", "Paz"]
        gun_layout = QHBoxLayout()
        gun_cbs = []
        for g in gun_isimleri:
            cb = QCheckBox(g)
            cb.setStyleSheet(f"color: {self.theme.get('text')};")
            cb.setChecked(True)
            gun_layout.addWidget(cb)
            gun_cbs.append(cb)
        dlg_layout.addLayout(gun_layout)

        # Butonlar
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        cancel_btn = QPushButton("İptal")
        cancel_btn.setStyleSheet(self._button_style())
        cancel_btn.clicked.connect(dialog.reject)
        btn_layout.addWidget(cancel_btn)

        apply_btn = QPushButton("Uygula")
        apply_btn.setStyleSheet(f"""
            QPushButton {{
                background: {self.theme.get('primary')};
                color: white; border: none; border-radius: 6px;
                padding: 8px 24px; font-weight: bold;
            }}
        """)
        apply_btn.clicked.connect(dialog.accept)
        btn_layout.addWidget(apply_btn)
        dlg_layout.addLayout(btn_layout)

        if dialog.exec() != QDialog.Accepted:
            return

        new_vardiya_id = combo.currentData()

        for day_idx in range(7):
            if not gun_cbs[day_idx].isChecked():
                continue
            gun = self.current_week_start + timedelta(days=day_idx)
            if (per_id, gun) in self.izin_map:
                continue
            self.changes[(per_id, gun)] = new_vardiya_id

        self._apply_changes_to_table()

    # ──────────────────── Toplu Atama ────────────────────

    def _open_toplu_atama(self):
        """Toplu vardiya atama dialogunu aç"""
        # Tablodaki mevcut personelleri al
        per_list = [(p[0], p[1]) for p in self.personeller]
        var_list = [(v[0], v[1]) for v in self.vardiyalar]

        if not per_list:
            QMessageBox.warning(self, "Uyarı", "Tabloda personel bulunmuyor.")
            return

        dialog = TopluAtamaDialog(per_list, var_list, self.theme, self)
        if dialog.exec() == QDialog.Accepted and dialog.result_data:
            data = dialog.result_data
            for per_id in data["personeller"]:
                for day_offset in data["gunler"]:
                    gun = self.current_week_start + timedelta(days=day_offset)
                    # İzinli günleri atla
                    if (per_id, gun) in self.izin_map:
                        continue
                    self.changes[(per_id, gun)] = data["vardiya_id"]

            # Tabloyu görsel olarak güncelle
            self._apply_changes_to_table()

    def _apply_changes_to_table(self):
        """Değişiklikleri tablodaki hücrelere yansıt"""
        for row_idx in range(self.table.rowCount()):
            name_item = self.table.item(row_idx, self.COL_PERSONEL)
            if not name_item:
                continue
            per_id = name_item.data(Qt.UserRole)

            for day_idx in range(7):
                gun = self.current_week_start + timedelta(days=day_idx)
                col = self.COL_DAY_START + day_idx
                key = (per_id, gun)

                if key not in self.changes:
                    continue

                # İzinli günleri atla
                if key in self.izin_map:
                    continue

                item = self.table.item(row_idx, col)
                if not item:
                    continue

                vardiya_id = self.changes[key]
                if vardiya_id is not None:
                    vardiya_adi = self._vardiya_adi(vardiya_id)
                    item.setText(vardiya_adi)
                    item.setForeground(QColor(self.theme.get('text')))
                    renk = _vardiya_renk(vardiya_adi)
                    if renk:
                        item.setBackground(renk)
                else:
                    item.setText("–")
                    item.setForeground(QColor(self.theme.get('text_muted')))
                    item.setBackground(QColor(self.theme.get('bg_card')))

    # ──────────────────── Kaydetme ────────────────────

    def _save_changes(self):
        """Değişiklikleri veritabanına kaydet"""
        if not self.changes:
            QMessageBox.information(self, "Bilgi", "Kaydedilecek değişiklik yok.")
            return

        try:
            conn = get_db_connection()
            cursor = conn.cursor()

            for (per_id, gun), vardiya_id in self.changes.items():
                if vardiya_id is not None:
                    # MERGE: varsa güncelle, yoksa ekle
                    cursor.execute("""
                        MERGE ik.vardiya_planlama AS hedef
                        USING (SELECT ? AS personel_id, ? AS tarih) AS kaynak
                        ON hedef.personel_id = kaynak.personel_id AND hedef.tarih = kaynak.tarih
                        WHEN MATCHED THEN
                            UPDATE SET vardiya_id = ?
                        WHEN NOT MATCHED THEN
                            INSERT (personel_id, tarih, vardiya_id)
                            VALUES (?, ?, ?);
                    """, (per_id, gun, vardiya_id, per_id, gun, vardiya_id))
                else:
                    # Vardiya kaldırma: kaydı sil
                    cursor.execute("""
                        DELETE FROM ik.vardiya_planlama
                        WHERE personel_id = ? AND tarih = ?
                    """, (per_id, gun))

            conn.commit()
            degisiklik_sayisi = len(self.changes)
            LogManager.log_update('ik', 'ik.vardiya_planlama', None,
                                  f'Vardiya planlama guncellendi: {degisiklik_sayisi} atama')
            conn.close()

            self.changes = {}

            QMessageBox.information(
                self, "Başarılı",
                f"{degisiklik_sayisi} vardiya ataması kaydedildi."
            )

            # Verileri yeniden yükle
            self._load_data()

        except Exception as e:
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "Hata", f"Kaydetme hatası: {e}")

    # ──────────────────── Ortak Gruplama ────────────────────

    def _grupla_personeller(self):
        """Personelleri baskın vardiyalarına göre grupla.
        Returns: (vardiya_gruplari dict, atanmamis list)
        """
        vardiya_gruplari = {}
        atanmamis = []

        for per in self.personeller:
            per_id, ad_soyad, dept_ad, poz_ad, varsayilan_vardiya_id = per
            vardiya_sayac = {}
            for day_idx in range(7):
                gun = self.current_week_start + timedelta(days=day_idx)
                key = (per_id, gun)
                if key in self.izin_map:
                    continue
                v_id = self.plan_map.get(key)
                if v_id:
                    vardiya_sayac[v_id] = vardiya_sayac.get(v_id, 0) + 1

            if vardiya_sayac:
                baskin = max(vardiya_sayac, key=vardiya_sayac.get)
                vardiya_gruplari.setdefault(baskin, []).append(per)
            elif varsayilan_vardiya_id:
                vardiya_gruplari.setdefault(varsayilan_vardiya_id, []).append(per)
            else:
                atanmamis.append(per)

        return vardiya_gruplari, atanmamis

    def _gun_metni(self, per_id, gun, grup_vardiya_id=None):
        """Bir personelin belirli gün için hücre metnini döndür.
        grup_vardiya_id verilirse, aynı vardiya için kısa gösterim (checkmark) kullanılır.
        """
        key = (per_id, gun)
        if key in self.izin_map:
            return "İZİN"
        elif key in self.plan_map:
            v_id = self.plan_map[key]
            if grup_vardiya_id and v_id == grup_vardiya_id:
                return "\u2713"  # ✓
            return self._vardiya_adi(v_id)
        return "–"

    # ──────────────────── PDF Dışa Aktarma ────────────────────

    def _export_pdf(self):
        """Haftalık vardiya planını vardiya bazında gruplandırarak PDF olarak dışa aktar"""
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import mm
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, HRFlowable,
        )
        from reportlab.lib.styles import ParagraphStyle

        if not self.personeller:
            QMessageBox.warning(self, "Uyarı", "Dışa aktarılacak veri bulunmuyor.")
            return

        has_font = _register_dejavu_fonts()
        font_name = 'NexorFont' if has_font else 'Helvetica'
        font_bold = 'NexorFont-Bold' if has_font else 'Helvetica-Bold'

        os.makedirs(REPORT_OUTPUT_DIR, exist_ok=True)
        week_end = self.current_week_start + timedelta(days=6)
        # Vardiya filtresi seçiliyse dosya adına ekle (çakışma önlenir)
        filtre_vardiya_id = self.vardiya_combo.currentData()
        vardiya_suffix = ""
        if filtre_vardiya_id is not None:
            vardiya_suffix = f"_{self.vardiya_combo.currentText().replace(' ', '_')}"
        filename = (
            f"vardiya_plani_{self.current_week_start.strftime('%Y%m%d')}"
            f"_{week_end.strftime('%Y%m%d')}{vardiya_suffix}.pdf"
        )
        pdf_path = os.path.join(str(REPORT_OUTPUT_DIR), filename)
        firma = get_firma_bilgileri()

        doc = SimpleDocTemplate(
            pdf_path, pagesize=landscape(A4),
            leftMargin=12 * mm, rightMargin=12 * mm,
            topMargin=12 * mm, bottomMargin=12 * mm,
        )

        # Stiller
        title_style = ParagraphStyle(
            'VPTitle', fontName=font_bold, fontSize=14, alignment=1, spaceAfter=2,
        )
        subtitle_style = ParagraphStyle(
            'VPSub', fontName=font_name, fontSize=9, alignment=1, spaceAfter=2,
            textColor=colors.HexColor('#555555'),
        )
        info_style = ParagraphStyle(
            'VPInfo', fontName=font_name, fontSize=8, alignment=1, spaceAfter=4,
            textColor=colors.HexColor('#777777'),
        )
        section_style = ParagraphStyle(
            'VPSection', fontName=font_bold, fontSize=11, spaceAfter=2, spaceBefore=2,
            textColor=colors.HexColor('#1a1a1a'),
        )
        count_style = ParagraphStyle(
            'VPCount', fontName=font_name, fontSize=8, spaceAfter=4,
            textColor=colors.HexColor('#555555'),
        )
        # Hücre içi Paragraph stili
        cell_style = ParagraphStyle('VPCell', fontName=font_name, fontSize=7, leading=9)
        cell_bold_style = ParagraphStyle('VPCellB', fontName=font_bold, fontSize=7, leading=9)
        cell_center = ParagraphStyle('VPCellC', fontName=font_name, fontSize=7, leading=9, alignment=1)
        cell_center_bold = ParagraphStyle('VPCellCB', fontName=font_bold, fontSize=7, leading=9, alignment=1)

        # Gün başlıkları
        gun_isimleri = ["Pzt", "Sal", "Çar", "Per", "Cum", "Cmt", "Paz"]
        header_labels = ["#", "Personel", "Departman", "Pozisyon"]
        for i in range(7):
            gun = self.current_week_start + timedelta(days=i)
            header_labels.append(f"{gun_isimleri[i]}\n{gun.strftime('%d.%m')}")

        headers = [Paragraph(h.replace('\n', '<br/>'), cell_center_bold) for h in header_labels]

        # Grupla
        vardiya_gruplari, atanmamis = self._grupla_personeller()

        # Filtre bilgisi
        dept_text = self.dept_combo.currentText()
        poz_text = self.poz_combo.currentText()
        filtre_vardiya_text = self.vardiya_combo.currentText()
        filtre = (
            f"Departman: {dept_text}  |  Pozisyon: {poz_text}  |  "
            f"Vardiya: {filtre_vardiya_text}  |  "
            f"Oluşturulma: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        )

        # Sütun genişlikleri - sabit ve orantılı
        page_w = landscape(A4)[0] - 24 * mm
        col_widths = [
            page_w * 0.03,   # #
            page_w * 0.14,   # Personel
            page_w * 0.10,   # Departman
            page_w * 0.10,   # Pozisyon
        ]
        day_w = (page_w - sum(col_widths)) / 7
        col_widths += [day_w] * 7

        def _build_rows(per_list, grup_vardiya_id=None):
            rows = []
            for sira, per in enumerate(per_list, 1):
                per_id, ad_soyad, dept_ad, poz_ad, _ = per
                row = [
                    Paragraph(str(sira), cell_center),
                    Paragraph(ad_soyad, cell_bold_style),
                    Paragraph(dept_ad, cell_style),
                    Paragraph(poz_ad, cell_style),
                ]
                for day_idx in range(7):
                    gun = self.current_week_start + timedelta(days=day_idx)
                    metin = self._gun_metni(per_id, gun, grup_vardiya_id)
                    row.append(Paragraph(metin, cell_center_bold if metin == "İZİN" else cell_center))
                rows.append(row)
            return rows

        def _build_table_style(data_len, per_list):
            cmds = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#5a7a9b')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#d0d0d0')),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ]
            for i in range(1, data_len):
                bg = '#f5f7fa' if i % 2 == 0 else '#ffffff'
                cmds.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor(bg)))
                if i - 1 < len(per_list):
                    per_id = per_list[i - 1][0]
                    for day_idx in range(7):
                        gun = self.current_week_start + timedelta(days=day_idx)
                        col_idx = 4 + day_idx
                        if (per_id, gun) in self.izin_map:
                            cmds.append(('BACKGROUND', (col_idx, i), (col_idx, i),
                                         colors.HexColor('#fef3c7')))
                        elif (per_id, gun) in self.plan_map:
                            cmds.append(('BACKGROUND', (col_idx, i), (col_idx, i),
                                         colors.HexColor('#e8f5e9')))
            return TableStyle(cmds)

        # ── Sayfa oluştur ──
        elements = []
        vardiya_sira = [(v[0], v[1]) for v in self.vardiyalar]
        filtre_vardiya_id = self.vardiya_combo.currentData()
        ilk_bolum = True

        for v_id, v_adi in vardiya_sira:
            per_list = vardiya_gruplari.get(v_id, [])
            if not per_list:
                continue
            if filtre_vardiya_id is not None and v_id != filtre_vardiya_id:
                continue

            if not ilk_bolum:
                elements.append(PageBreak())

            # Başlık bloğu
            elements.append(Paragraph(firma.get('name', 'NEXOR ERP'), title_style))
            elements.append(Paragraph(
                f"Haftalık Vardiya Planı &mdash; "
                f"{self.current_week_start.strftime('%d.%m.%Y')} - {week_end.strftime('%d.%m.%Y')}",
                subtitle_style,
            ))
            elements.append(Paragraph(filtre, info_style))
            elements.append(HRFlowable(
                width="100%", thickness=1.5, color=colors.HexColor('#5a7a9b'),
                spaceBefore=2, spaceAfter=3,
            ))
            elements.append(Paragraph(f"{v_adi} Vardiyası", section_style))
            elements.append(Paragraph(
                f"{len(per_list)} personel  |  "
                f"\u2713 = {v_adi}  |  İZİN = İzinli",
                count_style,
            ))

            table_data = [headers] + _build_rows(per_list, v_id)
            tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
            tbl.setStyle(_build_table_style(len(table_data), per_list))
            elements.append(tbl)
            elements.append(Spacer(1, 4 * mm))
            ilk_bolum = False

        # Atanmamış
        if atanmamis and filtre_vardiya_id is None:
            if not ilk_bolum:
                elements.append(PageBreak())
            elements.append(Paragraph(firma.get('name', 'NEXOR ERP'), title_style))
            elements.append(Paragraph(
                f"Haftalık Vardiya Planı &mdash; "
                f"{self.current_week_start.strftime('%d.%m.%Y')} - {week_end.strftime('%d.%m.%Y')}",
                subtitle_style,
            ))
            elements.append(Paragraph(filtre, info_style))
            elements.append(HRFlowable(
                width="100%", thickness=1.5, color=colors.HexColor('#5a7a9b'),
                spaceBefore=2, spaceAfter=3,
            ))
            elements.append(Paragraph("Vardiya Atanmamış", section_style))
            elements.append(Paragraph(f"{len(atanmamis)} personel", count_style))
            table_data = [headers] + _build_rows(atanmamis)
            tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
            tbl.setStyle(_build_table_style(len(table_data), atanmamis))
            elements.append(tbl)
            ilk_bolum = False

        if not elements:
            QMessageBox.warning(self, "Uyarı", "Seçili filtreye uygun personel bulunmuyor.")
            return

        try:
            doc.build(elements)
            os.startfile(pdf_path)
        except Exception as e:
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "Hata", f"PDF oluşturulamadı: {e}")

    # ──────────────────── Excel Dışa Aktarma ────────────────────

    def _export_excel(self):
        """Haftalık vardiya planını vardiya bazında gruplandırarak Excel'e aktar"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            QMessageBox.critical(self, "Hata", "openpyxl kütüphanesi bulunamadı.\npip install openpyxl")
            return

        if not self.personeller:
            QMessageBox.warning(self, "Uyarı", "Dışa aktarılacak veri bulunmuyor.")
            return

        os.makedirs(REPORT_OUTPUT_DIR, exist_ok=True)
        week_end = self.current_week_start + timedelta(days=6)
        filtre_vardiya_id = self.vardiya_combo.currentData()
        vardiya_suffix = ""
        if filtre_vardiya_id is not None:
            vardiya_suffix = f"_{self.vardiya_combo.currentText().replace(' ', '_')}"
        filename = (
            f"vardiya_plani_{self.current_week_start.strftime('%Y%m%d')}"
            f"_{week_end.strftime('%Y%m%d')}{vardiya_suffix}.xlsx"
        )
        excel_path = os.path.join(str(REPORT_OUTPUT_DIR), filename)
        firma = get_firma_bilgileri()

        wb = Workbook()
        # Varsayılan sayfayı kaldır
        wb.remove(wb.active)

        # Stiller
        header_font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='5A7A9B')
        section_font = Font(name='Calibri', bold=True, size=12, color='1A1A1A')
        title_font = Font(name='Calibri', bold=True, size=14)
        info_font = Font(name='Calibri', size=9, color='777777')
        data_font = Font(name='Calibri', size=10)
        bold_font = Font(name='Calibri', bold=True, size=10)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0'),
        )
        izin_fill = PatternFill('solid', fgColor='FEF3C7')
        izin_font = Font(name='Calibri', bold=True, size=10, color='D97706')
        check_fill = PatternFill('solid', fgColor='E8F5E9')
        check_font = Font(name='Calibri', bold=True, size=10, color='2E7D32')
        alt_fill = PatternFill('solid', fgColor='F5F7FA')

        # Gün başlıkları
        gun_isimleri = ["Pzt", "Sal", "Çar", "Per", "Cum", "Cmt", "Paz"]
        header_labels = ["#", "Personel", "Departman", "Pozisyon"]
        for i in range(7):
            gun = self.current_week_start + timedelta(days=i)
            header_labels.append(f"{gun_isimleri[i]} {gun.strftime('%d.%m')}")

        # Grupla
        vardiya_gruplari, atanmamis = self._grupla_personeller()
        vardiya_sira = [(v[0], v[1]) for v in self.vardiyalar]
        filtre_vardiya_id = self.vardiya_combo.currentData()

        dept_text = self.dept_combo.currentText()
        poz_text = self.poz_combo.currentText()
        filtre_vardiya_text = self.vardiya_combo.currentText()

        def _write_sheet(ws, sheet_title, per_list, grup_vardiya_id=None):
            """Bir vardiya grubunu Excel sayfasına yaz"""
            # Başlık
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
            ws.cell(1, 1, firma.get('name', 'NEXOR ERP')).font = title_font
            ws.cell(1, 1).alignment = Alignment(horizontal='center')

            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11)
            ws.cell(2, 1,
                    f"Haftalık Vardiya Planı - "
                    f"{self.current_week_start.strftime('%d.%m.%Y')} - {week_end.strftime('%d.%m.%Y')}"
                    ).font = info_font
            ws.cell(2, 1).alignment = Alignment(horizontal='center')

            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=11)
            ws.cell(3, 1,
                    f"Departman: {dept_text}  |  Pozisyon: {poz_text}  |  "
                    f"Vardiya: {filtre_vardiya_text}  |  "
                    f"Oluşturulma: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
                    ).font = info_font
            ws.cell(3, 1).alignment = Alignment(horizontal='center')

            # Bölüm başlığı
            ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=11)
            ws.cell(4, 1, f"{sheet_title}  ({len(per_list)} personel)").font = section_font

            # Tablo başlığı (satır 5)
            hdr_row = 5
            for col_idx, label in enumerate(header_labels, 1):
                c = ws.cell(hdr_row, col_idx, label)
                c.font = header_font
                c.fill = header_fill
                c.alignment = center
                c.border = thin_border

            # Veriler
            for sira, per in enumerate(per_list, 1):
                row = hdr_row + sira
                per_id, ad_soyad, dept_ad, poz_ad, _ = per

                ws.cell(row, 1, sira).font = data_font
                ws.cell(row, 1).alignment = center
                ws.cell(row, 1).border = thin_border

                ws.cell(row, 2, ad_soyad).font = bold_font
                ws.cell(row, 2).alignment = left
                ws.cell(row, 2).border = thin_border

                ws.cell(row, 3, dept_ad).font = data_font
                ws.cell(row, 3).alignment = left
                ws.cell(row, 3).border = thin_border

                ws.cell(row, 4, poz_ad).font = data_font
                ws.cell(row, 4).alignment = left
                ws.cell(row, 4).border = thin_border

                # Alternatif satır rengi
                if sira % 2 == 0:
                    for c in range(1, 12):
                        ws.cell(row, c).fill = alt_fill

                for day_idx in range(7):
                    gun = self.current_week_start + timedelta(days=day_idx)
                    col = 5 + day_idx
                    metin = self._gun_metni(per_id, gun, grup_vardiya_id)
                    c = ws.cell(row, col, metin)
                    c.alignment = center
                    c.border = thin_border

                    if metin == "İZİN":
                        c.font = izin_font
                        c.fill = izin_fill
                    elif metin == "\u2713":
                        c.font = check_font
                        c.fill = check_fill
                    else:
                        c.font = data_font

            # Sütun genişlikleri
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 22
            ws.column_dimensions['C'].width = 18
            ws.column_dimensions['D'].width = 20
            for i, letter in enumerate(['E', 'F', 'G', 'H', 'I', 'J', 'K']):
                ws.column_dimensions[letter].width = 12

        # ── Her vardiya için ayrı sayfa ──
        for v_id, v_adi in vardiya_sira:
            per_list = vardiya_gruplari.get(v_id, [])
            if not per_list:
                continue
            if filtre_vardiya_id is not None and v_id != filtre_vardiya_id:
                continue

            ws = wb.create_sheet(title=v_adi[:31])
            _write_sheet(ws, f"{v_adi} Vardiyası", per_list, v_id)

        # Atanmamış
        if atanmamis and filtre_vardiya_id is None:
            ws = wb.create_sheet(title="Atanmamış")
            _write_sheet(ws, "Vardiya Atanmamış", atanmamis)

        if len(wb.sheetnames) == 0:
            QMessageBox.warning(self, "Uyarı", "Seçili filtreye uygun personel bulunmuyor.")
            return

        try:
            wb.save(excel_path)
            os.startfile(excel_path)
        except Exception as e:
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "Hata", f"Excel oluşturulamadı: {e}")

    # ──────────────────── Canlı Vardiya Monitörü ────────────────────

    def _setup_monitor_panel(self, parent_layout):
        """Canlı vardiya monitör panelini oluştur"""
        self.monitor_frame = QFrame()
        self.monitor_frame.setObjectName("monitorFrame")
        self.monitor_frame.setFixedHeight(180)
        self.monitor_frame.setStyleSheet(f"""
            QFrame#monitorFrame {{
                background: {self.theme.get('bg_card')};
                border: 1px solid {self.theme.get('border')};
                border-radius: 10px;
            }}
        """)

        m_lay = QVBoxLayout(self.monitor_frame)
        m_lay.setContentsMargins(16, 12, 16, 12)
        m_lay.setSpacing(8)

        # Başlık
        m_title = QLabel("Aktif Vardiya Durumu")
        m_title.setStyleSheet(
            f"color: {self.theme.get('text')}; font-weight: bold; font-size: 13px; border: none;"
        )
        m_lay.addWidget(m_title)

        # Özet kartları
        cards = QHBoxLayout()
        cards.setSpacing(12)

        def _card(title, color):
            f = QFrame()
            f.setFixedSize(180, 52)
            f.setStyleSheet(
                f"QFrame {{ background: {self.theme.get('bg_input')}; "
                f"border-radius: 8px; border-left: 3px solid {color}; }}"
            )
            vl = QVBoxLayout(f)
            vl.setContentsMargins(10, 6, 10, 6)
            vl.setSpacing(2)
            t = QLabel(title)
            t.setStyleSheet(f"color: {self.theme.get('text_muted')}; font-size: 10px;")
            vl.addWidget(t)
            v = QLabel("–")
            v.setStyleSheet(f"color: {color}; font-size: 14px; font-weight: bold;")
            vl.addWidget(v)
            cards.addWidget(f)
            return v

        self._mv_vardiya = _card("Vardiya", self.theme.get('primary', '#C41E1E'))
        self._mv_beklenen = _card("Beklenen", '#3b82f6')
        self._mv_gelen = _card("Gelen", '#22c55e')
        self._mv_gelmeyen = _card("Gelmeyen", '#ef4444')

        cards.addStretch()
        m_lay.addLayout(cards)

        # Gelmeyen personel listesi
        self._mv_absent = QLabel("")
        self._mv_absent.setWordWrap(True)
        self._mv_absent.setStyleSheet(
            f"color: {self.theme.get('text_muted')}; font-size: 12px; border: none;"
        )
        m_lay.addWidget(self._mv_absent)

        m_lay.addStretch()
        parent_layout.addWidget(self.monitor_frame)

    def _load_monitor_data(self):
        """Canlı vardiya monitörünü güncelle"""
        try:
            now = datetime.now()
            current_time = now.time()

            # Aktif vardiyayı bul
            aktif = None
            for v in self.vardiyalar:
                v_id, v_ad, bas, bit = v
                # timedelta → time dönüşümü (pyodbc bazen timedelta döner)
                if isinstance(bas, timedelta):
                    total = int(bas.total_seconds())
                    bas = datetime(2000, 1, 1, total // 3600, (total % 3600) // 60).time()
                if isinstance(bit, timedelta):
                    total = int(bit.total_seconds())
                    bit = datetime(2000, 1, 1, total // 3600, (total % 3600) // 60).time()

                if bas <= bit:
                    # Normal vardiya
                    if bas <= current_time < bit:
                        aktif = (v_id, v_ad, bas, bit)
                        break
                else:
                    # Gece vardiyası (bitis < baslangic)
                    if current_time >= bas or current_time < bit:
                        aktif = (v_id, v_ad, bas, bit)
                        break

            if not aktif:
                self._mv_vardiya.setText("Aktif vardiya yok")
                self._mv_beklenen.setText("–")
                self._mv_gelen.setText("–")
                self._mv_gelmeyen.setText("–")
                self._mv_absent.setText("Şu anda aktif vardiya bulunmuyor.")
                return

            a_id, a_ad, a_bas, a_bit = aktif
            self._mv_vardiya.setText(f"{a_ad} {a_bas.strftime('%H:%M')}-{a_bit.strftime('%H:%M')}")

            conn = get_db_connection()
            cursor = conn.cursor()

            # Bu vardiyaya atanmış personeller (plan veya varsayılan)
            cursor.execute("""
                SELECT DISTINCT p.id, p.ad + ' ' + p.soyad,
                       p.gunluk_durum, ISNULL(d.ad, '')
                FROM ik.personeller p
                LEFT JOIN ik.departmanlar d ON p.departman_id = d.id
                LEFT JOIN ik.vardiya_planlama vp
                    ON vp.personel_id = p.id AND vp.tarih = CAST(GETDATE() AS DATE)
                WHERE p.aktif_mi = 1
                  AND (vp.vardiya_id = ? OR (vp.id IS NULL AND p.varsayilan_vardiya_id = ?))
            """, (a_id, a_id))
            personeller = cursor.fetchall()

            # Bugün izinli olanlar
            cursor.execute("""
                SELECT personel_id FROM ik.izin_talepleri
                WHERE durum = 'ONAYLANDI'
                  AND baslangic_tarihi <= CAST(GETDATE() AS DATE)
                  AND bitis_tarihi >= CAST(GETDATE() AS DATE)
            """)
            izinli_ids = {row[0] for row in cursor.fetchall()}

            conn.close()

            beklenen = []
            gelenler = []
            gelmeyenler = []

            for per in personeller:
                per_id, ad_soyad, durum, dept = per
                if per_id in izinli_ids:
                    continue
                beklenen.append(per)
                if durum == 1:
                    gelenler.append(per)
                else:
                    gelmeyenler.append(per)

            self._mv_beklenen.setText(str(len(beklenen)))
            self._mv_gelen.setText(str(len(gelenler)))
            self._mv_gelmeyen.setText(str(len(gelmeyenler)))

            if gelmeyenler:
                parts = [f"● {p[1]} ({p[3]})" for p in gelmeyenler]
                self._mv_absent.setText("Gelmeyen Personel:  " + "    ".join(parts))
            else:
                self._mv_absent.setText("Tüm personel mevcut.")

        except Exception as e:
            print(f"Monitör güncelleme hatası: {e}")
