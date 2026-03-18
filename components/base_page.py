# -*- coding: utf-8 -*-
"""
NEXOR ERP - Base Page v2.0
Tum sayfalarin miras alacagi temel sinif
Modern ve tutarli stil tanimlari
"""
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QFrame,
    QPushButton, QLineEdit, QComboBox, QTableWidget,
    QHeaderView, QScrollArea, QGroupBox, QFileDialog,
    QMessageBox, QMenu
)
from PySide6.QtCore import Qt
from PySide6.QtGui import QFont
from datetime import datetime


class BasePage(QWidget):
    """Tum sayfalar icin temel sinif"""

    def __init__(self, theme: dict):
        super().__init__()
        self.theme = theme
        self._apply_global_style()

    def update_theme(self, theme: dict):
        """Temayi guncelle (sayfayi yeniden olusturmadan)"""
        self.theme = theme
        self._apply_global_style()
    
    def _apply_global_style(self):
        """Global stil - Tum bilesenler icin"""
        t = self.theme
        
        self.setStyleSheet(f"""
            /* ========================================
               GENEL
            ======================================== */
            QWidget {{
                background: transparent;
                color: {t['text']};
                font-family: 'Segoe UI', 'Inter', -apple-system, sans-serif;
                font-size: 13px;
            }}
            
            /* ========================================
               LABEL
            ======================================== */
            QLabel {{
                background: transparent;
                color: {t['text']};
                padding: 0;
            }}
            
            /* ========================================
               FRAME / KART
            ======================================== */
            QFrame {{
                background: transparent;
                border: none;
            }}
            
            QFrame[frameShape="StyledPanel"] {{
                background: {t['bg_card']};
                border: 1px solid {t['border']};
                border-radius: 10px;
            }}
            
            /* ========================================
               GROUP BOX
            ======================================== */
            QGroupBox {{
                background: {t['bg_card']};
                border: 1px solid {t['border']};
                border-radius: 10px;
                margin-top: 20px;
                padding: 20px;
                padding-top: 35px;
                font-weight: 600;
                font-size: 14px;
                color: {t['text']};
            }}
            
            QGroupBox::title {{
                subcontrol-origin: margin;
                subcontrol-position: top left;
                left: 16px;
                top: 8px;
                padding: 0 8px;
                color: {t['text']};
                background: {t['bg_card']};
            }}
            
            /* ========================================
               INPUT - LINE EDIT
            ======================================== */
            QLineEdit {{
                background: {t['bg_input']};
                border: 1px solid {t['border_input']};
                border-radius: 8px;
                padding: 10px 14px;
                color: {t['text']};
                font-size: 13px;
                selection-background-color: {t['primary']};
                selection-color: {t['text_inverse']};
            }}
            
            QLineEdit:hover {{
                border-color: {t['border_light']};
            }}
            
            QLineEdit:focus {{
                border-color: {t['primary']};
                background: {t['bg_input_focus']};
            }}
            
            QLineEdit:disabled {{
                background: {t['bg_hover']};
                color: {t['text_disabled']};
            }}
            
            QLineEdit[readOnly="true"] {{
                background: {t['bg_hover']};
            }}
            
            /* ========================================
               INPUT - SPINBOX
            ======================================== */
            QSpinBox, QDoubleSpinBox {{
                background: {t['bg_input']};
                border: 1px solid {t['border_input']};
                border-radius: 8px;
                padding: 8px 12px;
                color: {t['text']};
                font-size: 13px;
            }}
            
            QSpinBox:hover, QDoubleSpinBox:hover {{
                border-color: {t['border_light']};
            }}
            
            QSpinBox:focus, QDoubleSpinBox:focus {{
                border-color: {t['primary']};
            }}
            
            QSpinBox::up-button, QDoubleSpinBox::up-button,
            QSpinBox::down-button, QDoubleSpinBox::down-button {{
                border: none;
                background: transparent;
                width: 20px;
            }}
            
            QSpinBox::up-arrow, QDoubleSpinBox::up-arrow {{
                image: none;
                border-left: 4px solid transparent;
                border-right: 4px solid transparent;
                border-bottom: 5px solid {t['text_muted']};
                width: 0;
                height: 0;
            }}
            
            QSpinBox::down-arrow, QDoubleSpinBox::down-arrow {{
                image: none;
                border-left: 4px solid transparent;
                border-right: 4px solid transparent;
                border-top: 5px solid {t['text_muted']};
                width: 0;
                height: 0;
            }}
            
            /* ========================================
               INPUT - COMBOBOX
            ======================================== */
            QComboBox {{
                background: {t['bg_input']};
                border: 1px solid {t['border_input']};
                border-radius: 8px;
                padding: 8px 14px;
                padding-right: 30px;
                color: {t['text']};
                font-size: 13px;
                min-height: 20px;
            }}
            
            QComboBox:hover {{
                border-color: {t['border_light']};
            }}
            
            QComboBox:focus {{
                border-color: {t['primary']};
            }}
            
            QComboBox::drop-down {{
                border: none;
                width: 30px;
            }}
            
            QComboBox::down-arrow {{
                image: none;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent;
                border-top: 6px solid {t['text_muted']};
                width: 0;
                height: 0;
                margin-right: 10px;
            }}
            
            QComboBox QAbstractItemView {{
                background: {t['bg_dropdown']};
                border: 1px solid {t['border']};
                border-radius: 8px;
                padding: 4px;
                selection-background-color: {t['bg_hover']};
                selection-color: {t['text']};
                outline: none;
            }}
            
            QComboBox QAbstractItemView::item {{
                padding: 8px 12px;
                border-radius: 4px;
                min-height: 20px;
            }}
            
            QComboBox QAbstractItemView::item:hover {{
                background: {t['bg_hover']};
            }}
            
            QComboBox QAbstractItemView::item:selected {{
                background: {t['bg_selected']};
                color: {t['primary']};
            }}
            
            /* ========================================
               INPUT - TEXTAREA
            ======================================== */
            QTextEdit, QPlainTextEdit {{
                background: {t['bg_input']};
                border: 1px solid {t['border_input']};
                border-radius: 8px;
                padding: 10px 14px;
                color: {t['text']};
                font-size: 13px;
                selection-background-color: {t['primary']};
            }}
            
            QTextEdit:focus, QPlainTextEdit:focus {{
                border-color: {t['primary']};
            }}
            
            /* ========================================
               INPUT - DATE/TIME
            ======================================== */
            QDateEdit, QTimeEdit, QDateTimeEdit {{
                background: {t['bg_input']};
                border: 1px solid {t['border_input']};
                border-radius: 8px;
                padding: 8px 12px;
                color: {t['text']};
                font-size: 13px;
            }}
            
            QDateEdit:focus, QTimeEdit:focus, QDateTimeEdit:focus {{
                border-color: {t['primary']};
            }}
            
            QDateEdit::drop-down, QTimeEdit::drop-down, QDateTimeEdit::drop-down {{
                border: none;
                width: 25px;
            }}
            
            QCalendarWidget {{
                background: {t['bg_card']};
            }}
            
            QCalendarWidget QToolButton {{
                color: {t['text']};
                background: transparent;
                border: none;
                border-radius: 4px;
                padding: 6px;
            }}
            
            QCalendarWidget QToolButton:hover {{
                background: {t['bg_hover']};
            }}
            
            /* ========================================
               CHECKBOX
            ======================================== */
            QCheckBox {{
                color: {t['text']};
                spacing: 10px;
                font-size: 13px;
            }}
            
            QCheckBox::indicator {{
                width: 20px;
                height: 20px;
                border-radius: 4px;
                border: 2px solid {t['border']};
                background: {t['bg_input']};
            }}
            
            QCheckBox::indicator:hover {{
                border-color: {t['primary']};
            }}
            
            QCheckBox::indicator:checked {{
                background: {t['primary']};
                border-color: {t['primary']};
            }}
            
            QCheckBox::indicator:checked:hover {{
                background: {t['primary_hover']};
            }}
            
            /* ========================================
               RADIO BUTTON
            ======================================== */
            QRadioButton {{
                color: {t['text']};
                spacing: 10px;
                font-size: 13px;
            }}
            
            QRadioButton::indicator {{
                width: 20px;
                height: 20px;
                border-radius: 10px;
                border: 2px solid {t['border']};
                background: {t['bg_input']};
            }}
            
            QRadioButton::indicator:hover {{
                border-color: {t['primary']};
            }}
            
            QRadioButton::indicator:checked {{
                background: {t['primary']};
                border-color: {t['primary']};
            }}
            
            /* ========================================
               BUTTON - STANDART
            ======================================== */
            QPushButton {{
                background: {t['bg_card']};
                border: 1px solid {t['border']};
                border-radius: 8px;
                padding: 10px 20px;
                color: {t['text']};
                font-size: 13px;
                font-weight: 500;
                min-height: 18px;
            }}
            
            QPushButton:hover {{
                background: {t['bg_hover']};
                border-color: {t['border_light']};
            }}
            
            QPushButton:pressed {{
                background: {t['bg_selected']};
            }}
            
            QPushButton:disabled {{
                background: {t['bg_hover']};
                color: {t['text_disabled']};
                border-color: {t['border']};
            }}
            
            /* ========================================
               BUTTON - PRIMARY (Ozel class ile)
            ======================================== */
            QPushButton[class="primary"] {{
                background: {t['primary']};
                border: none;
                color: {t['text_inverse']};
                font-weight: 600;
            }}
            
            QPushButton[class="primary"]:hover {{
                background: {t['primary_hover']};
            }}
            
            /* ========================================
               BUTTON - SUCCESS
            ======================================== */
            QPushButton[class="success"] {{
                background: {t['success']};
                border: none;
                color: white;
                font-weight: 600;
            }}
            
            QPushButton[class="success"]:hover {{
                background: {t.get('success_dark', t['success'])};
            }}
            
            /* ========================================
               BUTTON - DANGER
            ======================================== */
            QPushButton[class="danger"] {{
                background: {t['error']};
                border: none;
                color: white;
                font-weight: 600;
            }}
            
            QPushButton[class="danger"]:hover {{
                background: {t.get('error_dark', t['error'])};
            }}
            
            /* ========================================
               TABLE
            ======================================== */
            QTableWidget, QTableView {{
                background: {t['table_row_bg']};
                border: 1px solid {t['border']};
                border-radius: 10px;
                gridline-color: {t['table_border']};
                selection-background-color: {t['table_row_selected']};
                selection-color: {t['text']};
                outline: none;
            }}
            
            QTableWidget::item, QTableView::item {{
                padding: 12px 16px;
                border: none;
            }}
            
            QTableWidget::item:hover, QTableView::item:hover {{
                background: {t['table_row_hover']};
            }}
            
            QTableWidget::item:selected, QTableView::item:selected {{
                background: {t['table_row_selected']};
            }}
            
            QTableWidget::item:alternate, QTableView::item:alternate {{
                background: {t['table_row_alt']};
            }}
            
            QHeaderView::section {{
                background: {t['table_header_bg']};
                color: {t['table_header_text']};
                padding: 12px 16px;
                border: none;
                border-bottom: 2px solid {t['primary']};
                font-weight: 600;
                font-size: 12px;
            }}
            
            QHeaderView::section:first {{
                border-top-left-radius: 12px;
            }}
            
            QHeaderView::section:last {{
                border-top-right-radius: 12px;
            }}
            
            /* ========================================
               SCROLLBAR - VERTICAL
            ======================================== */
            QScrollBar:vertical {{
                background: {t['scrollbar_bg']};
                width: 10px;
                border-radius: 5px;
                margin: 0;
            }}
            
            QScrollBar::handle:vertical {{
                background: {t['scrollbar_thumb']};
                border-radius: 5px;
                min-height: 30px;
            }}
            
            QScrollBar::handle:vertical:hover {{
                background: {t['scrollbar_thumb_hover']};
            }}
            
            QScrollBar::add-line:vertical,
            QScrollBar::sub-line:vertical {{
                height: 0;
                background: none;
            }}
            
            QScrollBar::add-page:vertical,
            QScrollBar::sub-page:vertical {{
                background: none;
            }}
            
            /* ========================================
               SCROLLBAR - HORIZONTAL
            ======================================== */
            QScrollBar:horizontal {{
                background: {t['scrollbar_bg']};
                height: 10px;
                border-radius: 5px;
                margin: 0;
            }}
            
            QScrollBar::handle:horizontal {{
                background: {t['scrollbar_thumb']};
                border-radius: 5px;
                min-width: 30px;
            }}
            
            QScrollBar::handle:horizontal:hover {{
                background: {t['scrollbar_thumb_hover']};
            }}
            
            QScrollBar::add-line:horizontal,
            QScrollBar::sub-line:horizontal {{
                width: 0;
                background: none;
            }}
            
            /* ========================================
               TAB WIDGET
            ======================================== */
            QTabWidget::pane {{
                background: {t['bg_card']};
                border: 1px solid {t['border']};
                border-radius: 10px;
                padding: 16px;
                top: -1px;
            }}
            
            QTabBar::tab {{
                background: transparent;
                color: {t['text_muted']};
                padding: 12px 24px;
                margin-right: 4px;
                border: none;
                border-bottom: 2px solid transparent;
                font-weight: 500;
            }}
            
            QTabBar::tab:hover {{
                color: {t['text']};
                background: {t['bg_hover']};
            }}
            
            QTabBar::tab:selected {{
                color: {t['primary']};
                border-bottom-color: {t['primary']};
            }}
            
            /* ========================================
               TOOLTIP
            ======================================== */
            QToolTip {{
                background: {t['bg_tooltip']};
                color: {t.get('text_inverse', '#FFFFFF') if t['mode'] == 'light' else t['text']};
                border: 1px solid {t['border']};
                border-radius: 6px;
                padding: 8px 12px;
                font-size: 12px;
            }}
            
            /* ========================================
               MENU
            ======================================== */
            QMenu {{
                background: {t['bg_dropdown']};
                border: 1px solid {t['border']};
                border-radius: 8px;
                padding: 6px;
            }}
            
            QMenu::item {{
                padding: 10px 16px;
                border-radius: 4px;
            }}
            
            QMenu::item:selected {{
                background: {t['bg_hover']};
            }}
            
            QMenu::separator {{
                height: 1px;
                background: {t['border']};
                margin: 6px 0;
            }}
            
            /* ========================================
               PROGRESS BAR
            ======================================== */
            QProgressBar {{
                background: {t['bg_hover']};
                border: none;
                border-radius: 6px;
                height: 10px;
                text-align: center;
            }}
            
            QProgressBar::chunk {{
                background: {t['primary']};
                border-radius: 6px;
            }}
            
            /* ========================================
               SLIDER
            ======================================== */
            QSlider::groove:horizontal {{
                background: {t['bg_hover']};
                height: 6px;
                border-radius: 3px;
            }}
            
            QSlider::handle:horizontal {{
                background: {t['primary']};
                width: 18px;
                height: 18px;
                margin: -6px 0;
                border-radius: 9px;
            }}
            
            QSlider::handle:horizontal:hover {{
                background: {t['primary_hover']};
            }}
            
            /* ========================================
               SCROLL AREA
            ======================================== */
            QScrollArea {{
                background: transparent;
                border: none;
            }}
            
            QScrollArea > QWidget > QWidget {{
                background: transparent;
            }}
            
            /* ========================================
               SPLITTER
            ======================================== */
            QSplitter::handle {{
                background: {t['border']};
            }}
            
            QSplitter::handle:horizontal {{
                width: 2px;
            }}
            
            QSplitter::handle:vertical {{
                height: 2px;
            }}
        """)
    
    # =========================================================================
    # YARDIMCI METODLAR
    # =========================================================================
    
    def create_card(self, title: str = None) -> QFrame:
        """Standart kart olustur"""
        card = QFrame()
        card.setFrameShape(QFrame.StyledPanel)
        
        layout = QVBoxLayout(card)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(16)
        
        if title:
            title_label = QLabel(title)
            title_label.setStyleSheet(f"""
                font-size: 16px;
                font-weight: 600;
                color: {self.theme['text']};
                padding-bottom: 8px;
                border-bottom: 1px solid {self.theme['border']};
            """)
            layout.addWidget(title_label)
        
        return card
    
    def create_stat_card(self, label: str, value: str, icon: str = None,
                         color: str = None) -> QFrame:
        """Istatistik karti olustur (PRAXIS: sol accent cizgi + uppercase label)"""
        t = self.theme
        accent = color or t['primary']

        card = QFrame()
        card.setStyleSheet(f"""
            QFrame {{
                background: {t['bg_card']};
                border: 1px solid {t['border']};
                border-left: 3px solid {accent};
                border-radius: 10px;
            }}
            QFrame:hover {{
                border-color: {accent};
                border-left: 3px solid {accent};
            }}
        """)

        layout = QHBoxLayout(card)
        layout.setContentsMargins(16, 14, 16, 14)
        layout.setSpacing(12)

        if icon:
            icon_label = QLabel(icon)
            icon_label.setStyleSheet(f"""
                background: {t.get('bg_hover', t['bg_card'])};
                padding: 8px;
                border-radius: 8px;
                font-size: 18px;
                border: none;
            """)
            layout.addWidget(icon_label)

        text_layout = QVBoxLayout()
        text_layout.setSpacing(2)

        lbl = QLabel(label.upper())
        lbl.setStyleSheet(f"""
            color: {t['text_muted']};
            font-size: 10px;
            font-weight: 600;
            letter-spacing: 1px;
            border: none;
        """)
        text_layout.addWidget(lbl)

        value_label = QLabel(value)
        value_label.setObjectName("stat_value")
        value_label.setStyleSheet(f"""
            color: {accent};
            font-size: 22px;
            font-weight: 700;
            border: none;
        """)
        text_layout.addWidget(value_label)

        layout.addLayout(text_layout)
        layout.addStretch()

        return card
    
    def create_primary_button(self, text: str) -> QPushButton:
        """Primary buton olustur"""
        btn = QPushButton(text)
        btn.setCursor(Qt.PointingHandCursor)
        btn.setProperty("class", "primary")
        return btn
    
    def create_success_button(self, text: str) -> QPushButton:
        """Success buton olustur"""
        btn = QPushButton(text)
        btn.setCursor(Qt.PointingHandCursor)
        btn.setProperty("class", "success")
        return btn
    
    def create_danger_button(self, text: str) -> QPushButton:
        """Danger buton olustur"""
        btn = QPushButton(text)
        btn.setCursor(Qt.PointingHandCursor)
        btn.setProperty("class", "danger")
        return btn
    
    def create_badge(self, text: str, variant: str = "default") -> QLabel:
        """Badge olustur (success, warning, error, info, default)"""
        t = self.theme
        
        colors = {
            "success": (t['success_bg'], t['success_text']),
            "warning": (t['warning_bg'], t['warning_text']),
            "error": (t['error_bg'], t['error_text']),
            "info": (t['info_bg'], t['info_text']),
            "default": (t['badge_bg'], t['text_secondary']),
        }
        
        bg, fg = colors.get(variant, colors["default"])
        
        badge = QLabel(text)
        badge.setStyleSheet(f"""
            background: {bg};
            color: {fg};
            padding: 5px 12px;
            border-radius: 6px;
            font-size: 12px;
            font-weight: 500;
        """)
        badge.setAlignment(Qt.AlignCenter)
        
        return badge
    
    def create_section_title(self, text: str, icon: str = None) -> QLabel:
        """Bolum basligi olustur"""
        display_text = f"{icon} {text}" if icon else text

        label = QLabel(display_text)
        label.setStyleSheet(f"""
            color: {self.theme['text']};
            font-size: 18px;
            font-weight: 600;
            padding: 8px 0;
        """)

        return label

    def create_section_label(self, text: str) -> QLabel:
        """PRAXIS uppercase bolum etiketi (OZET, FILTRELER vb.)"""
        label = QLabel(text.upper())
        label.setStyleSheet(f"""
            color: {self.theme['text_muted']};
            font-size: 10px;
            font-weight: 600;
            letter-spacing: 1px;
        """)
        return label

    def create_page_header(self, title: str, subtitle: str = None) -> QHBoxLayout:
        """PRAXIS sayfa basligi: kirmizi accent cizgi + baslik + alt baslik"""
        t = self.theme
        header = QHBoxLayout()
        header.setSpacing(0)

        accent = QFrame()
        accent.setFixedSize(4, 36)
        accent.setStyleSheet(f"background: {t['primary']}; border-radius: 2px;")
        header.addWidget(accent)

        title_section = QVBoxLayout()
        title_section.setSpacing(2)

        title_lbl = QLabel(title)
        title_lbl.setStyleSheet(f"""
            color: {t['text']};
            font-size: 22px;
            font-weight: 600;
            margin-left: 12px;
        """)
        title_section.addWidget(title_lbl)

        if subtitle:
            sub_lbl = QLabel(subtitle)
            sub_lbl.setObjectName("page_subtitle")
            sub_lbl.setStyleSheet(f"""
                color: {t['text_secondary']};
                font-size: 13px;
                margin-left: 12px;
            """)
            title_section.addWidget(sub_lbl)

        header.addLayout(title_section)
        header.addStretch()

        return header

    def create_action_buttons(self, buttons: list) -> QWidget:
        """
        Tablo islem sutunu icin PRAXIS kurumsal stil butonlar olustur.
        Tema uyumlu (dark + light), hover + pressed state destegi.

        Args:
            buttons: [(icon, tooltip, callback, type), ...]
                     type: "edit", "delete", "view", "print", "photo",
                           "primary", "success", "danger", "warning", "info"

        Ornek:
            widget = self.create_action_buttons([
                ("✏️", "Duzenle", lambda: self.edit(rid), "edit"),
                ("🗑️", "Sil", lambda: self.delete(rid), "delete"),
                ("👁️", "Goruntule", lambda: self.view(rid), "view"),
            ])
            self.table.setCellWidget(row, col, widget)
        """
        t = self.theme

        # Parlak solid renkler
        color_map = {
            "edit":      "#2563EB",
            "delete":    "#EF4444",
            "view":      "#8B5CF6",
            "print":     "#6366F1",
            "photo":     "#0EA5E9",
            "primary":   t.get('primary', '#DC2626'),
            "success":   "#10B981",
            "danger":    "#EF4444",
            "warning":   "#F59E0B",
            "info":      "#3B82F6",
            "secondary": "#64748B",
        }

        # Hover icin daha koyu ton
        hover_map = {
            "edit":      "#1D4ED8",
            "delete":    "#DC2626",
            "view":      "#7C3AED",
            "print":     "#4F46E5",
            "photo":     "#0284C7",
            "primary":   t.get('primary_hover', '#9B1818'),
            "success":   "#059669",
            "danger":    "#DC2626",
            "warning":   "#D97706",
            "info":      "#2563EB",
            "secondary": "#475569",
        }

        # Tip -> metin etiketi (emoji render problemini onler)
        label_map = {
            "edit":      "Duzenle",
            "delete":    "Sil",
            "view":      "Detay",
            "print":     "Yazdir",
            "photo":     "Foto",
            "primary":   "Islem",
            "success":   "Onayla",
            "danger":    "Sil",
            "warning":   "Cikis",
            "info":      "Detay",
            "secondary": "...",
        }

        widget = QWidget()

        # Tum buton stillerini container uzerinde tanimla
        # BasePage global QPushButton stilini ezmek icin en yakin parent kazanir
        style_parts = ["QWidget { background: transparent; }"]
        btn_styles = []

        for i, btn_config in enumerate(buttons):
            if len(btn_config) == 4:
                icon, tooltip, callback, btn_type = btn_config
            elif len(btn_config) == 3:
                icon, tooltip, callback = btn_config
                btn_type = "edit"
            else:
                continue

            c = color_map.get(btn_type, "#2563EB")
            h = hover_map.get(btn_type, "#1D4ED8")
            btn_styles.append((btn_type, c, h, tooltip, callback, i))

        # Container stylesheet - BasePage'in QPushButton stilini tamamen ezecek
        css = "QWidget { background: transparent; }\n"
        css += "QPushButton {\n"
        css += "    border: none;\n"
        css += "    border-radius: 4px;\n"
        css += "    font-size: 11px;\n"
        css += "    font-weight: 600;\n"
        css += "    font-family: 'Segoe UI';\n"
        css += "    padding: 2px 10px;\n"
        css += "    min-width: 40px;\n"
        css += "    min-height: 0px;\n"
        css += "    max-height: 24px;\n"
        css += "    color: white;\n"
        css += "}\n"
        css += "QPushButton:hover { opacity: 0.9; }\n"
        widget.setStyleSheet(css)

        layout = QHBoxLayout(widget)
        layout.setContentsMargins(4, 3, 4, 3)
        layout.setSpacing(6)
        layout.setAlignment(Qt.AlignCenter)

        for btn_type, c, h, tooltip, callback, i in btn_styles:
            display = label_map.get(btn_type, tooltip[:8])

            btn = QPushButton(display)
            btn.setToolTip(tooltip)
            btn.setCursor(Qt.PointingHandCursor)

            # Sadece renk farkli - diger stiller container'dan geliyor
            btn.setStyleSheet(f"""
                QPushButton {{ background: {c}; }}
                QPushButton:hover {{ background: {h}; }}
                QPushButton:pressed {{ background: {h}; }}
            """)

            if callback:
                btn.clicked.connect(callback)

            layout.addWidget(btn)

        return widget

    # ══════════════════════════════════════════════
    # EXPORT ALTYAPISI
    # ══════════════════════════════════════════════

    def export_table_to_excel(self, table: QTableWidget, title: str = "Veri",
                               skip_last_col: bool = True):
        """
        Herhangi bir QTableWidget'i Excel'e aktar.
        NEXOR kurumsal temasina uygun, otomatik kolon genisligi.

        Args:
            table: Aktarilacak QTableWidget
            title: Sayfa ve dosya adi
            skip_last_col: Son sutunu atla (genelde islem butonu kolonu)
        """
        if table is None:
            QMessageBox.warning(self, "Hata", "Aktarilacak tablo bulunamadi.")
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            QMessageBox.warning(self, "Hata",
                                "openpyxl kutuphanesi yuklu degil.\npip install openpyxl")
            return

        safe_title = "".join(c for c in title if c.isalnum() or c in (' ', '_', '-')).strip()
        filepath, _ = QFileDialog.getSaveFileName(
            self, "Excel Kaydet",
            f"{safe_title}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel Dosyasi (*.xlsx)"
        )
        if not filepath:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = safe_title[:31]

            col_count = table.columnCount()
            if skip_last_col:
                col_count -= 1

            # -- Stiller --
            primary = self.theme.get('primary', '#DC2626').lstrip('#')
            header_font = Font(name='Segoe UI', bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color=primary, end_color=primary, fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

            data_font = Font(name='Segoe UI', size=10)
            data_align = Alignment(vertical="center", wrap_text=True)
            number_align = Alignment(horizontal="right", vertical="center")

            thin_border = Border(
                left=Side(style='thin', color='D1D5DB'),
                right=Side(style='thin', color='D1D5DB'),
                top=Side(style='thin', color='D1D5DB'),
                bottom=Side(style='thin', color='D1D5DB')
            )

            alt_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

            # -- Baslik satiri --
            title_font = Font(name='Segoe UI', bold=True, size=14, color=primary)
            ws.cell(row=1, column=1, value=title).font = title_font
            ws.cell(row=2, column=1,
                    value=f"Olusturma: {datetime.now().strftime('%d.%m.%Y %H:%M')}").font = Font(
                name='Segoe UI', size=9, color='6B7280')
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(col_count, 1))
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(col_count, 1))

            header_row = 4
            col_max_widths = {}

            # -- Tablo baslik satirlari --
            excel_col = 1
            for col in range(col_count):
                header = table.horizontalHeaderItem(col)
                val = header.text() if header else f"Sutun {col + 1}"
                cell = ws.cell(row=header_row, column=excel_col, value=val)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border
                col_max_widths[excel_col] = len(val) + 4
                excel_col += 1

            # -- Veri satirlari --
            for row in range(table.rowCount()):
                excel_row = header_row + 1 + row
                excel_col = 1
                for col in range(col_count):
                    item = table.item(row, col)
                    val = item.text() if item else ""

                    cell = ws.cell(row=excel_row, column=excel_col, value=val)
                    cell.font = data_font
                    cell.border = thin_border

                    # Sayisal deger kontrolu
                    clean = val.replace('.', '').replace(',', '.').replace('%', '').strip()
                    try:
                        num = float(clean)
                        cell.value = num
                        cell.alignment = number_align
                        if '%' in val:
                            cell.number_format = '0.00%'
                            cell.value = num / 100 if num > 1 else num
                    except (ValueError, TypeError):
                        cell.alignment = data_align

                    # Zebra renklendirme
                    if row % 2 == 1:
                        cell.fill = alt_fill

                    # Kolon genisligi hesapla
                    w = len(val) + 2
                    if w > col_max_widths.get(excel_col, 0):
                        col_max_widths[excel_col] = w

                    excel_col += 1

            # -- Kolon genislikleri --
            for col_idx, width in col_max_widths.items():
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width, 10), 50)

            # -- Filtre ekle --
            if table.rowCount() > 0:
                last_letter = get_column_letter(max(col_count, 1))
                ws.auto_filter.ref = f"A{header_row}:{last_letter}{header_row + table.rowCount()}"

            # -- Satir yuksekligi --
            ws.row_dimensions[header_row].height = 28

            # -- Sayfa yapisi --
            ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(
                fitToPage=True)
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.page_setup.orientation = 'landscape'

            wb.save(filepath)

            QMessageBox.information(self, "Basarili",
                                    f"Excel dosyasi olusturuldu:\n{filepath}")

            # Export logla
            try:
                from core.log_manager import LogManager
                LogManager.log_export(title, f"{title} Excel olarak disari aktarildi ({table.rowCount()} satir)")
            except Exception:
                pass

            # Dosyayi ac
            import os
            try:
                os.startfile(filepath)
            except Exception:
                pass

        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Excel olusturma hatasi:\n{str(e)}")

    def create_export_button(self, table=None, title: str = "Veri",
                               table_attr: str = "table") -> QPushButton:
        """
        Excel/CSV/PDF export menu butonu olustur.
        Toolbar'a eklenmeye hazir.

        Args:
            table: QTableWidget referansi (None ise table_attr kullanilir)
            title: Export dosya adi
            table_attr: self uzerindeki tablo attribute adi (lazy binding)

        Kullanim:
            btn = self.create_export_button(title="Personel Listesi")
            toolbar_layout.addWidget(btn)
        """
        t = self.theme
        btn = QPushButton("Disa Aktar")
        btn.setCursor(Qt.PointingHandCursor)
        btn.setStyleSheet(f"""
            QPushButton {{
                background: {t.get('bg_card', '#1E293B')};
                color: {t.get('text', '#E2E8F0')};
                border: 1px solid {t.get('border', '#334155')};
                padding: 8px 16px;
                border-radius: 6px;
                font-size: 13px;
                font-weight: 500;
            }}
            QPushButton:hover {{
                background: {t.get('bg_hover', '#334155')};
                border-color: {t.get('primary', '#DC2626')};
            }}
            QPushButton::menu-indicator {{
                subcontrol-position: right center;
                subcontrol-origin: padding;
                right: 8px;
            }}
        """)

        menu = QMenu(self)
        menu.setStyleSheet(f"""
            QMenu {{
                background: {t.get('bg_card', '#1E293B')};
                color: {t.get('text', '#E2E8F0')};
                border: 1px solid {t.get('border', '#334155')};
                border-radius: 6px;
                padding: 4px;
            }}
            QMenu::item {{
                padding: 8px 24px;
                border-radius: 4px;
            }}
            QMenu::item:selected {{
                background: {t.get('bg_hover', '#334155')};
            }}
        """)

        def _get_table():
            if table is not None:
                return table
            return getattr(self, table_attr, None)

        menu.addAction("Excel (.xlsx)", lambda: self.export_table_to_excel(_get_table(), title))
        menu.addAction("CSV (.csv)", lambda: self._export_table_to_csv(_get_table(), title))

        btn.setMenu(menu)
        return btn

    def _export_table_to_csv(self, table: QTableWidget, title: str = "Veri"):
        """QTableWidget'i CSV'ye aktar"""
        if table is None:
            QMessageBox.warning(self, "Hata", "Aktarilacak tablo bulunamadi.")
            return

        import csv

        safe_title = "".join(c for c in title if c.isalnum() or c in (' ', '_', '-')).strip()
        filepath, _ = QFileDialog.getSaveFileName(
            self, "CSV Kaydet",
            f"{safe_title}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            "CSV Dosyasi (*.csv)"
        )
        if not filepath:
            return

        try:
            col_count = table.columnCount() - 1  # Son sutun (islem) atla

            with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f, delimiter=';')

                headers = []
                for col in range(col_count):
                    header = table.horizontalHeaderItem(col)
                    headers.append(header.text() if header else "")
                writer.writerow(headers)

                for row in range(table.rowCount()):
                    row_data = []
                    for col in range(col_count):
                        item = table.item(row, col)
                        row_data.append(item.text() if item else "")
                    writer.writerow(row_data)

            QMessageBox.information(self, "Basarili", f"CSV kaydedildi:\n{filepath}")

            import os
            try:
                os.startfile(filepath)
            except Exception:
                pass

        except Exception as e:
            QMessageBox.critical(self, "Hata", f"CSV olusturma hatasi:\n{str(e)}")
