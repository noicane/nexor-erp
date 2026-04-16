# -*- coding: utf-8 -*-
"""
REDLINE NEXOR ERP - Styled Table
Tüm tablolar için merkezi stil ve davranış yönetimi

Özellikler:
- Tema entegrasyonu
- Otomatik sütun genişlikleri
- İşlem butonları
- Sıralama
- Arama/Filtreleme
- Sayfalama
- Export (Excel, PDF, CSV)
- Context menu
- Loading/Empty state
- Sütun gizleme/gösterme
"""
import os
from datetime import datetime
from typing import Optional, Callable, Any

from core.nexor_brand import brand
from PySide6.QtWidgets import (
    QTableWidget, QTableWidgetItem, QHeaderView, QWidget,
    QHBoxLayout, QVBoxLayout, QPushButton, QAbstractItemView,
    QMenu, QLabel, QLineEdit, QComboBox, QFrame, QSpinBox,
    QFileDialog, QMessageBox, QApplication, QProgressBar,
    QStyledItemDelegate, QStyle
)
from PySide6.QtCore import Qt, Signal, QTimer, QSize
from PySide6.QtGui import QColor, QIcon, QAction, QCursor


# ============================================
# GLOBAL TABLO AYARLARI
# ============================================
DEFAULT_CONFIG = {
    # Minimum sütun genişlikleri
    "min_column_width": 60,
    "min_button_column_width": 90,
    
    # Buton ayarları
    "action_button_size": (28, 24),
    "action_column_min_width": 80,
    
    # Satır ayarları
    "row_height": 36,
    "alternating_row_colors": True,
    
    # Sayfalama
    "default_page_size": 50,
    "page_size_options": [25, 50, 100, 250, 500],
    
    # Özel sütun genişlikleri
    "column_widths": {
        # Dar sütunlar
        "ID": 50, "id": 50, "No": 50, "#": 40, "%": 50,
        "Ort.": 55, "Oran": 55, "Adet": 60, "Saat": 60,
        # Orta sütunlar
        "Tarih": 85, "Durum": 80, "Seviye": 70, "Kod": 80, "Tip": 80,
        # Geniş sütunlar
        "Ad": 150, "Açıklama": 200, "Personel": 150, "Departman": 120,
        # İşlem sütunları
        "İşlem": 90, "İşlemler": 110, "Aksiyon": 90,
    },
}


class TableConfig:
    """
    Tablo konfigürasyonu - Merkezi ayar yönetimi
    themes.py veya harici config ile entegre edilebilir
    """
    _instance = None
    _config = DEFAULT_CONFIG.copy()
    
    @classmethod
    def get_instance(cls):
        if cls._instance is None:
            cls._instance = cls()
        return cls._instance
    
    @classmethod
    def get(cls, key: str, default=None):
        return cls._config.get(key, default)
    
    @classmethod
    def set(cls, key: str, value):
        cls._config[key] = value
    
    @classmethod
    def update(cls, config: dict):
        """Konfigürasyonu güncelle"""
        cls._config.update(config)
    
    @classmethod
    def get_column_width(cls, header: str) -> int:
        """Sütun başlığına göre genişlik döndür"""
        widths = cls._config.get("column_widths", {})
        return widths.get(header, cls._config.get("min_column_width", 60))
    
    @classmethod
    def set_column_width(cls, header: str, width: int):
        """Sütun genişliği tanımla"""
        if "column_widths" not in cls._config:
            cls._config["column_widths"] = {}
        cls._config["column_widths"][header] = width
    
    @classmethod
    def load_from_theme(cls, theme: dict):
        """Tema'dan tablo ayarlarını yükle"""
        if "table_config" in theme:
            cls.update(theme["table_config"])


class LoadingOverlay(QWidget):
    """Tablo üzerinde loading göstergesi"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents, False)
        self.setStyleSheet("background: rgba(0, 0, 0, 0.5); border-radius: 12px;")
        
        layout = QVBoxLayout(self)
        layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        self.label = QLabel("Yükleniyor...")
        self.label.setStyleSheet(f"color: {brand.TEXT}; font-size: {brand.FS_BODY_LG}px; font-weight: bold;")
        self.label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.label)

        self.progress = QProgressBar()
        self.progress.setFixedWidth(200)
        self.progress.setTextVisible(False)
        self.progress.setRange(0, 0)  # Indeterminate
        self.progress.setStyleSheet(f"""
            QProgressBar {{
                border: none;
                border-radius: {brand.R_SM}px;
                background: rgba(255,255,255,0.2);
                height: 6px;
            }}
            QProgressBar::chunk {{
                background: {brand.INFO};
                border-radius: {brand.R_SM}px;
            }}
        """)
        layout.addWidget(self.progress)
        
        self.hide()
    
    def set_message(self, message: str):
        self.label.setText(message)
    
    def set_progress(self, value: int, maximum: int = 100):
        """Belirli progress göster"""
        self.progress.setRange(0, maximum)
        self.progress.setValue(value)


class EmptyStateWidget(QWidget):
    """Veri yokken gösterilen widget"""
    
    def __init__(self, theme: dict, parent=None):
        super().__init__(parent)
        self.theme = theme
        self.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents)
        
        layout = QVBoxLayout(self)
        layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        self.icon_label = QLabel("📭")
        self.icon_label.setStyleSheet("font-size: 48px;")
        self.icon_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.icon_label)
        
        self.message_label = QLabel("Görüntülenecek veri bulunamadı")
        self.message_label.setStyleSheet(f"""
            color: {brand.TEXT_DIM};
            font-size: {brand.FS_BODY_LG}px;
            margin-top: {brand.SP_2}px;
        """)
        self.message_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.message_label)

        self.sub_label = QLabel("")
        self.sub_label.setStyleSheet(f"""
            color: {brand.TEXT_DIM};
            font-size: {brand.FS_BODY_SM}px;
        """)
        self.sub_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.sub_label)
        
        self.hide()
    
    def set_message(self, message: str, sub_message: str = "", icon: str = "📭"):
        self.icon_label.setText(icon)
        self.message_label.setText(message)
        self.sub_label.setText(sub_message)
        self.sub_label.setVisible(bool(sub_message))


class PaginationWidget(QFrame):
    """Sayfalama kontrolleri"""
    
    page_changed = Signal(int)  # Sayfa değişti sinyali
    page_size_changed = Signal(int)  # Sayfa boyutu değişti sinyali
    
    def __init__(self, theme: dict, parent=None):
        super().__init__(parent)
        self.theme = theme
        self.current_page = 1
        self.total_pages = 1
        self.total_records = 0
        self.page_size = TableConfig.get("default_page_size", 50)
        
        self._setup_ui()
    
    def _setup_ui(self):
        self.setStyleSheet(f"""
            QFrame {{
                background: {brand.BG_CARD};
                border: 1px solid {brand.BORDER};
                border-radius: {brand.R_MD}px;
                padding: {brand.SP_1}px;
            }}
        """)

        layout = QHBoxLayout(self)
        layout.setContentsMargins(brand.SP_3, brand.SP_2, brand.SP_3, brand.SP_2)
        layout.setSpacing(brand.SP_2)

        # Sol: Kayit bilgisi
        self.info_label = QLabel()
        self.info_label.setStyleSheet(f"color: {brand.TEXT_MUTED}; font-size: {brand.FS_BODY_SM}px;")
        layout.addWidget(self.info_label)

        layout.addStretch()

        # Orta: Sayfa boyutu
        layout.addWidget(QLabel("Sayfa:"))
        self.page_size_combo = QComboBox()
        self.page_size_combo.setFixedWidth(70)
        for size in TableConfig.get("page_size_options", [25, 50, 100]):
            self.page_size_combo.addItem(str(size), size)
        self.page_size_combo.setCurrentText(str(self.page_size))
        self.page_size_combo.currentIndexChanged.connect(self._on_page_size_changed)
        layout.addWidget(self.page_size_combo)

        layout.addSpacing(brand.SP_5)

        # Sag: Navigasyon butonlari
        btn_style = f"""
            QPushButton {{
                background: {brand.BG_INPUT};
                border: 1px solid {brand.BORDER};
                border-radius: {brand.R_SM}px;
                padding: {brand.SP_1}px {brand.SP_2}px;
                color: {brand.TEXT};
                min-width: 28px;
            }}
            QPushButton:hover {{
                background: {brand.BG_HOVER};
                border-color: {brand.PRIMARY};
            }}
            QPushButton:disabled {{
                color: {brand.TEXT_DISABLED};
            }}
        """

        self.btn_first = QPushButton("\u23ee")
        self.btn_first.setToolTip("Ilk Sayfa")
        self.btn_first.setStyleSheet(btn_style)
        self.btn_first.clicked.connect(lambda: self.go_to_page(1))
        layout.addWidget(self.btn_first)

        self.btn_prev = QPushButton("\u25c0")
        self.btn_prev.setToolTip("Onceki Sayfa")
        self.btn_prev.setStyleSheet(btn_style)
        self.btn_prev.clicked.connect(lambda: self.go_to_page(self.current_page - 1))
        layout.addWidget(self.btn_prev)

        self.page_label = QLabel("1 / 1")
        self.page_label.setStyleSheet(f"color: {brand.TEXT}; font-weight: bold; margin: 0 {brand.SP_2}px;")
        layout.addWidget(self.page_label)

        self.btn_next = QPushButton("\u25b6")
        self.btn_next.setToolTip("Sonraki Sayfa")
        self.btn_next.setStyleSheet(btn_style)
        self.btn_next.clicked.connect(lambda: self.go_to_page(self.current_page + 1))
        layout.addWidget(self.btn_next)

        self.btn_last = QPushButton("\u23ed")
        self.btn_last.setToolTip("Son Sayfa")
        self.btn_last.setStyleSheet(btn_style)
        self.btn_last.clicked.connect(lambda: self.go_to_page(self.total_pages))
        layout.addWidget(self.btn_last)
        
        self._update_ui()
    
    def set_total_records(self, total: int):
        """Toplam kayıt sayısını ayarla"""
        self.total_records = total
        self.total_pages = max(1, (total + self.page_size - 1) // self.page_size)
        if self.current_page > self.total_pages:
            self.current_page = self.total_pages
        self._update_ui()
    
    def go_to_page(self, page: int):
        """Belirtilen sayfaya git"""
        if 1 <= page <= self.total_pages and page != self.current_page:
            self.current_page = page
            self._update_ui()
            self.page_changed.emit(page)
    
    def _on_page_size_changed(self):
        self.page_size = self.page_size_combo.currentData()
        self.total_pages = max(1, (self.total_records + self.page_size - 1) // self.page_size)
        self.current_page = 1
        self._update_ui()
        self.page_size_changed.emit(self.page_size)
    
    def _update_ui(self):
        # Info label
        start = (self.current_page - 1) * self.page_size + 1
        end = min(self.current_page * self.page_size, self.total_records)
        if self.total_records > 0:
            self.info_label.setText(f"{start}-{end} / {self.total_records} kayıt")
        else:
            self.info_label.setText("0 kayıt")
        
        # Page label
        self.page_label.setText(f"{self.current_page} / {self.total_pages}")
        
        # Button states
        self.btn_first.setEnabled(self.current_page > 1)
        self.btn_prev.setEnabled(self.current_page > 1)
        self.btn_next.setEnabled(self.current_page < self.total_pages)
        self.btn_last.setEnabled(self.current_page < self.total_pages)
    
    def get_offset(self) -> int:
        """SQL OFFSET değerini döndür"""
        return (self.current_page - 1) * self.page_size
    
    def get_limit(self) -> int:
        """SQL LIMIT değerini döndür"""
        return self.page_size


class SearchFilterWidget(QFrame):
    """Arama ve filtreleme kontrolleri"""
    
    search_changed = Signal(str)  # Arama metni değişti
    filter_changed = Signal(str, object)  # Filtre değişti (alan, değer)
    
    def __init__(self, theme: dict, parent=None):
        super().__init__(parent)
        self.theme = theme
        self.filters = {}  # {alan: QComboBox}
        
        self._setup_ui()
    
    def _setup_ui(self):
        self.setStyleSheet("QFrame { background: transparent; }")

        self.layout = QHBoxLayout(self)
        self.layout.setContentsMargins(0, 0, 0, 0)
        self.layout.setSpacing(brand.SP_3)

        # Arama kutusu
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("\U0001f50d Ara...")
        self.search_input.setFixedWidth(250)
        self.search_input.setStyleSheet(f"""
            QLineEdit {{
                background: {brand.BG_INPUT};
                border: 1px solid {brand.BORDER};
                border-radius: {brand.R_MD}px;
                padding: {brand.SP_2}px {brand.SP_3}px;
                color: {brand.TEXT};
            }}
            QLineEdit:focus {{
                border-color: {brand.PRIMARY};
            }}
        """)
        
        # Debounce için timer
        self._search_timer = QTimer()
        self._search_timer.setSingleShot(True)
        self._search_timer.timeout.connect(self._emit_search)
        self.search_input.textChanged.connect(self._on_search_changed)
        
        self.layout.addWidget(self.search_input)
        self.layout.addStretch()
    
    def _on_search_changed(self):
        """Arama değişti - debounce ile bekle"""
        self._search_timer.stop()
        self._search_timer.start(300)  # 300ms bekle
    
    def _emit_search(self):
        self.search_changed.emit(self.search_input.text())
    
    def add_filter(self, name: str, label: str, options: list):
        """
        Filtre combobox ekle
        
        Args:
            name: Filtre adı (alan adı)
            label: Görünen etiket
            options: [(value, text), ...] veya [text, ...] listesi
        """
        combo = QComboBox()
        combo.setFixedWidth(150)
        combo.setStyleSheet(f"""
            QComboBox {{
                background: {brand.BG_INPUT};
                border: 1px solid {brand.BORDER};
                border-radius: {brand.R_MD}px;
                padding: {brand.SP_2}px {brand.SP_3}px;
                color: {brand.TEXT};
            }}
        """)
        
        # Tümü seçeneği
        combo.addItem(f"Tüm {label}", None)
        
        # Seçenekleri ekle
        for opt in options:
            if isinstance(opt, tuple):
                combo.addItem(opt[1], opt[0])
            else:
                combo.addItem(str(opt), opt)
        
        combo.currentIndexChanged.connect(lambda: self.filter_changed.emit(name, combo.currentData()))
        
        # Layout'a ekle (stretch'ten önce)
        self.layout.insertWidget(self.layout.count() - 1, QLabel(f"{label}:"))
        self.layout.insertWidget(self.layout.count() - 1, combo)
        
        self.filters[name] = combo
    
    def get_search_text(self) -> str:
        return self.search_input.text()
    
    def get_filter_value(self, name: str):
        if name in self.filters:
            return self.filters[name].currentData()
        return None
    
    def clear_all(self):
        """Tüm filtreleri temizle"""
        self.search_input.clear()
        for combo in self.filters.values():
            combo.setCurrentIndex(0)


class ColumnVisibilityMenu(QMenu):
    """Sütun görünürlük menüsü"""
    
    column_toggled = Signal(int, bool)  # (sütun_index, görünür)
    
    def __init__(self, table: 'StyledTable', parent=None):
        super().__init__("Sütunlar", parent)
        self.table = table
        self._build_menu()
    
    def _build_menu(self):
        self.clear()
        for col in range(self.table.columnCount()):
            header = self.table.horizontalHeaderItem(col)
            text = header.text() if header else f"Sütun {col}"
            
            action = QAction(text, self)
            action.setCheckable(True)
            action.setChecked(not self.table.isColumnHidden(col))
            action.triggered.connect(lambda checked, c=col: self._toggle_column(c, checked))
            self.addAction(action)
    
    def _toggle_column(self, col: int, visible: bool):
        self.table.setColumnHidden(col, not visible)
        self.column_toggled.emit(col, visible)


class StyledTable(QTableWidget):
    """
    Merkezi stil ve davranış yönetimine sahip tablo widget'ı.
    
    Özellikler:
    - Tema entegrasyonu
    - Otomatik sütun genişlikleri
    - İşlem butonları helper'ı
    - Sıralama desteği
    - Context menu
    - Loading/Empty state
    - Export fonksiyonları
    - Sütun gizleme/gösterme
    
    Kullanım:
        table = StyledTable(self.theme)
        table.setup_columns(["ID", "Ad", "Durum", "Tarih", "İşlem"])
        # ... veri doldur ...
        table.auto_resize()
    """
    
    # Sinyaller
    row_double_clicked = Signal(int, dict)  # (row_index, row_data)
    action_triggered = Signal(str, int, dict)  # (action_name, row_index, row_data)
    data_changed = Signal()  # Veri değişti
    
    def __init__(self, theme: dict, parent=None, enable_sorting: bool = True,
                 enable_pagination: bool = False, enable_search: bool = False):
        super().__init__(parent)
        
        self.theme = theme
        self._enable_sorting = enable_sorting
        self._enable_pagination = enable_pagination
        self._enable_search = enable_search
        
        # Veri
        self._all_data = []  # Tüm veri (filtreleme için)
        self._hidden_columns = set()
        self._column_formatters = {}  # {col_index: formatter_func}
        
        # Alt widget'lar
        self._loading_overlay = None
        self._empty_state = None
        self._pagination = None
        self._search_filter = None
        self._column_menu = None
        
        # Kurulum
        self._apply_style()
        self._setup_defaults()
        self._setup_signals()
    
    def _apply_style(self):
        """Brand bazli stil uygula"""
        self.setStyleSheet(f"""
            QTableWidget {{
                background: {brand.BG_CARD};
                border: 1px solid {brand.BORDER};
                border-radius: {brand.R_LG}px;
                gridline-color: {brand.BORDER};
                color: {brand.TEXT};
                font-size: {brand.FS_BODY_SM}px;
                selection-background-color: {brand.PRIMARY_SOFT};
                outline: none;
            }}
            QTableWidget::item {{
                padding: {brand.SP_2}px;
                border-bottom: 1px solid {brand.BORDER};
            }}
            QTableWidget::item:selected {{
                background: {brand.PRIMARY_SOFT};
                color: {brand.TEXT};
            }}
            QTableWidget::item:hover {{
                background: {brand.BG_HOVER};
            }}
            QHeaderView::section {{
                background: {brand.BG_SURFACE};
                color: {brand.TEXT};
                padding: {brand.SP_3}px {brand.SP_2}px;
                border: none;
                border-bottom: 2px solid {brand.PRIMARY};
                font-weight: bold;
                font-size: {brand.FS_BODY_SM}px;
            }}
            QHeaderView::section:hover {{
                background: {brand.BG_HOVER};
            }}
            QTableCornerButton::section {{
                background: {brand.BG_SURFACE};
                border: none;
            }}
            QScrollBar:vertical {{
                background: {brand.BG_CARD};
                width: 10px;
                border-radius: 5px;
                margin: 0;
            }}
            QScrollBar::handle:vertical {{
                background: {brand.BORDER_HARD};
                border-radius: 5px;
                min-height: 30px;
            }}
            QScrollBar::handle:vertical:hover {{
                background: {brand.PRIMARY};
            }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                height: 0px;
            }}
            QScrollBar:horizontal {{
                background: {brand.BG_CARD};
                height: 10px;
                border-radius: 5px;
            }}
            QScrollBar::handle:horizontal {{
                background: {brand.BORDER_HARD};
                border-radius: 5px;
                min-width: 30px;
            }}
            QScrollBar::handle:horizontal:hover {{
                background: {brand.PRIMARY};
            }}
            QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {{
                width: 0px;
            }}
        """)
    
    def _setup_defaults(self):
        """Varsayılan tablo ayarları"""
        # Seçim davranışı
        self.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        
        # Alternatif satır renkleri
        self.setAlternatingRowColors(TableConfig.get("alternating_row_colors", True))
        
        # Grid
        self.setShowGrid(True)
        
        # Satır numaralarını gizle
        self.verticalHeader().setVisible(False)
        
        # Header ayarları
        header = self.horizontalHeader()
        header.setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
        header.setHighlightSections(False)
        header.setStretchLastSection(False)
        
        # Sıralama
        if self._enable_sorting:
            self.setSortingEnabled(True)
            header.setSortIndicatorShown(True)
        
        # Varsayılan satır yüksekliği
        self.verticalHeader().setDefaultSectionSize(TableConfig.get("row_height", 36))
        
        # Context menu
        self.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.customContextMenuRequested.connect(self._show_context_menu)
        
        # Edit disabled
        self.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
    
    def _setup_signals(self):
        """Sinyalleri bağla"""
        self.cellDoubleClicked.connect(self._on_double_click)
    
    def _on_double_click(self, row: int, col: int):
        """Çift tıklama"""
        data = self.get_row_data(row)
        self.row_double_clicked.emit(row, data)
    
    # ==========================================
    # PUBLIC API - Setup
    # ==========================================
    
    def setup_columns(self, headers: list, widths: list = None):
        """
        Sütunları ayarla
        
        Args:
            headers: Sütun başlıkları
            widths: Sütun genişlikleri (opsiyonel)
        """
        self.setColumnCount(len(headers))
        self.setHorizontalHeaderLabels(headers)
        
        if widths:
            for i, w in enumerate(widths):
                if w:
                    self.setColumnWidth(i, w)
    
    def set_stretch_last_section(self, stretch: bool):
        """Son sütunu genişlet"""
        self.horizontalHeader().setStretchLastSection(stretch)
    
    def set_column_formatter(self, column: int, formatter: Callable):
        """
        Sütun için formatter tanımla
        
        Args:
            column: Sütun indeksi
            formatter: func(value) -> str
        """
        self._column_formatters[column] = formatter
    
    def update_theme(self, theme: dict):
        """Tema değiştiğinde güncelle"""
        self.theme = theme
        self._apply_style()
        
        # Alt widget'ları güncelle
        if self._loading_overlay:
            self._loading_overlay.deleteLater()
            self._loading_overlay = None
        if self._empty_state:
            self._empty_state.deleteLater()
            self._empty_state = None
    
    # ==========================================
    # PUBLIC API - Loading / Empty State
    # ==========================================
    
    def show_loading(self, message: str = "Yükleniyor..."):
        """Loading göster"""
        if not self._loading_overlay:
            self._loading_overlay = LoadingOverlay(self)
        
        self._loading_overlay.set_message(message)
        self._loading_overlay.setGeometry(self.rect())
        self._loading_overlay.show()
        self._loading_overlay.raise_()
        QApplication.processEvents()
    
    def hide_loading(self):
        """Loading gizle"""
        if self._loading_overlay:
            self._loading_overlay.hide()
    
    def show_empty_state(self, message: str = "Görüntülenecek veri bulunamadı",
                         sub_message: str = "", icon: str = "📭"):
        """Empty state göster"""
        if not self._empty_state:
            self._empty_state = EmptyStateWidget(self.theme, self)
        
        self._empty_state.set_message(message, sub_message, icon)
        self._empty_state.setGeometry(self.rect())
        self._empty_state.show()
        self._empty_state.raise_()
    
    def hide_empty_state(self):
        """Empty state gizle"""
        if self._empty_state:
            self._empty_state.hide()
    
    def resizeEvent(self, event):
        """Resize olduğunda overlay'ları yeniden boyutlandır"""
        super().resizeEvent(event)
        if self._loading_overlay:
            self._loading_overlay.setGeometry(self.rect())
        if self._empty_state:
            self._empty_state.setGeometry(self.rect())
    
    # ==========================================
    # PUBLIC API - Pagination
    # ==========================================
    
    def get_pagination_widget(self) -> PaginationWidget:
        """Sayfalama widget'ını döndür (yoksa oluştur)"""
        if not self._pagination:
            self._pagination = PaginationWidget(self.theme)
        return self._pagination
    
    def set_total_records(self, total: int):
        """Toplam kayıt sayısını ayarla (pagination için)"""
        if self._pagination:
            self._pagination.set_total_records(total)
    
    # ==========================================
    # PUBLIC API - Search/Filter
    # ==========================================
    
    def get_search_widget(self) -> SearchFilterWidget:
        """Arama widget'ını döndür (yoksa oluştur)"""
        if not self._search_filter:
            self._search_filter = SearchFilterWidget(self.theme)
        return self._search_filter
    
    # ==========================================
    # PUBLIC API - Column Visibility
    # ==========================================
    
    def get_column_menu(self) -> ColumnVisibilityMenu:
        """Sütun görünürlük menüsünü döndür"""
        if not self._column_menu:
            self._column_menu = ColumnVisibilityMenu(self)
        return self._column_menu
    
    def hide_column(self, column: int):
        """Sütunu gizle"""
        self.setColumnHidden(column, True)
        self._hidden_columns.add(column)
    
    def show_column(self, column: int):
        """Sütunu göster"""
        self.setColumnHidden(column, False)
        self._hidden_columns.discard(column)
    
    def hide_columns_by_name(self, *names):
        """İsme göre sütunları gizle"""
        for col in range(self.columnCount()):
            header = self.horizontalHeaderItem(col)
            if header and header.text() in names:
                self.hide_column(col)
    
    # ==========================================
    # PUBLIC API - Data
    # ==========================================
    
    def auto_resize(self):
        """Sütun genişliklerini otomatik ayarla"""
        # Sorting'i geçici kapat (performans için)
        sorting_enabled = self.isSortingEnabled()
        self.setSortingEnabled(False)
        
        # Önce içeriğe göre boyutlandır
        self.resizeColumnsToContents()
        
        # Sonra minimum/özel genişlikleri uygula
        for col in range(self.columnCount()):
            if col in self._hidden_columns:
                continue
                
            current_width = self.columnWidth(col)
            header_item = self.horizontalHeaderItem(col)
            header_text = header_item.text() if header_item else ""
            
            # Özel genişlik tanımlı mı?
            if header_text in TableConfig.get("column_widths", {}):
                min_width = TableConfig.get_column_width(header_text)
            # Widget içeren sütun mu?
            elif self._column_has_widget(col):
                min_width = TableConfig.get("min_button_column_width", 90)
            else:
                min_width = TableConfig.get("min_column_width", 60)
            
            # Minimum genişliği uygula
            if current_width < min_width:
                self.setColumnWidth(col, min_width)
        
        # Sorting'i geri aç
        self.setSortingEnabled(sorting_enabled)
        
        # Empty state kontrolü
        if self.rowCount() == 0:
            self.show_empty_state()
        else:
            self.hide_empty_state()
    
    def _column_has_widget(self, col: int) -> bool:
        """Sütunda widget var mı"""
        for row in range(min(3, self.rowCount())):
            if self.cellWidget(row, col):
                return True
        return False
    
    def fill_row(self, row: int, data: list, alignments: list = None):
        """
        Satırı verilerle doldur
        
        Args:
            row: Satır indeksi
            data: Hücre verileri
            alignments: Hizalama ("L", "C", "R")
        """
        for col, value in enumerate(data):
            # Formatter varsa uygula
            if col in self._column_formatters:
                display_value = self._column_formatters[col](value)
            else:
                display_value = str(value) if value is not None else "-"
            
            item = QTableWidgetItem(display_value)
            item.setData(Qt.ItemDataRole.UserRole, value)  # Orijinal değeri sakla
            
            # Hizalama
            if alignments and col < len(alignments):
                align = alignments[col]
                if align == "C":
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                elif align == "R":
                    item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                else:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
            
            self.setItem(row, col, item)
    
    def fill_table(self, data: list, headers: list = None):
        """
        Tabloyu tamamen doldur
        
        Args:
            data: [[col1, col2, ...], ...] şeklinde veri
            headers: Sütun başlıkları (opsiyonel)
        """
        self.show_loading()
        self.setSortingEnabled(False)
        
        try:
            if headers:
                self.setup_columns(headers)
            
            self.setRowCount(len(data))
            
            for row_idx, row_data in enumerate(data):
                self.fill_row(row_idx, row_data)
            
            self._all_data = data
            self.auto_resize()
            
        finally:
            self.setSortingEnabled(self._enable_sorting)
            self.hide_loading()
    
    def clear_data(self):
        """Tablo verilerini temizle"""
        self.setRowCount(0)
        self._all_data = []
        self.show_empty_state()
    
    def get_row_data(self, row: int) -> dict:
        """Satır verisini dictionary olarak al"""
        data = {}
        for col in range(self.columnCount()):
            header = self.horizontalHeaderItem(col)
            key = header.text() if header else f"col_{col}"
            
            item = self.item(row, col)
            if item:
                # UserRole'da orijinal değer varsa onu al
                original = item.data(Qt.ItemDataRole.UserRole)
                data[key] = original if original is not None else item.text()
            else:
                # Widget olabilir
                widget = self.cellWidget(row, col)
                data[key] = widget if widget else None
        
        return data
    
    def get_selected_row_data(self) -> Optional[dict]:
        """Seçili satırın verisini al"""
        row = self.currentRow()
        if row >= 0:
            return self.get_row_data(row)
        return None
    
    def get_selected_id(self, column: int = 0):
        """Seçili satırın ID'sini al"""
        row = self.currentRow()
        if row >= 0:
            item = self.item(row, column)
            if item:
                user_data = item.data(Qt.ItemDataRole.UserRole)
                return user_data if user_data is not None else item.text()
        return None
    
    def get_all_data(self) -> list:
        """Tüm tablo verisini liste olarak al"""
        data = []
        for row in range(self.rowCount()):
            row_data = []
            for col in range(self.columnCount()):
                item = self.item(row, col)
                if item:
                    row_data.append(item.text())
                else:
                    row_data.append("")
            data.append(row_data)
        return data
    
    # ==========================================
    # PUBLIC API - Action Buttons
    # ==========================================
    
    def create_action_widget(self, buttons: list, row_data: dict = None) -> QWidget:
        """
        İşlem sütunu için buton widget'ı oluştur
        
        Args:
            buttons: [(icon, tooltip, callback, color), ...]
                     color: "primary", "success", "danger", "warning", "info" veya hex
            row_data: Callback'lere geçirilecek satır verisi
        
        Örnek:
            widget = table.create_action_widget([
                ("✏️", "Düzenle", lambda: self.edit(row_id), "primary"),
                ("🗑️", "Sil", lambda: self.delete(row_id), "danger"),
            ])
        """
        widget = QWidget()
        layout = QHBoxLayout(widget)
        layout.setContentsMargins(4, 2, 4, 2)
        layout.setSpacing(4)
        layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        btn_width, btn_height = TableConfig.get("action_button_size", (28, 24))

        # Koyu tonlu arka planlar (kurumsal)
        color_map = {
            "edit": "#1E3A5F",
            "delete": "#5F1E1E",
            "view": "#3D1E5F",
            "print": "#2E2E5F",
            "photo": "#1E3A4F",
            "primary": "#4F1E1E",
            "success": "#1E4F2E",
            "danger": "#5F1E1E",
            "warning": "#4F3A1E",
            "info": "#1E3A5F",
            "secondary": "#2A3040",
        }

        # Hover'da parlak renkler
        hover_map = {
            "edit": "#2563EB",
            "delete": "#EF4444",
            "view": "#8B5CF6",
            "print": "#6366F1",
            "photo": "#0EA5E9",
            "primary": brand.PRIMARY,
            "success": "#10B981",
            "danger": "#EF4444",
            "warning": "#F59E0B",
            "info": "#3B82F6",
            "secondary": "#4B5563",
        }

        for btn_config in buttons:
            if len(btn_config) == 4:
                icon, tooltip, callback, color = btn_config
            else:
                icon, tooltip, callback = btn_config
                color = "edit"

            btn = QPushButton(icon)
            btn.setToolTip(tooltip)
            btn.setFixedSize(btn_width, btn_height)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)

            bg_color = color_map.get(color, color)
            hover_color = hover_map.get(color, bg_color)

            btn.setStyleSheet(f"""
                QPushButton {{
                    background: {bg_color};
                    color: white;
                    border: none;
                    border-radius: 4px;
                    font-size: 13px;
                    padding: 0;
                }}
                QPushButton:hover {{
                    background: {hover_color};
                }}
            """)
            
            if callback:
                btn.clicked.connect(callback)
            
            layout.addWidget(btn)
        
        return widget
    
    def create_status_item(self, text: str, status: str = "default") -> QTableWidgetItem:
        """
        Renkli durum hücresi oluştur
        
        Args:
            text: Görüntülenecek metin
            status: "success", "warning", "danger", "info", "default"
        """
        item = QTableWidgetItem(text)
        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        
        status_colors = {
            "success": (brand.SUCCESS, brand.TEXT_INVERSE),
            "warning": (brand.WARNING, brand.TEXT_INVERSE),
            "danger": (brand.ERROR, brand.TEXT_INVERSE),
            "info": (brand.INFO, brand.TEXT_INVERSE),
            "default": (brand.BG_HOVER, brand.TEXT),
        }
        
        bg, fg = status_colors.get(status, status_colors["default"])
        item.setBackground(QColor(bg))
        item.setForeground(QColor(fg))
        
        return item
    
    # ==========================================
    # PUBLIC API - Context Menu
    # ==========================================
    
    def _show_context_menu(self, pos):
        """Sağ tık menüsü"""
        menu = QMenu(self)
        menu.setStyleSheet(f"""
            QMenu {{
                background: {brand.BG_CARD};
                border: 1px solid {brand.BORDER};
                border-radius: {brand.R_MD}px;
                padding: {brand.SP_1}px;
            }}
            QMenu::item {{
                padding: {brand.SP_2}px {brand.SP_6}px;
                color: {brand.TEXT};
            }}
            QMenu::item:selected {{
                background: {brand.PRIMARY};
                border-radius: {brand.R_SM}px;
            }}
            QMenu::separator {{
                height: 1px;
                background: {brand.BORDER};
                margin: {brand.SP_1}px {brand.SP_2}px;
            }}
        """)
        
        # Kopyala
        copy_action = menu.addAction("📋 Seçili Satırı Kopyala")
        copy_action.triggered.connect(self._copy_selected_row)
        
        copy_all_action = menu.addAction("📄 Tümünü Kopyala")
        copy_all_action.triggered.connect(self._copy_all)
        
        menu.addSeparator()
        
        # Export
        export_menu = menu.addMenu("📤 Dışa Aktar")
        
        excel_action = export_menu.addAction("Excel (.xlsx)")
        excel_action.triggered.connect(lambda: self.export_to_excel())
        
        csv_action = export_menu.addAction("CSV (.csv)")
        csv_action.triggered.connect(lambda: self.export_to_csv())
        
        pdf_action = export_menu.addAction("PDF (.pdf)")
        pdf_action.triggered.connect(lambda: self.export_to_pdf())
        
        menu.addSeparator()
        
        # Sütun görünürlük
        col_menu = self.get_column_menu()
        col_menu._build_menu()  # Refresh
        menu.addMenu(col_menu)
        
        menu.exec(self.mapToGlobal(pos))
    
    def _copy_selected_row(self):
        """Seçili satırı kopyala"""
        row = self.currentRow()
        if row >= 0:
            texts = []
            for col in range(self.columnCount()):
                item = self.item(row, col)
                texts.append(item.text() if item else "")
            
            clipboard = QApplication.clipboard()
            clipboard.setText("\t".join(texts))
    
    def _copy_all(self):
        """Tüm tabloyu kopyala"""
        lines = []
        
        # Headers
        headers = []
        for col in range(self.columnCount()):
            header = self.horizontalHeaderItem(col)
            headers.append(header.text() if header else "")
        lines.append("\t".join(headers))
        
        # Data
        for row in range(self.rowCount()):
            texts = []
            for col in range(self.columnCount()):
                item = self.item(row, col)
                texts.append(item.text() if item else "")
            lines.append("\t".join(texts))
        
        clipboard = QApplication.clipboard()
        clipboard.setText("\n".join(lines))
    
    # ==========================================
    # PUBLIC API - Export
    # ==========================================
    
    def export_to_excel(self, filepath: str = None):
        """Excel'e aktar"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            QMessageBox.warning(self, "Hata", "openpyxl kütüphanesi yüklü değil.\npip install openpyxl")
            return
        
        if not filepath:
            filepath, _ = QFileDialog.getSaveFileName(
                self, "Excel Kaydet",
                f"Tablo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel (*.xlsx)"
            )
        
        if not filepath:
            return
        
        self.show_loading("Excel oluşturuluyor...")
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Veri"
            
            # Stiller
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Headers
            for col in range(self.columnCount()):
                if col in self._hidden_columns:
                    continue
                header = self.horizontalHeaderItem(col)
                cell = ws.cell(row=1, column=col + 1)
                cell.value = header.text() if header else ""
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border
            
            # Data
            for row in range(self.rowCount()):
                for col in range(self.columnCount()):
                    if col in self._hidden_columns:
                        continue
                    item = self.item(row, col)
                    cell = ws.cell(row=row + 2, column=col + 1)
                    cell.value = item.text() if item else ""
                    cell.border = thin_border
            
            # Sütun genişlikleri
            for col in range(self.columnCount()):
                if col in self._hidden_columns:
                    continue
                ws.column_dimensions[openpyxl.utils.get_column_letter(col + 1)].width = 15
            
            wb.save(filepath)
            QMessageBox.information(self, "Başarılı", f"Excel kaydedildi:\n{filepath}")
            
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Excel oluşturma hatası:\n{str(e)}")
        finally:
            self.hide_loading()
    
    def export_to_csv(self, filepath: str = None):
        """CSV'ye aktar"""
        import csv
        
        if not filepath:
            filepath, _ = QFileDialog.getSaveFileName(
                self, "CSV Kaydet",
                f"Tablo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                "CSV (*.csv)"
            )
        
        if not filepath:
            return
        
        try:
            with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f, delimiter=';')
                
                # Headers
                headers = []
                for col in range(self.columnCount()):
                    if col in self._hidden_columns:
                        continue
                    header = self.horizontalHeaderItem(col)
                    headers.append(header.text() if header else "")
                writer.writerow(headers)
                
                # Data
                for row in range(self.rowCount()):
                    row_data = []
                    for col in range(self.columnCount()):
                        if col in self._hidden_columns:
                            continue
                        item = self.item(row, col)
                        row_data.append(item.text() if item else "")
                    writer.writerow(row_data)
            
            QMessageBox.information(self, "Başarılı", f"CSV kaydedildi:\n{filepath}")
            
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"CSV oluşturma hatası:\n{str(e)}")
    
    def export_to_pdf(self, filepath: str = None, title: str = "Tablo Raporu"):
        """PDF'e aktar"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
        except ImportError:
            QMessageBox.warning(self, "Hata", "reportlab kütüphanesi yüklü değil.\npip install reportlab")
            return
        
        if not filepath:
            filepath, _ = QFileDialog.getSaveFileName(
                self, "PDF Kaydet",
                f"Tablo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                "PDF (*.pdf)"
            )
        
        if not filepath:
            return
        
        self.show_loading("PDF oluşturuluyor...")
        
        try:
            # Font
            try:
                pdfmetrics.registerFont(TTFont('DejaVu', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
                font_name = 'DejaVu'
            except Exception:
                font_name = 'Helvetica'
            
            doc = SimpleDocTemplate(filepath, pagesize=landscape(A4))
            elements = []
            styles = getSampleStyleSheet()
            
            # Başlık
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontName=font_name,
                fontSize=16,
                spaceAfter=10
            )
            elements.append(Paragraph(title, title_style))
            elements.append(Paragraph(
                f"Tarih: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
                ParagraphStyle('Date', fontName=font_name, fontSize=10)
            ))
            elements.append(Spacer(1, 15))
            
            # Headers
            headers = []
            visible_cols = []
            for col in range(self.columnCount()):
                if col not in self._hidden_columns:
                    header = self.horizontalHeaderItem(col)
                    headers.append(header.text() if header else "")
                    visible_cols.append(col)
            
            data = [headers]
            
            # Data
            for row in range(self.rowCount()):
                row_data = []
                for col in visible_cols:
                    item = self.item(row, col)
                    row_data.append(item.text() if item else "")
                data.append(row_data)
            
            # Tablo
            col_count = len(visible_cols)
            col_width = 700 / col_count if col_count > 0 else 100
            table = Table(data, colWidths=[col_width] * col_count)
            
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
            ]))
            
            elements.append(table)
            doc.build(elements)
            
            QMessageBox.information(self, "Başarılı", f"PDF kaydedildi:\n{filepath}")
            
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"PDF oluşturma hatası:\n{str(e)}")
        finally:
            self.hide_loading()
    
    def get_export_data(self) -> tuple:
        """
        Export için hazır veri döndür (headers, data)
        Harici export fonksiyonları için kullanılabilir
        """
        headers = []
        for col in range(self.columnCount()):
            if col not in self._hidden_columns:
                header = self.horizontalHeaderItem(col)
                headers.append(header.text() if header else "")
        
        data = []
        for row in range(self.rowCount()):
            row_data = []
            for col in range(self.columnCount()):
                if col not in self._hidden_columns:
                    item = self.item(row, col)
                    row_data.append(item.text() if item else "")
            data.append(row_data)
        
        return headers, data
